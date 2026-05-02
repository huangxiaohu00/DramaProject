from __future__ import annotations

import importlib
import argparse
import sys
import time
import random
import inspect
from concurrent.futures import ThreadPoolExecutor
from collections import Counter
from copy import deepcopy
from dataclasses import replace
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Mapping, MutableMapping, Sequence, Tuple

import openpyxl


if __package__ in (None, ""):
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))


config = importlib.import_module("drama_pipeline.2_config")
models = importlib.import_module("drama_pipeline.3_models")
rules = importlib.import_module("drama_pipeline.8_recommendation_rules")
excel_io = importlib.import_module("drama_pipeline.5_excel_io")
client_factory = importlib.import_module("drama_pipeline.platform_clients.factory")
runtime = importlib.import_module("drama_pipeline.10_runtime")
runtime.configure_utf8_runtime()


MaterialMap = Mapping[Tuple[str, str, str], int | models.MaterialResult]
TodayInputBundle = Dict[str, List[object]]
RunStats = Dict[str, object]


TODAY_STAGE_FILES = {
    "mobo_new": (1, "Mobo新剧", "Mobo新剧"),
    "beidou_new": (2, "北斗新剧", "北斗新剧"),
    "mobo_recommend": (3, "Mobo推荐剧", "Mobo推荐剧"),
    "beidou_hot": (4, "北斗飞书爆款", "北斗飞书爆款"),
    "duole_recommend": (5, "多乐推荐剧", "多乐推荐剧"),
    "published": (6, "飞书已发布", "飞书已发布"),
}

SOURCE_DISPLAY_NAMES = {
    "mobo_new": "Mobo新剧",
    "beidou_new": "北斗新剧",
    "mobo_recommend": "Mobo推荐",
    "beidou_hot": "飞书北斗热榜",
    "beidou_income": "北斗收入榜",
    "duole_recommend": "多乐推荐榜",
}


def _log_info(logger, message: str) -> None:
    if logger is not None:
        logger.info(message)


def _log_step_result(logger, name: str, count: int | None = None, elapsed: float | None = None) -> None:
    if logger is None:
        return
    parts = [name]
    if count is not None:
        parts.append(f"count={count}")
    if elapsed is not None:
        parts.append(f"elapsed={elapsed:.1f}s")
    logger.info("；".join(parts))


def collect_today_inputs(
    mobo_client,
    beidou_client,
    feishu_client,
    duole_client,
    selected_languages: Sequence[str] | None = None,
    logger=None,
) -> TodayInputBundle:
    languages = normalize_selected_languages(selected_languages)
    bundle: TodayInputBundle = {
        "mobo_new": [],
        "beidou_new": [],
        "mobo_recommend": [],
        "beidou_hot": [],
        "duole_recommend": [],
        "published": [],
    }

    return _collect_today_inputs_concurrent(
        languages=languages,
        mobo_client=mobo_client,
        beidou_client=beidou_client,
        feishu_client=feishu_client,
        duole_client=duole_client,
        bundle=bundle,
        logger=logger,
    )


def _collect_today_inputs_concurrent(
    languages: Sequence[str],
    mobo_client,
    beidou_client,
    feishu_client,
    duole_client,
    bundle: TodayInputBundle,
    logger=None,
) -> TodayInputBundle:
    _log_info(logger, f"开始并发采集：languages={len(languages)}")
    started = time.perf_counter()
    workers = min(max(int(getattr(config, "TODAY_COLLECT_WORKERS", 4) or 4), 1), 4)
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = [
            executor.submit(_collect_mobo_inputs, mobo_client, languages, logger),
            executor.submit(_collect_beidou_inputs, beidou_client, languages, logger),
            executor.submit(_collect_feishu_inputs, feishu_client, languages, logger),
            executor.submit(_collect_duole_inputs, duole_client, languages, logger),
        ]
        for future in futures:
            _merge_today_bundle(bundle, future.result())
    _log_step_result(logger, f"并发采集完成(workers={workers})", elapsed=time.perf_counter() - started)
    return bundle


def _merge_today_bundle(bundle: TodayInputBundle, partial: TodayInputBundle) -> None:
    for key, rows in partial.items():
        bundle.setdefault(key, []).extend(rows)


def _empty_today_bundle() -> TodayInputBundle:
    return {
        "mobo_new": [],
        "beidou_new": [],
        "mobo_recommend": [],
        "beidou_hot": [],
        "duole_recommend": [],
        "published": [],
    }


def _collect_mobo_inputs(mobo_client, languages: Sequence[str], logger=None) -> TodayInputBundle:
    partial = _empty_today_bundle()
    started = time.perf_counter()
    _log_info(logger, f"[采集] Mobo 开始：languages={','.join(languages)}")
    for language in languages:
        language_started = time.perf_counter()
        new_rows = _fetch_mobo_for_target_theaters(mobo_client, "fetch_new_dramas", language)
        partial["mobo_new"].extend(new_rows)
        _log_info(
            logger,
            f"[采集] Mobo新剧 {language} 完成：count={len(new_rows)} elapsed={time.perf_counter() - language_started:.1f}s",
        )
        language_started = time.perf_counter()
        recommend_rows = _fetch_mobo_for_target_theaters(mobo_client, "fetch_recommend_dramas", language)
        partial["mobo_recommend"].extend(recommend_rows)
        _log_info(
            logger,
            f"[采集] Mobo推荐 {language} 完成：count={len(recommend_rows)} elapsed={time.perf_counter() - language_started:.1f}s",
        )
    _log_info(
        logger,
        f"[采集] Mobo 完成：new_total={len(partial['mobo_new'])} recommend_total={len(partial['mobo_recommend'])} elapsed={time.perf_counter() - started:.1f}s",
    )
    return partial


def _fetch_mobo_for_target_theaters(mobo_client, method_name: str, language: str) -> List[models.DramaRecord]:
    method = getattr(mobo_client, method_name)
    theaters = _mobo_supported_target_theaters(language)
    if not theaters or not _callable_accepts_parameter(method, "theater"):
        return method(language)
    rows: List[models.DramaRecord] = []
    for theater in theaters:
        rows.extend(method(language, theater=theater))
    return rows


def _mobo_supported_target_theaters(language: str) -> List[str]:
    platform_ids = dict(getattr(config, "MOBO_PLATFORM_IDS", {}) or {})
    return [theater for theater in _target_theaters_for_language(language) if theater in platform_ids]


def _collect_beidou_inputs(beidou_client, languages: Sequence[str], logger=None) -> TodayInputBundle:
    partial = _empty_today_bundle()
    english = config.LANGUAGE_CONFIG[2]
    started = time.perf_counter()
    _log_info(logger, f"[采集] 北斗 开始：languages={','.join(languages)}")
    for language in languages:
        language_started = time.perf_counter()
        beidou_new = _fetch_beidou_new_for_target_theaters(beidou_client, language)
        partial["beidou_new"].extend(row for row in beidou_new if row.theater != "MoboReels")
        _log_info(logger, f"[采集] 北斗新剧 {language} 完成：raw={len(beidou_new)} kept={len([row for row in beidou_new if row.theater != 'MoboReels'])} elapsed={time.perf_counter() - language_started:.1f}s")
    for language in languages:
        if language == english:
            continue
        language_started = time.perf_counter()
        partial["beidou_hot"].extend(beidou_client.fetch_income_dramas(language))
        _log_info(logger, f"[采集] 北斗收入榜 {language} 完成：count={len([row for row in partial['beidou_hot'] if row.language == language])} elapsed={time.perf_counter() - language_started:.1f}s")
    _log_info(
        logger,
        f"[采集] 北斗 完成：new_total={len(partial['beidou_new'])} hot_total={len(partial['beidou_hot'])} elapsed={time.perf_counter() - started:.1f}s",
    )
    return partial


def _fetch_beidou_new_for_target_theaters(beidou_client, language: str) -> List[models.DramaRecord]:
    theaters = _target_theaters_for_language(language, include_moboreels=False)
    if not theaters or not _callable_accepts_parameter(getattr(beidou_client, "fetch_new_dramas", None), "theater"):
        return beidou_client.fetch_new_dramas(language)
    rows: List[models.DramaRecord] = []
    for theater in theaters:
        rows.extend(beidou_client.fetch_new_dramas(language, theater=theater))
    return rows


def _target_theaters_for_language(language: str, include_moboreels: bool = True) -> List[str]:
    quotas = rules.normalized_language_quotas(language)
    theaters = [theater for theater, limit in quotas.items() if int(limit or 0) > 0]
    if not include_moboreels:
        theaters = [theater for theater in theaters if theater != "MoboReels"]
    return theaters


def _callable_accepts_parameter(func, parameter_name: str) -> bool:
    if not callable(func):
        return False
    try:
        signature = inspect.signature(func)
    except (TypeError, ValueError):
        return False
    return parameter_name in signature.parameters


def _collect_feishu_inputs(feishu_client, languages: Sequence[str], logger=None) -> TodayInputBundle:
    partial = _empty_today_bundle()
    english = config.LANGUAGE_CONFIG[2]
    started = time.perf_counter()
    _log_info(logger, f"[采集] 飞书 开始：languages={','.join(languages)}")
    for language in languages:
        language_started = time.perf_counter()
        rows = feishu_client.fetch_published(language)
        partial["published"].extend(rows)
        _log_info(logger, f"[采集] 飞书已发布 {language} 完成：count={len(rows)} elapsed={time.perf_counter() - language_started:.1f}s")
    if english in languages:
        started_hot = time.perf_counter()
        hot_rows = feishu_client.fetch_beidou_hot_dramas()
        partial["beidou_hot"].extend(hot_rows)
        _log_info(logger, f"[采集] 飞书北斗热榜 完成：count={len(hot_rows)} elapsed={time.perf_counter() - started_hot:.1f}s")
    _log_info(
        logger,
        f"[采集] 飞书 完成：published_total={len(partial['published'])} hot_total={len(partial['beidou_hot'])} elapsed={time.perf_counter() - started:.1f}s",
    )
    return partial


def _collect_duole_inputs(duole_client, languages: Sequence[str], logger=None) -> TodayInputBundle:
    partial = _empty_today_bundle()
    started = time.perf_counter()
    _log_info(logger, f"[采集] 多乐 开始：languages={','.join(languages)}")
    all_rows = duole_client.fetch_recommend_dramas()
    duole_rows = [row for row in all_rows if row.language in languages]
    limited_rows = limit_duole_rows(duole_rows)
    partial["duole_recommend"].extend(limited_rows)
    _log_duole_sheet_summary(
        logger,
        getattr(duole_client, "fetch_stats", []),
        language_rows=duole_rows,
        limited_rows=limited_rows,
    )
    _log_info(
        logger,
        f"[采集] 多乐 完成：raw={len(all_rows)} language_kept={len(duole_rows)} limited={len(partial['duole_recommend'])} elapsed={time.perf_counter() - started:.1f}s",
    )
    return partial


def build_default_clients(logger=None) -> Dict[str, object]:
    return client_factory.build_today_clients(
        logger=logger,
        material_enabled=_is_realtime_material_check_enabled(),
    )


def load_local_adult_filter(app_dir: Path | str | None = None, logger=None) -> List[models.TitleBlockRecord]:
    path = runtime.ensure_adult_filter_workbook(runtime.default_adult_filter_path(app_dir))
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if runtime.ADULT_FILTER_SHEET not in workbook.sheetnames:
            return []
        worksheet = workbook[runtime.ADULT_FILTER_SHEET]
        headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
        index = {str(header).strip(): column - 1 for column, header in enumerate(headers, 1) if header is not None}
        title_keys = ("剧集名称", "剧名")
        output: List[models.TitleBlockRecord] = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            language = runtime.normalize_language_name(_sheet_value(row, index, "语言"))
            title = _sheet_value(row, index, *title_keys).strip()
            theater = config.normalize_theater_name(_sheet_value(row, index, "剧场"))
            if not language or not title:
                continue
            output.append(
                models.TitleBlockRecord(
                    title=title,
                    language=language,
                    theater=theater,
                    source="adult_filter",
                    raw={"path": str(path), "theater": theater},
                )
            )
        if logger is not None:
            logger.info(f"已加载本地成人过滤 {len(output)} 条: {path}")
        return output
    finally:
        workbook.close()


def normalize_selected_languages(selected_languages: Sequence[str] | None = None) -> List[str]:
    if not selected_languages:
        return list(config.LANGUAGE_ORDER)
    allowed = set(config.LANGUAGE_ORDER)
    output: List[str] = []
    for language in selected_languages:
        normalized = runtime.normalize_language_name(language)
        if normalized in allowed and normalized not in output:
            output.append(normalized)
    return output


def _sheet_value(row: Sequence[object], index: Mapping[str, int], *keys: str) -> str:
    for key in keys:
        column = index.get(key)
        if column is None or column >= len(row):
            continue
        text = "" if row[column] is None else str(row[column]).strip()
        if text:
            return text
    return ""


def limit_duole_rows(rows: Sequence[models.DramaRecord]) -> List[models.DramaRecord]:
    kept_per_sheet: Dict[str, List[models.DramaRecord]] = {}
    passthrough: List[models.DramaRecord] = []
    for row in rows:
        sheet_name = str((row.raw or {}).get("sheet_name") or "")
        limit = config.DUOLE_SHEET_LIMITS.get(sheet_name)
        if limit is None:
            passthrough.append(row)
            continue
        bucket = kept_per_sheet.setdefault(sheet_name, [])
        if len(bucket) < limit:
            bucket.append(row)

    limited: List[models.DramaRecord] = []
    for sheet_name in config.DUOLE_TARGET_SHEETS:
        sheet_rows = kept_per_sheet.get(sheet_name, [])
        for index, row in enumerate(sheet_rows, 1):
            limited.append(replace(row, rank=index))
    limited.extend(passthrough)
    return limited


def _log_duole_sheet_summary(
    logger,
    fetch_stats_rows: Sequence[Dict[str, object]],
    language_rows: Sequence[models.DramaRecord],
    limited_rows: Sequence[models.DramaRecord],
) -> None:
    raw_by_sheet = _duole_fetch_stage_counts(fetch_stats_rows, "原始返回")
    parsed_by_sheet = _duole_fetch_stage_counts(fetch_stats_rows, "解析成功")
    language_by_sheet = _count_duole_rows_by_sheet(language_rows)
    limited_by_sheet = _count_duole_rows_by_sheet(limited_rows)
    ordered_sheets = list(dict.fromkeys([*config.DUOLE_TARGET_SHEETS, *raw_by_sheet.keys(), *parsed_by_sheet.keys()]))
    for sheet_name in ordered_sheets:
        _log_info(
            logger,
            "[采集] 多乐 Sheet 汇总："
            f"sheet={sheet_name} "
            f"raw={raw_by_sheet.get(sheet_name, 0)} "
            f"parsed={parsed_by_sheet.get(sheet_name, 0)} "
            f"language_kept={language_by_sheet.get(sheet_name, 0)} "
            f"limited={limited_by_sheet.get(sheet_name, 0)}",
        )


def _duole_fetch_stage_counts(fetch_stats_rows: Sequence[Dict[str, object]], stage: str) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for row in fetch_stats_rows:
        if not isinstance(row, dict):
            continue
        if str(row.get("来源") or "") != "duole_recommend":
            continue
        if str(row.get("阶段") or "") != stage:
            continue
        if str(row.get("维度") or "") != "Sheet":
            continue
        sheet_name = str(row.get("值") or "").strip()
        if not sheet_name:
            continue
        counts[sheet_name] = counts.get(sheet_name, 0) + int(row.get("条数") or 0)
    return counts


def _count_duole_rows_by_sheet(rows: Sequence[models.DramaRecord]) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for row in rows:
        sheet_name = str((row.raw or {}).get("sheet_name") or "").strip()
        if not sheet_name:
            continue
        counts[sheet_name] = counts.get(sheet_name, 0) + 1
    return counts


def export_today_stage_files(bundle: TodayInputBundle, output_root: Path, date_text: str) -> Dict[str, Path]:
    paths: Dict[str, Path] = {}
    for key, (stage, filename, sheet_name) in TODAY_STAGE_FILES.items():
        path = excel_io.build_output_path(output_root, date_text, stage, filename)
        rows = [serialize_record(row) for row in bundle.get(key, [])]
        excel_io.write_workbook(path, {sheet_name: rows})
        paths[key] = path
    return paths


def build_recommendation_from_bundle(
    bundle: TodayInputBundle,
    material_results: MaterialMap | None = None,
    material_client=None,
    material_start_date: str | None = None,
    material_end_date: str | None = None,
    fetch_stats_rows: Sequence[Dict[str, object]] | None = None,
    reference_date: str | None = None,
    material_failed_registry_path: Path | str | None = None,
    logger=None,
) -> models.TodayRecommendationRun:
    candidates: List[models.DramaRecord] = []
    for key in ("mobo_new", "beidou_new", "mobo_recommend", "beidou_hot", "duole_recommend"):
        candidates.extend(bundle.get(key, []))
    published = [row for row in bundle.get("published", []) if isinstance(row, models.PublishedRecord)]
    adult_blocks = [row for row in bundle.get("adult_filter", []) if isinstance(row, models.TitleBlockRecord)]
    stats = build_bundle_stats(bundle, fetch_stats_rows=fetch_stats_rows)
    metadata_index = build_metadata_index(bundle)
    return build_offline_recommendation(
        candidates,
        published,
        material_results or {},
        adult_blocks=adult_blocks,
        material_client=material_client,
        material_start_date=material_start_date,
        material_end_date=material_end_date,
        initial_stats=stats,
        reference_date=reference_date,
        material_failed_registry_path=material_failed_registry_path,
        metadata_index=metadata_index,
        logger=logger,
    )


def run_today_recommendation(
    selected_languages: Sequence[str] | None = None,
    date_text: str | None = None,
    output_root: Path | None = None,
    material_results: MaterialMap | None = None,
    material_start_date: str | None = None,
    material_end_date: str | None = None,
    clients: Dict[str, object] | None = None,
    logger=None,
) -> Dict[str, Path]:
    active_clients = clients or build_default_clients(logger=logger)
    reset_client_fetch_stats(active_clients)
    if logger is not None:
        logger.step("开始采集今日推荐源数据")
    bundle = collect_today_inputs(
        mobo_client=active_clients["mobo_client"],
        beidou_client=active_clients["beidou_client"],
        feishu_client=active_clients["feishu_client"],
        duole_client=active_clients["duole_client"],
        selected_languages=selected_languages,
        logger=logger,
    )
    bundle["adult_filter"] = load_local_adult_filter(logger=logger)
    fetch_stats_rows = collect_client_fetch_stats(active_clients)
    date_value = date_text or datetime.now().strftime("%Y-%m-%d")
    if logger is not None:
        logger.set_run_date(date_value)
    start_date, end_date = resolve_material_date_range(material_start_date, material_end_date, date_value)
    root = Path(output_root or config.TODAY_OUTPUT_ROOT)
    material_client = active_clients.get("material_client")
    cache_setter = getattr(material_client, "set_daily_cache_path", None)
    if callable(cache_setter):
        cache_setter(root / date_value / "material_cache.json")
    if logger is not None:
        logger.info(f"输出目录: {root}")
        logger.step("写出阶段原始数据")
    paths = export_today_stage_files(bundle, root, date_value)
    if logger is not None:
        logger.step("执行去重、过滤、打分与素材校验")
    result = build_recommendation_from_bundle(
        bundle,
        material_results=material_results,
        material_client=material_client,
        material_start_date=start_date,
        material_end_date=end_date,
        fetch_stats_rows=fetch_stats_rows,
        reference_date=date_value,
        material_failed_registry_path=config.PIPELINE_DIR / "material_failed_records.xlsx",
        logger=logger,
    )
    if logger is not None:
        _log_recommendation_health(logger, result)
        logger.step("写出推荐结果")
    paths.update(export_offline_result(result, root, date_value))
    return paths


def parse_args(argv: Sequence[str] | None = None):
    parser = argparse.ArgumentParser(description="今日剧集推荐")
    parser.add_argument("--languages", default="", help="逗号分隔语言，例如: 德语,法语")
    parser.add_argument("--date", default="", help="输出日期目录，默认当天，格式 YYYY-MM-DD")
    parser.add_argument("--material-start", default="", help="素材开始日期 YYYY-MM-DD，默认输出日期向前两年")
    parser.add_argument("--material-end", default="", help="素材结束日期 YYYY-MM-DD，默认输出日期")
    parsed = parser.parse_args(argv)
    parsed.languages = [item.strip() for item in parsed.languages.split(",") if item.strip()]
    return parsed


def serialize_record(row: object) -> Dict[str, object]:
    if hasattr(row, "to_dict"):
        return row.to_dict()
    if isinstance(row, dict):
        return dict(row)
    return {"value": row}


def build_offline_recommendation(
    candidates: Sequence[models.DramaRecord],
    published: Sequence[models.PublishedRecord],
    material_results: MaterialMap,
    adult_blocks: Sequence[models.TitleBlockRecord] = (),
    material_client=None,
    material_start_date: str | None = None,
    material_end_date: str | None = None,
    initial_stats: RunStats | None = None,
    reference_date: str | None = None,
    metadata_index: Mapping[Tuple[str, str], Dict[str, object]] | None = None,
    material_failed_registry_path: Path | str | None = None,
    logger=None,
) -> models.TodayRecommendationRun:
    stats = prepare_run_stats(initial_stats, candidates, published)
    _log_info(logger, f"[处理] 候选准备开始: candidates={len(candidates)} published={len(published)} adult_blocks={len(adult_blocks)}")
    started = time.perf_counter()
    deduped = rules.dedupe_candidates(candidates)
    _log_step_result(logger, "[处理] 去重完成", count=len(deduped), elapsed=time.perf_counter() - started)
    started = time.perf_counter()
    deduped, metadata_backfilled = _enrich_candidates_with_metadata(
        deduped,
        metadata_index=metadata_index,
        logger=logger,
    )
    _log_step_result(logger, "[处理] 元数据回填完成", count=metadata_backfilled, elapsed=time.perf_counter() - started)
    started = time.perf_counter()
    unpublished, published_blocks = rules.filter_published(deduped, published)
    _log_step_result(logger, "[处理] 飞书已发布过滤完成", count=len(unpublished), elapsed=time.perf_counter() - started)
    started = time.perf_counter()
    not_adult, adult_filter_blocks = rules.filter_title_blocks(unpublished, adult_blocks, "本地成人过滤")
    _log_step_result(logger, "[处理] 本地成人过滤完成", count=len(not_adult), elapsed=time.perf_counter() - started)
    started = time.perf_counter()
    language_cleaned, language_blocks = _filter_language_mismatches(not_adult)
    _log_step_result(logger, "[处理] 语言校验完成", count=len(language_cleaned), elapsed=time.perf_counter() - started)
    started = time.perf_counter()
    clean_candidates, content_blocks = rules.filter_blocked_content(language_cleaned)
    _log_step_result(logger, "[处理] 内容过滤完成", count=len(clean_candidates), elapsed=time.perf_counter() - started)
    started = time.perf_counter()
    scored = rules.score_candidates(clean_candidates, reference_date=reference_date)
    _log_step_result(logger, "[处理] 打分完成", count=len(scored), elapsed=time.perf_counter() - started)
    started = time.perf_counter()
    shortlist, shortlist_blocks, shortlist_seen = _build_material_shortlist(scored)
    _log_step_result(logger, "[处理] shortlist 预选完成", count=len(shortlist), elapsed=time.perf_counter() - started)

    _set_stat(stats, "processed", "deduped", len(deduped))
    _set_stat(stats, "processed", "metadata_backfilled", metadata_backfilled)
    _set_stat(stats, "processed", "after_published_filter", len(unpublished))
    _set_stat(stats, "processed", "after_adult_filter", len(not_adult))
    _set_stat(stats, "processed", "after_language_filter", len(language_cleaned))
    _set_stat(stats, "processed", "after_content_filter", len(clean_candidates))
    _set_stat(stats, "processed", "scored", len(scored))
    shortlist_total = len(shortlist)
    _set_stat(stats, "processed", "shortlist", shortlist_total)
    _record_stage_counts(stats, "输入候选", candidates)
    _record_stage_counts(stats, "候选去重后", deduped)
    _record_stage_counts(stats, "元数据回填后", deduped)
    _record_stage_counts(stats, "飞书已发布过滤后", unpublished)
    _record_stage_counts(stats, "本地成人过滤后", not_adult)
    _record_stage_counts(stats, "语言校验后", language_cleaned)
    _record_stage_counts(stats, "内容过滤后", clean_candidates)
    _record_stage_counts(stats, "进入打分", scored)
    _record_stage_counts(stats, "shortlist", shortlist)
    _record_layer_counts(stats, scored)

    recommendations: List[models.RecommendationResult] = []
    filter_records: List[models.FilterRecord] = (
        list(published_blocks)
        + list(adult_filter_blocks)
        + list(language_blocks)
        + list(content_blocks)
        + list(shortlist_blocks)
    )
    theater_counts: Dict[str, Dict[str, int]] = {}
    language_counts: Dict[str, int] = {}
    selected_title_keys: set[Tuple[str, str]] = set()
    material_cache: Dict[Tuple[str, str, str], int | models.MaterialResult] = dict(material_results)
    material_checked = 0
    material_skipped = 0
    material_prefetched = 0
    material_qualified: List[models.DramaRecord] = []
    material_failed_rows: List[Dict[str, object]] = []
    material_cooldown_keys = load_material_failed_cooldown_keys(
        material_failed_registry_path,
        reference_date=reference_date,
        cooldown_days=int(getattr(config, "MATERIAL_FAILED_COOLDOWN_DAYS", 7) or 0),
    )
    if material_cooldown_keys:
        _log_info(logger, f"[处理] 已加载不达标素材冷却: count={len(material_cooldown_keys)}")
    ai_anime_counts: Dict[str, int] = {}
    english_origin_counts: Dict[str, int] = {"local": 0, "translated": 0, "unknown": 0}
    remaining_english_origin = _build_english_origin_remaining(
        [candidate for candidate in scored if _is_material_pool_supported(candidate)]
    )
    deferred_english_balance: List[models.DramaRecord] = []
    deferred_pool_counts: Dict[str, Dict[str, int]] = {}
    progress_started = time.perf_counter()
    last_progress = progress_started
    progress_interval = 200 if config.MATERIAL_CHECK_ENABLED else 500
    _log_info(
        logger,
        f"[处理] 开始最终选取: scored={len(scored)} shortlist={len(shortlist)} material_check={'on' if config.MATERIAL_CHECK_ENABLED else 'off'}",
    )
    exhausted_pools: set[Tuple[str, str]] = set()

    def try_select_candidate(
        candidate: models.DramaRecord,
        allow_english_balance_overflow: bool = False,
    ) -> bool:
        nonlocal material_checked, material_skipped, last_progress
        language_target = rules.language_target(candidate.language)
        if language_counts.get(candidate.language, 0) >= language_target:
            return False
        theater_block_reason = rules.theater_quota_block_reason(candidate, theater_counts)
        if theater_block_reason:
            blocked = _filter(candidate, theater_block_reason)
            filter_records.append(blocked)
            _increment_stat(stats, "filtered", theater_block_reason)
            return False
        if candidate.match_key in selected_title_keys:
            blocked = _filter(candidate, "同语言剧名重复")
            filter_records.append(blocked)
            _increment_stat(stats, "filtered", "同语言剧名重复")
            return False
        ai_anime_block_reason = _ai_anime_limit_block_reason(candidate, ai_anime_counts)
        if ai_anime_block_reason:
            blocked = _filter(candidate, ai_anime_block_reason)
            filter_records.append(blocked)
            _increment_stat(stats, "filtered", ai_anime_block_reason)
            return False
        english_balance_reason = ""
        if not allow_english_balance_overflow:
            english_balance_reason = _english_origin_balance_block_reason(
                candidate,
                language_counts,
                english_origin_counts,
                remaining_english_origin,
            )
        if english_balance_reason:
            deferred_english_balance.append(candidate)
            rules.increment_theater_quota(candidate, deferred_pool_counts)
            return False

        if _candidate_material_key(candidate) in material_cooldown_keys:
            reason = "素材冷却期未达标"
            filter_records.append(_filter(candidate, reason))
            _increment_stat(stats, "filtered", reason)
            return False

        qualified_count, total_count, checked, skipped = _resolve_material_counts(
            material_cache,
            candidate,
            material_client,
            material_start_date,
            material_end_date,
        )
        if checked:
            material_checked += 1
            if logger is not None and (
                material_checked % 20 == 0 or time.perf_counter() - last_progress >= 30
            ):
                last_progress = time.perf_counter()
                logger.info(
                    f"[处理] 素材校验进度: checked={material_checked} selected={len(recommendations)} current_language={candidate.language} elapsed={last_progress - progress_started:.1f}s"
                )
        if skipped:
            material_skipped += 1
        if qualified_count >= config.VIDEO_THRESHOLD:
            material_qualified.append(candidate)
            rank = language_counts.get(candidate.language, 0) + 1
            recommendations.append(
                models.RecommendationResult(
                    title=candidate.title,
                    language=candidate.language,
                    theater=candidate.theater,
                    rank=rank,
                    score=candidate.score,
                    layer=candidate.layer,
                    qualified_count=qualified_count,
                    total_count=total_count,
                    content_label=build_content_label(candidate),
                    source=candidate.source,
                    promotion_time=build_promotion_time(candidate),
                    pre_rank_summary=build_pre_rank_summary(candidate),
                    recommend_reason=build_recommend_reason(
                        candidate,
                        qualified_count,
                        total_count,
                        material_skipped=skipped,
                    ),
                    score_breakdown=build_score_breakdown(candidate),
                    rule_hits=build_rule_hits(candidate, qualified_count, total_count, skipped),
                    filter_trace=build_filter_trace(candidate),
                    source_dates=build_source_dates(candidate),
                    quality_flags=build_quality_flags(candidate, qualified_count, total_count, skipped),
                )
            )
            language_counts[candidate.language] = rank
            selected_title_keys.add(candidate.match_key)
            rules.increment_theater_quota(candidate, theater_counts)
            if rules.is_ai_or_anime(candidate):
                ai_anime_counts[candidate.language] = ai_anime_counts.get(candidate.language, 0) + 1
            if candidate.language == config.LANGUAGE_CONFIG[2]:
                origin = rules.content_origin(candidate)
                english_origin_counts[origin] = english_origin_counts.get(origin, 0) + 1
            return True

        reason = f"素材不达标:{qualified_count}"
        filter_records.append(_filter(candidate, reason))
        _increment_stat(stats, "filtered", reason)
        if checked and not skipped:
            material_failed_rows.append(_material_failed_row(candidate, reference_date, qualified_count, total_count))
        return False

    processed_shortlist = 0
    pool_wave_counts: Dict[Tuple[str, str], int] = {}
    pending_shortlist_by_pool = _group_shortlist_by_pool(shortlist)
    prefetch_chunk_size = _material_prefetch_chunk_size()
    for pool_key in _material_pool_order(scored):
        while True:
            pool_shortage = int(_recommendation_pool_shortages(theater_counts).get(pool_key, 0) or 0)
            if pool_shortage <= 0:
                break
            pending_shortlist = list(pending_shortlist_by_pool.get(pool_key) or [])
            if not pending_shortlist:
                extension_shortage = int(
                    _recommendation_pool_shortages(theater_counts, deferred_pool_counts).get(pool_key, 0) or 0
                )
                if extension_shortage <= 0:
                    _increment_stat(stats, "processed", f"暂存已覆盖缺口:{pool_key[0]}/{pool_key[1]}")
                    break
                next_wave = pool_wave_counts.get(pool_key, 0) + 1
                if next_wave > _material_max_expansion_waves_per_pool():
                    exhausted_pools.add(pool_key)
                    _increment_stat(stats, "processed", f"扩容上限:{pool_key[0]}/{pool_key[1]}")
                    _log_info(logger, f"[处理] shortlist 扩容达到上限，停止 {pool_key[0]}/{pool_key[1]} waves={next_wave - 1}")
                    break
                pending_shortlist = _extend_material_shortlist(scored, shortlist_seen, {pool_key: extension_shortage})
                if not pending_shortlist:
                    exhausted_pools.add(pool_key)
                    _increment_stat(stats, "processed", f"候选耗尽:{pool_key[0]}/{pool_key[1]}")
                    _log_info(logger, f"[处理] shortlist 候选耗尽，停止 {pool_key[0]}/{pool_key[1]} shortage={pool_shortage}")
                    break
                pending_shortlist_by_pool[pool_key] = list(pending_shortlist)
                shortlist_total += len(pending_shortlist)
                _set_stat(stats, "processed", "shortlist", shortlist_total)
                _record_stage_counts(
                    stats,
                    f"shortlist-扩容-{pool_key[0]}-{pool_key[1]}-{next_wave}",
                    pending_shortlist,
                )
                _log_info(
                    logger,
                    f"[处理] shortlist 不足，扩容 {pool_key[0]}/{pool_key[1]} 第{next_wave}轮: added={len(pending_shortlist)} shortage={pool_shortage}",
                )
            if not pending_shortlist:
                break
            wave = pool_wave_counts.get(pool_key, 0) + 1
            current_batch = list(pending_shortlist[:prefetch_chunk_size])
            pending_shortlist_by_pool[pool_key] = list(pending_shortlist[prefetch_chunk_size:])
            if not current_batch:
                break

            started = time.perf_counter()
            prefetched = _prefetch_shortlist_materials(
                current_batch,
                material_cache,
                material_client,
                material_start_date,
                material_end_date,
                material_cooldown_keys=material_cooldown_keys,
                logger=logger,
            )
            material_prefetched += prefetched
            if prefetched:
                _log_step_result(
                    logger,
                    f"[处理] {pool_key[0]}/{pool_key[1]} 第{wave}轮 shortlist 素材批量校验完成",
                    count=prefetched,
                    elapsed=time.perf_counter() - started,
                )

            for candidate in current_batch:
                _consume_english_origin(candidate, remaining_english_origin)
                try_select_candidate(candidate)
                processed_shortlist += 1
                now = time.perf_counter()
                if logger is not None and (
                    processed_shortlist % progress_interval == 0 or now - last_progress >= 30
                ):
                    last_progress = now
                    logger.info(
                        f"[处理] 选取进度: processed={processed_shortlist} selected={len(recommendations)} material_checked={material_checked} material_skipped={material_skipped} elapsed={now - progress_started:.1f}s"
                    )
            pool_wave_counts[pool_key] = wave
            if (
                pool_key not in exhausted_pools
                and pending_shortlist_by_pool.get(pool_key)
                and int(_recommendation_pool_shortages(theater_counts, deferred_pool_counts).get(pool_key, 0) or 0) > 0
            ):
                _pause_between_material_batches()

    english = config.LANGUAGE_CONFIG[2]
    if language_counts.get(english, 0) < rules.language_target(english):
        for candidate in deferred_english_balance:
            if language_counts.get(english, 0) >= rules.language_target(english):
                break
            if try_select_candidate(candidate, allow_english_balance_overflow=True):
                _decrement_deferred_pool_count(candidate, deferred_pool_counts)
    _log_step_result(logger, "[处理] 最终选取完成", count=len(recommendations), elapsed=time.perf_counter() - progress_started)

    _set_stat(stats, "materials", "checked", material_checked)
    _set_stat(stats, "materials", "skipped", material_skipped)
    _set_stat(stats, "materials", "prefetched", material_prefetched)
    _set_stat(stats, "materials", "qualified", len(material_qualified))
    _set_stat(stats, "materials", "rejected", max(material_checked - len(material_qualified), 0))
    _record_stage_counts(stats, "素材达标", material_qualified)
    _record_stage_counts(stats, "最终推荐", recommendations)
    _record_filter_reason_counts(stats, filter_records)
    if material_failed_registry_path and material_failed_rows:
        appended = append_material_failed_records(material_failed_registry_path, material_failed_rows)
        _set_stat(stats, "materials", "failed_recorded", appended)
        _log_info(logger, f"[处理] 不达标素材落表完成: count={appended}")
    for reason, count in _count_filter_reasons(filter_records).items():
        _set_stat(stats, "filtered", reason, count)
    _set_stat(stats, "final", "推荐结果", len(recommendations))
    _set_stat(stats, "final", "重复推荐", _count_duplicate_recommendations(recommendations))
    stats["quality_rows"] = build_quality_rows(recommendations, filter_records, scored)
    stats["coverage_rows"] = build_language_theater_coverage_rows(
        candidates=candidates,
        deduped=deduped,
        scored=scored,
        material_qualified=material_qualified,
        recommendations=recommendations,
        filters=filter_records,
    )
    for language, count in _count_languages(recommendations).items():
        _set_stat(stats, "final", f"语言:{language}", count)
    for language, shortage in _recommendation_shortages(recommendations).items():
        _set_stat(stats, "final", f"缺口:{language}", shortage)

    return models.TodayRecommendationRun(
        recommendations=recommendations,
        filters=filter_records,
        candidates=scored,
        stats=stats,
    )

def build_metadata_index(bundle: TodayInputBundle) -> Dict[Tuple[str, str], Dict[str, object]]:
    records: List[models.DramaRecord] = []
    for key in ("mobo_new", "beidou_new", "mobo_recommend", "beidou_hot", "duole_recommend"):
        records.extend(row for row in bundle.get(key, []) if isinstance(row, models.DramaRecord))
    return _build_metadata_index(rules.dedupe_candidates(records))


def _build_metadata_index(records: Sequence[models.DramaRecord]) -> Dict[Tuple[str, str], Dict[str, object]]:
    index: Dict[Tuple[str, str], Dict[str, object]] = {}
    for record in records:
        if not record.language or not record.title_norm:
            continue
        publish_at = rules.resolve_publish_at(record)
        tags = [str(tag).strip() for tag in record.tags if str(tag).strip()]
        raw = record.raw or {}
        content_tags = list(raw.get(rules.ALL_CONTENT_TAG_VALUES_RAW_KEY) or [])
        primary_content_tags = list(raw.get(rules.PRIMARY_CONTENT_TAG_VALUES_RAW_KEY) or [])
        if not publish_at and not tags and not content_tags:
            continue
        key = record.match_key
        index[key] = {
            "publish_at": publish_at,
            "tags": list(tags),
            "theater": rules.primary_theater(record),
            "sources": [source.strip() for source in str(record.source or "").split(",") if source.strip()],
            "content_tag_values": list(content_tags),
            "primary_content_tag_values": list(primary_content_tags),
            "merged_theaters": list(raw.get(rules.MERGED_THEATERS_RAW_KEY) or []),
            "earliest_publish_at": _clean_text(raw.get(rules.EARLIEST_PUBLISH_AT_RAW_KEY)),
            "latest_publish_at": _clean_text(raw.get(rules.LATEST_PUBLISH_AT_RAW_KEY)),
            "moboreels_publish_at": _clean_text(raw.get(rules.MOBOREELS_PUBLISH_AT_RAW_KEY)),
            "freshness_publish_at": _clean_text(raw.get(rules.FRESHNESS_PUBLISH_AT_RAW_KEY)),
        }
    return index


def _enrich_candidates_with_metadata(
    candidates: Sequence[models.DramaRecord],
    metadata_index: Mapping[Tuple[str, str], Dict[str, object]] | None = None,
    logger=None,
) -> Tuple[List[models.DramaRecord], int]:
    metadata_map = dict(metadata_index or {})
    if not metadata_map:
        return list(candidates), 0

    enriched: List[models.DramaRecord] = []
    changed = 0
    missing_before = 0
    for candidate in candidates:
        if not _needs_metadata_backfill(candidate):
            enriched.append(candidate)
            continue
        missing_before += 1
        metadata = metadata_map.get(candidate.match_key) or {}
        updated = _apply_candidate_metadata(candidate, metadata)
        if updated != candidate:
            changed += 1
        enriched.append(updated)
    if logger is not None:
        logger.info(f"元数据回填完成：before_missing={missing_before} changed={changed}")
    return enriched, changed


def _needs_metadata_backfill(candidate: models.DramaRecord) -> bool:
    publish_at = rules.resolve_publish_at(candidate)
    tags = [tag for tag in candidate.tags if str(tag).strip()]
    raw = candidate.raw or {}
    content_tags = list(raw.get(rules.ALL_CONTENT_TAG_VALUES_RAW_KEY) or [])
    return not publish_at or not tags or not content_tags


def _apply_candidate_metadata(candidate: models.DramaRecord, metadata: Mapping[str, object]) -> models.DramaRecord:
    publish_at = rules.resolve_publish_at(candidate)
    metadata_publish_at = _clean_text((metadata or {}).get("publish_at"))
    merged_publish_at = publish_at or metadata_publish_at
    metadata_tags = [str(tag).strip() for tag in list((metadata or {}).get("tags") or []) if str(tag).strip()]
    merged_tags = sorted({str(tag).strip() for tag in list(candidate.tags) + metadata_tags if str(tag).strip()})
    raw = dict(candidate.raw or {})
    updated = False
    metadata_content_tags = [str(tag).strip() for tag in list((metadata or {}).get("content_tag_values") or []) if str(tag).strip()]
    metadata_primary_content_tags = [
        str(tag).strip() for tag in list((metadata or {}).get("primary_content_tag_values") or []) if str(tag).strip()
    ]

    if metadata_publish_at:
        raw["metadata_publish_at"] = metadata_publish_at
        raw["metadata_lookup_sources"] = list((metadata or {}).get("sources") or [])
        raw["metadata_lookup_tags"] = list(metadata_tags)
        publish_values = list(raw.get(rules.PUBLISH_VALUES_RAW_KEY) or [])
        if metadata_publish_at not in publish_values:
            publish_values.append(metadata_publish_at)
            raw[rules.PUBLISH_VALUES_RAW_KEY] = publish_values
            updated = True
        if merged_publish_at:
            raw[rules.PRIMARY_PUBLISH_AT_RAW_KEY] = merged_publish_at
            updated = True
    for raw_key, meta_key in (
        (rules.ALL_CONTENT_TAG_VALUES_RAW_KEY, metadata_content_tags),
        (rules.PRIMARY_CONTENT_TAG_VALUES_RAW_KEY, metadata_primary_content_tags),
        (rules.MERGED_THEATERS_RAW_KEY, list((metadata or {}).get("merged_theaters") or [])),
    ):
        if meta_key and list(raw.get(raw_key) or []) != list(meta_key):
            raw[raw_key] = list(meta_key)
            updated = True
    for raw_key, meta_value in (
        (rules.EARLIEST_PUBLISH_AT_RAW_KEY, _clean_text((metadata or {}).get("earliest_publish_at"))),
        (rules.LATEST_PUBLISH_AT_RAW_KEY, _clean_text((metadata or {}).get("latest_publish_at"))),
        (rules.MOBOREELS_PUBLISH_AT_RAW_KEY, _clean_text((metadata or {}).get("moboreels_publish_at"))),
        (rules.FRESHNESS_PUBLISH_AT_RAW_KEY, _clean_text((metadata or {}).get("freshness_publish_at"))),
    ):
        if meta_value and _clean_text(raw.get(raw_key)) != meta_value:
            raw[raw_key] = meta_value
            updated = True

    theater_text = candidate.theater
    metadata_theater = rules.normalize_theater(_clean_text((metadata or {}).get("theater")))
    reordered_theaters = _reorder_theaters(candidate, metadata_theater)
    if reordered_theaters:
        theater_text = ",".join(reordered_theaters)
        raw[rules.PRIMARY_THEATER_RAW_KEY] = reordered_theaters[0]
        raw[rules.MERGED_THEATERS_RAW_KEY] = reordered_theaters
        if theater_text != candidate.theater:
            updated = True

    if merged_publish_at and merged_publish_at != candidate.publish_at:
        updated = True
    if merged_tags != list(candidate.tags):
        updated = True
    if not updated:
        return candidate
    return replace(candidate, publish_at=merged_publish_at, tags=merged_tags, theater=theater_text, raw=raw)


def _reorder_theaters(candidate: models.DramaRecord, metadata_theater: str) -> List[str]:
    raw = candidate.raw or {}
    merged = [rules.normalize_theater(name) for name in list(raw.get(rules.MERGED_THEATERS_RAW_KEY) or []) if rules.normalize_theater(name)]
    if not merged:
        merged = [rules.normalize_theater(name) for name in str(candidate.theater or "").split(",") if rules.normalize_theater(name)]
    if not merged:
        return []
    primary = metadata_theater if metadata_theater in merged else rules.primary_theater(candidate)
    primary = rules.normalize_theater(primary)
    if not primary:
        return merged
    ordered = [primary]
    ordered.extend(name for name in merged if name != primary)
    return ordered


def _latest_datetime_text(*values: object) -> str:
    latest = None
    texts: List[str] = []
    for value in values:
        text = _clean_text(value)
        if not text:
            continue
        texts.append(text)
        parsed = _parse_datetime_text(text)
        if parsed is None:
            continue
        if latest is None or parsed > latest:
            latest = parsed
    if latest is not None:
        return latest.strftime("%Y-%m-%d") if latest.time() == datetime.min.time() else latest.strftime("%Y-%m-%d %H:%M:%S")
    return texts[0] if texts else ""


def _parse_datetime_text(value: object) -> datetime | None:
    text = _clean_text(value).replace("/", "-").replace("T", " ")
    if not text:
        return None
    if text.endswith("Z"):
        text = text[:-1]
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S.%f"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def _filter_language_mismatches(
    candidates: Sequence[models.DramaRecord],
) -> Tuple[List[models.DramaRecord], List[models.FilterRecord]]:
    kept: List[models.DramaRecord] = []
    blocked: List[models.FilterRecord] = []
    allowed = set(config.LANGUAGE_ORDER)
    for candidate in candidates:
        candidate_language = runtime.normalize_language_name(candidate.language)
        if candidate_language not in allowed:
            blocked.append(_filter(candidate, "语种无效"))
            continue
        expected_language = _expected_language_from_record(candidate)
        if expected_language and expected_language != candidate_language:
            blocked.append(_filter(candidate, "语言识别冲突"))
            continue
        if candidate_language != candidate.language:
            kept.append(replace(candidate, language=candidate_language))
            continue
        kept.append(candidate)
    return kept, blocked


def _expected_language_from_record(candidate: models.DramaRecord) -> str:
    raw = candidate.raw or {}
    for key in ("language_name", "languageName", "language_str", "语言", "语种"):
        expected = runtime.normalize_language_name(raw.get(key))
        if expected:
            return expected
    return ""


def _recommendation_shortages(
    recommendations: Sequence[models.RecommendationResult],
) -> Dict[str, int]:
    current = Counter(item.language for item in recommendations)
    shortages: Dict[str, int] = {}
    for language in config.LANGUAGE_ORDER:
        target = rules.language_target(language)
        shortage = max(target - int(current.get(language, 0)), 0)
        if shortage > 0:
            shortages[language] = shortage
    return shortages


def _ai_anime_limit_block_reason(
    candidate: models.DramaRecord,
    ai_anime_counts: Mapping[str, int],
) -> str:
    limit = int(getattr(config, "MAX_AI_ANIME_PER_LANGUAGE", 0) or 0)
    if limit <= 0 or not rules.is_ai_or_anime(candidate):
        return ""
    if int(ai_anime_counts.get(candidate.language, 0) or 0) >= limit:
        return "AI/漫剧语种上限"
    return ""


def _build_english_origin_remaining(candidates: Sequence[models.DramaRecord]) -> Dict[str, int]:
    english = config.LANGUAGE_CONFIG[2]
    remaining = {"local": 0, "translated": 0, "unknown": 0}
    for candidate in candidates:
        if candidate.language != english:
            continue
        origin = rules.content_origin(candidate)
        remaining[origin] = remaining.get(origin, 0) + 1
    return remaining


def _consume_english_origin(candidate: models.DramaRecord, remaining: MutableMapping[str, int]) -> None:
    english = config.LANGUAGE_CONFIG[2]
    if candidate.language != english:
        return
    origin = rules.content_origin(candidate)
    remaining[origin] = max(int(remaining.get(origin, 0) or 0) - 1, 0)


def _english_origin_balance_block_reason(
    candidate: models.DramaRecord,
    language_counts: Mapping[str, int],
    selected_origin_counts: Mapping[str, int],
    remaining_origin_counts: Mapping[str, int],
) -> str:
    english = config.LANGUAGE_CONFIG[2]
    if candidate.language != english:
        return ""
    ratio = float(getattr(config, "ENGLISH_LOCAL_TRANSLATED_RATIO", 0.5) or 0.0)
    if ratio <= 0 or ratio >= 1:
        return ""
    total_target = rules.language_target(english)
    translated_target = int(total_target * ratio)
    local_target = max(total_target - translated_target, 0)
    origin = rules.content_origin(candidate)
    selected_local = int(selected_origin_counts.get("local", 0) or 0)
    selected_translated = int(selected_origin_counts.get("translated", 0) or 0)
    selected_total = int(language_counts.get(english, 0) or 0)
    remaining_slots = max(total_target - selected_total, 0)
    remaining_local = int(remaining_origin_counts.get("local", 0) or 0)
    remaining_translated = int(remaining_origin_counts.get("translated", 0) or 0)
    local_reserve = max(min(local_target, selected_local + remaining_local) - selected_local, 0)
    translated_reserve = max(min(translated_target, selected_translated + remaining_translated) - selected_translated, 0)
    remaining_after_select = max(remaining_slots - 1, 0)

    if origin == "local":
        if remaining_after_select < translated_reserve:
            return "英语本土/翻译配比保留"
        return ""
    if origin == "translated":
        if remaining_after_select < local_reserve:
            return "英语本土/翻译配比保留"
        return ""

    if remaining_slots <= 0:
        return ""
    if remaining_after_select < local_reserve + translated_reserve:
        return "英语本土/翻译配比保留"
    return ""


def _log_recommendation_health(logger, result: models.TodayRecommendationRun) -> None:
    final_count = len(result.recommendations)
    target_total = sum(rules.language_target(language) for language in config.LANGUAGE_ORDER)
    logger.info(f"最终达标推荐: {final_count}/{target_total}")
    shortages = _recommendation_shortages(result.recommendations)
    for language, shortage in shortages.items():
        logger.warning(f"{language} 推荐缺口 {shortage}")
    duplicate_count = _count_duplicate_recommendations(result.recommendations)
    if duplicate_count:
        logger.warning(f"最终推荐存在重复剧名 {duplicate_count} 条")


def export_offline_result(result: models.TodayRecommendationRun, output_root: Path, date_text: str) -> Dict[str, Path]:
    recommendation_headers = [
        "语言", "剧场", "剧名", "排名", "综合得分", "达标视频", "内容标签", "推荐语",
        "分数明细", "命中规则", "过滤链路", "来源时间", "质量提示",
    ]
    paths = {
        "stats": excel_io.build_output_path(output_root, date_text, 0, "拉取与处理统计"),
        "filters": excel_io.build_output_path(output_root, date_text, 7, "候选过滤明细"),
        "candidates": excel_io.build_output_path(output_root, date_text, 8, "加权推荐候选"),
        "recommendations": excel_io.build_output_path(output_root, date_text, 9, "最终达标推荐"),
    }
    excel_io.write_workbook(
        paths["stats"],
        {
            "汇总统计": build_stats_summary_rows(result.stats),
            "来源统计": list(result.stats.get("source_rows", [])),
            "处理统计": list(result.stats.get("stage_rows", [])),
            "质量健康": list(result.stats.get("quality_rows", [])),
            "语言剧场覆盖": list(result.stats.get("coverage_rows", [])),
        },
        sheet_headers={
            "汇总统计": ["分类", "指标", "条数"],
            "来源统计": ["阶段", "来源", "维度", "值", "条数"],
            "处理统计": ["阶段", "来源", "维度", "值", "条数"],
            "质量健康": ["指标", "维度", "值", "说明"],
            "语言剧场覆盖": [
                "语言", "剧场", "目标配额", "原始候选", "去重后", "进入打分", "素材达标", "最终推荐", "缺口", "过滤数", "主要过滤原因", "状态",
            ],
        },
    )
    excel_io.write_workbook(paths["filters"], {"过滤明细": [item.to_dict() for item in result.filters]})
    excel_io.write_workbook(paths["candidates"], {"加权候选": [item.to_dict() for item in result.candidates]})
    excel_io.write_workbook(
        paths["recommendations"],
        {"最终推荐": [item.to_export_dict() for item in result.recommendations]},
        sheet_headers={"最终推荐": recommendation_headers},
    )
    return paths


def build_bundle_stats(
    bundle: TodayInputBundle,
    fetch_stats_rows: Sequence[Dict[str, object]] | None = None,
) -> RunStats:
    stats = _empty_stats()
    stats["source_rows"].extend(_copy_rows(fetch_stats_rows))
    all_candidates = 0
    for source in ("mobo_new", "beidou_new", "mobo_recommend", "beidou_hot", "duole_recommend"):
        rows = [row for row in bundle.get(source, []) if isinstance(row, models.DramaRecord)]
        count = len(rows)
        all_candidates += count
        _set_stat(stats, "fetched", f"入编排:{source}", count)
        _record_source_counts(stats, "入编排", source, rows)

    published_rows = [row for row in bundle.get("published", []) if isinstance(row, models.PublishedRecord)]
    _set_stat(stats, "fetched", "入编排:published", len(published_rows))
    _set_stat(stats, "fetched", "all_candidates", all_candidates)
    _record_source_counts(stats, "入编排", "published", published_rows)
    adult_filter_rows = [row for row in bundle.get("adult_filter", []) if isinstance(row, models.TitleBlockRecord)]
    _set_stat(stats, "fetched", "入编排:adult_filter", len(adult_filter_rows))
    _record_source_counts(stats, "入编排", "adult_filter", adult_filter_rows)
    _apply_fetch_summary(stats)
    return stats


def build_stats_summary_rows(stats: RunStats) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    for category in ("fetched", "processed", "filtered", "materials", "final"):
        values = stats.get(category, {})
        if not isinstance(values, dict):
            continue
        for key in sorted(values):
            rows.append({"分类": category, "指标": key, "条数": values[key]})
    return rows


def prepare_run_stats(
    initial_stats: RunStats | None,
    candidates: Sequence[models.DramaRecord],
    published: Sequence[models.PublishedRecord],
) -> RunStats:
    stats = deepcopy(initial_stats) if initial_stats else _empty_stats()
    fetched = stats.setdefault("fetched", {})
    if not isinstance(fetched, dict):
        fetched = {}
        stats["fetched"] = fetched
    fetched.setdefault("all_candidates", len(candidates))
    fetched.setdefault("published", len(published))
    if not stats.get("source_rows"):
        _record_candidate_source_counts(stats, candidates, published)
    _apply_fetch_summary(stats)
    return stats


def _empty_stats() -> RunStats:
    return {
        "fetched": {},
        "processed": {},
        "filtered": {},
        "materials": {},
        "final": {},
        "source_rows": [],
        "stage_rows": [],
    }


def _record_candidate_source_counts(
    stats: RunStats,
    candidates: Sequence[models.DramaRecord],
    published: Sequence[models.PublishedRecord],
) -> None:
    per_source: Dict[str, List[models.DramaRecord]] = {}
    for candidate in candidates:
        source = candidate.source or "unknown"
        per_source.setdefault(source, []).append(candidate)
    for source, rows in sorted(per_source.items()):
        _record_source_counts(stats, "入编排", source, rows)
    _record_source_counts(stats, "入编排", "published", published)


def _record_source_counts(stats: RunStats, stage: str, source: str, rows: Sequence[object]) -> None:
    _append_count_row(stats["source_rows"], stage=stage, source=source, dimension="全部", value="全部", count=len(rows))
    for language, count in _count_languages(rows).items():
        _append_count_row(stats["source_rows"], stage=stage, source=source, dimension="语言", value=language, count=count)
    for theater, count in _count_theaters(rows).items():
        _append_count_row(stats["source_rows"], stage=stage, source=source, dimension="剧场", value=theater, count=count)
    for pair_key, count in _count_language_theater_pairs(rows).items():
        _append_count_row(stats["source_rows"], stage=stage, source=source, dimension="语言+剧场", value=pair_key, count=count)
    for sheet_name, count in _count_sheet_names(rows).items():
        _append_count_row(stats["source_rows"], stage=stage, source=source, dimension="Sheet", value=sheet_name, count=count)


def _record_stage_counts(stats: RunStats, stage: str, rows: Sequence[object]) -> None:
    _append_count_row(stats["stage_rows"], stage=stage, dimension="全部", value="全部", count=len(rows))
    for language, count in _count_languages(rows).items():
        _append_count_row(stats["stage_rows"], stage=stage, dimension="语言", value=language, count=count)
    for theater, count in _count_theaters(rows).items():
        _append_count_row(stats["stage_rows"], stage=stage, dimension="剧场", value=theater, count=count)
    for pair_key, count in _count_language_theater_pairs(rows).items():
        _append_count_row(stats["stage_rows"], stage=stage, dimension="语言+剧场", value=pair_key, count=count)


def _record_filter_reason_counts(stats: RunStats, filters: Sequence[models.FilterRecord]) -> None:
    for reason, count in _count_filter_reasons(filters).items():
        _append_count_row(stats["stage_rows"], stage="过滤原因", dimension="原因", value=reason, count=count)


def _record_layer_counts(stats: RunStats, rows: Sequence[models.DramaRecord]) -> None:
    for layer in config.LAYER_ORDER:
        layer_rows = [row for row in rows if getattr(row, "layer", "") == layer]
        _set_stat(stats, "processed", f"layer:{layer}", len(layer_rows))
        _record_stage_counts(stats, f"分层:{layer}", layer_rows)


def _append_count_row(
    rows: object,
    stage: str,
    dimension: str,
    value: str,
    count: int,
    source: str = "",
) -> None:
    if not isinstance(rows, list):
        return
    rows.append(
        {
            "阶段": stage,
            "来源": source,
            "维度": dimension,
            "值": value,
            "条数": int(count),
        }
    )


def _set_stat(stats: RunStats, category: str, key: str, value: int) -> None:
    bucket = stats.setdefault(category, {})
    if isinstance(bucket, dict):
        bucket[key] = int(value)


def _increment_stat(stats: RunStats, category: str, key: str, amount: int = 1) -> None:
    bucket = stats.setdefault(category, {})
    if isinstance(bucket, dict):
        bucket[key] = int(bucket.get(key, 0)) + int(amount)


def _count_languages(rows: Sequence[object]) -> Dict[str, int]:
    counter = Counter(str(getattr(row, "language", "") or "").strip() for row in rows)
    counter.pop("", None)
    return {
        language: counter[language]
        for language in sorted(counter, key=lambda item: (rules.language_sort_index(item), item))
    }


def _count_theaters(rows: Sequence[object]) -> Dict[str, int]:
    counter = Counter(str(getattr(row, "theater", "") or "").strip() for row in rows)
    counter.pop("", None)
    return {theater: counter[theater] for theater in sorted(counter)}


def _count_language_theater_pairs(rows: Sequence[object]) -> Dict[str, int]:
    counter = Counter()
    for row in rows:
        language = str(getattr(row, "language", "") or "").strip()
        theater = str(getattr(row, "theater", "") or "").strip()
        if not language or not theater:
            continue
        counter[f"{language} / {theater}"] += 1
    return {
        pair_key: counter[pair_key]
        for pair_key in sorted(counter, key=lambda item: (rules.language_sort_index(item.split(" / ", 1)[0]), item))
    }


def _count_sheet_names(rows: Sequence[object]) -> Dict[str, int]:
    counter = Counter()
    for row in rows:
        raw = getattr(row, "raw", {}) or {}
        sheet_name = str(raw.get("sheet_name") or "").strip()
        if sheet_name:
            counter[sheet_name] += 1
    return {sheet_name: counter[sheet_name] for sheet_name in sorted(counter)}


def _count_filter_reasons(filters: Sequence[models.FilterRecord]) -> Dict[str, int]:
    counter = Counter(filter_record.reason for filter_record in filters if filter_record.reason)
    return {reason: counter[reason] for reason in sorted(counter)}


def _count_duplicate_recommendations(rows: Sequence[models.RecommendationResult]) -> int:
    counter = Counter((row.language, models.normalize_title(row.title)) for row in rows)
    return sum(count - 1 for count in counter.values() if count > 1)


def _copy_rows(rows: Sequence[Dict[str, object]] | None) -> List[Dict[str, object]]:
    return [dict(row) for row in (rows or []) if isinstance(row, dict)]


def reset_client_fetch_stats(clients: Dict[str, object]) -> None:
    for client in clients.values():
        reset = getattr(client, "reset_fetch_stats", None)
        if callable(reset):
            reset()


def collect_client_fetch_stats(clients: Dict[str, object]) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    for client in clients.values():
        consume = getattr(client, "consume_fetch_stats", None)
        if callable(consume):
            for row in consume() or []:
                if isinstance(row, dict):
                    rows.append(dict(row))
            continue
        raw_rows = getattr(client, "fetch_stats", None)
        if isinstance(raw_rows, list):
            for row in raw_rows:
                if isinstance(row, dict):
                    rows.append(dict(row))
    return rows


def _apply_fetch_summary(stats: RunStats) -> None:
    source_rows = stats.get("source_rows", [])
    if not isinstance(source_rows, list):
        return
    stage_source_totals: Dict[Tuple[str, str], int] = {}
    stage_totals: Dict[str, int] = {}
    for row in source_rows:
        if not isinstance(row, dict):
            continue
        if str(row.get("维度") or "") != "全部":
            continue
        stage = str(row.get("阶段") or "").strip()
        source = str(row.get("来源") or "").strip()
        count = int(row.get("条数") or 0)
        stage_source_totals[(stage, source)] = stage_source_totals.get((stage, source), 0) + count
        stage_totals[stage] = stage_totals.get(stage, 0) + count
    for (stage, source), count in sorted(stage_source_totals.items()):
        _set_stat(stats, "fetched", f"{stage}:{source}", count)
    for stage, count in sorted(stage_totals.items()):
        _set_stat(stats, "fetched", f"{stage}:总计", count)


def _material_counts(material_results: MaterialMap, candidate: models.DramaRecord) -> Tuple[int, int]:
    key = _candidate_material_key(candidate)
    value = material_results.get(key)
    if isinstance(value, models.MaterialResult):
        return value.qualified_count, value.total_count
    if isinstance(value, int):
        return value, value
    return 0, 0


def load_material_failed_cooldown_keys(
    registry_path: Path | str | None,
    reference_date: str | None,
    cooldown_days: int,
) -> set[Tuple[str, str, str]]:
    if not registry_path or cooldown_days <= 0:
        return set()
    path = Path(registry_path)
    if not path.exists():
        return set()
    try:
        current_date = datetime.strptime(str(reference_date or datetime.now().strftime("%Y-%m-%d")), "%Y-%m-%d").date()
    except ValueError:
        current_date = datetime.now().date()
    cutoff = current_date - timedelta(days=max(cooldown_days - 1, 0))
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    keys: set[Tuple[str, str, str]] = set()
    try:
        worksheet = _material_failed_worksheet(workbook)
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return set()
        for row in rows[1:]:
            date_text = _clean_text(_cell_value(row, 0))
            record_date = _parse_date_only(date_text)
            if record_date is None or record_date < cutoff or record_date > current_date:
                continue
            language = runtime.normalize_language_name(_cell_value(row, 1))
            theater = rules.normalize_theater(_cell_value(row, 2))
            title = _clean_text(_cell_value(row, 3))
            if language and theater and title:
                keys.add(rules.material_key(language, title, theater))
    finally:
        workbook.close()
    return keys


def append_material_failed_records(registry_path: Path | str, rows: Sequence[Mapping[str, object]]) -> int:
    if not rows:
        return 0
    path = Path(registry_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = ["日期", "语言", "剧场", "剧名", "达标视频", "素材总数"]
    existing_keys: set[Tuple[str, str, str, str]] = set()
    if path.exists():
        workbook = openpyxl.load_workbook(path)
        worksheet = _material_failed_worksheet(workbook)
        header_values = [str(cell.value or "").strip() for cell in worksheet[1]]
        if not any(header_values):
            worksheet.append(headers)
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            key = _material_failed_registry_key(row)
            if key[0]:
                existing_keys.add(key)
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "failed_materials"
        worksheet.append(headers)
    appended = 0
    for row in rows:
        values = _material_failed_values(row)
        key = _material_failed_registry_key(values)
        if not key[0] or key in existing_keys:
            continue
        worksheet.append(values)
        existing_keys.add(key)
        appended += 1
    workbook.save(path)
    workbook.close()
    return appended


def _material_failed_row(
    candidate: models.DramaRecord,
    reference_date: str | None,
    qualified_count: int,
    total_count: int,
) -> Dict[str, object]:
    return {
        "date": str(reference_date or datetime.now().strftime("%Y-%m-%d")),
        "language": candidate.language,
        "theater": _material_lookup_theater(candidate),
        "title": candidate.title,
        "qualified_count": int(qualified_count or 0),
        "total_count": int(total_count or 0),
        "日期": str(reference_date or datetime.now().strftime("%Y-%m-%d")),
        "语言": candidate.language,
        "剧场": _material_lookup_theater(candidate),
        "剧名": candidate.title,
        "达标视频": int(qualified_count or 0),
        "素材总数": int(total_count or 0),
    }


def _material_failed_worksheet(workbook):
    for sheet_name in ("不达标素材", "failed_materials"):
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
    return workbook.active


def _material_failed_values(row: Mapping[str, object]) -> List[object]:
    return [
        row.get("日期") or row.get("date") or "",
        row.get("语言") or row.get("language") or "",
        row.get("剧场") or row.get("theater") or "",
        row.get("剧名") or row.get("剧集名称") or row.get("title") or "",
        row.get("达标视频") or row.get("qualified_count") or 0,
        row.get("素材总数") or row.get("total_count") or 0,
    ]


def _cell_value(row: Sequence[object], index: int | None) -> object:
    if index is None or index < 0 or index >= len(row):
        return ""
    return row[index]


def _parse_date_only(value: object):
    if isinstance(value, datetime):
        return value.date()
    text = _clean_text(value)
    if not text:
        return None
    for pattern in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10].replace("/", "-"), "%Y-%m-%d").date()
        except ValueError:
            continue
    return None


def _material_failed_registry_key(row: Sequence[object]) -> Tuple[str, str, str, str]:
    date_text = _clean_text(_cell_value(row, 0))
    language = runtime.normalize_language_name(_cell_value(row, 1))
    theater = rules.normalize_theater(_cell_value(row, 2))
    title = _clean_text(_cell_value(row, 3))
    if not date_text or not language or not theater or not title:
        return ("", "", "", "")
    return (date_text[:10], language, theater, models.normalize_title(title))


def _resolve_material_counts(
    material_results: MutableMapping[Tuple[str, str, str], int | models.MaterialResult],
    candidate: models.DramaRecord,
    material_client,
    material_start_date: str | None,
    material_end_date: str | None,
) -> Tuple[int, int, bool, bool]:
    key = _candidate_material_key(candidate)
    if key in material_results:
        qualified_count, total_count = _material_counts(material_results, candidate)
        return qualified_count, total_count, True, False
    if _should_query_material_client(material_client) and material_client is not None and material_start_date and material_end_date:
        try:
            material_results[key] = material_client.fetch_material_result(
                candidate.language,
                candidate.title,
                _material_lookup_theater(candidate),
                material_start_date,
                material_end_date,
            )
            qualified_count, total_count = _material_counts(material_results, candidate)
            return qualified_count, total_count, True, False
        except Exception:
            return config.VIDEO_THRESHOLD, config.VIDEO_THRESHOLD, False, True
    if not _should_query_material_client(material_client):
        return config.VIDEO_THRESHOLD, config.VIDEO_THRESHOLD, False, True
    return 0, 0, False, False


def _build_material_shortlist(
    candidates: Sequence[models.DramaRecord],
) -> Tuple[List[models.DramaRecord], List[models.FilterRecord], set[int]]:
    multiplier = max(int(getattr(config, "MATERIAL_SHORTLIST_MULTIPLIER", 3) or 3), 1)
    shortlist: List[models.DramaRecord] = []
    blocked: List[models.FilterRecord] = []
    seen: set[int] = set()
    counts: Dict[str, Dict[str, int]] = {}
    for candidate in candidates:
        quotas = rules.normalized_language_quotas(candidate.language)
        theater = rules.primary_theater(candidate)
        limit = int(quotas.get(theater, 0) or 0)
        if limit <= 0:
            blocked.append(_filter(candidate, "剧场不在语言配额内"))
            continue
        shortlist_limit = limit * multiplier
        counts.setdefault(candidate.language, {})
        if counts[candidate.language].get(theater, 0) >= shortlist_limit:
            continue
        shortlist.append(candidate)
        seen.add(id(candidate))
        counts[candidate.language][theater] = counts[candidate.language].get(theater, 0) + 1
    return shortlist, blocked, seen


def _group_shortlist_by_pool(
    candidates: Sequence[models.DramaRecord],
) -> Dict[Tuple[str, str], List[models.DramaRecord]]:
    grouped: Dict[Tuple[str, str], List[models.DramaRecord]] = {}
    for candidate in candidates:
        grouped.setdefault(_material_pool_key(candidate), []).append(candidate)
    return grouped


def _material_pool_order(candidates: Sequence[models.DramaRecord]) -> List[Tuple[str, str]]:
    ordered: List[Tuple[str, str]] = []
    seen: set[Tuple[str, str]] = set()
    for candidate in candidates:
        if not _is_material_pool_supported(candidate):
            continue
        key = _material_pool_key(candidate)
        if key in seen:
            continue
        seen.add(key)
        ordered.append(key)
    return ordered


def _extend_material_shortlist(
    candidates: Sequence[models.DramaRecord],
    seen: set[int],
    pool_shortages: Mapping[Tuple[str, str], int],
) -> List[models.DramaRecord]:
    multiplier = max(int(getattr(config, "MATERIAL_SHORTLIST_MULTIPLIER", 3) or 3), 1)
    targets = {
        key: max(int(shortage or 0), 0) * multiplier
        for key, shortage in pool_shortages.items()
        if int(shortage or 0) > 0
    }
    if not targets:
        return []
    batch_counts: Dict[Tuple[str, str], int] = {}
    batch: List[models.DramaRecord] = []
    for candidate in candidates:
        candidate_id = id(candidate)
        if candidate_id in seen:
            continue
        if not _is_material_pool_supported(candidate):
            continue
        pool_key = _material_pool_key(candidate)
        target = int(targets.get(pool_key, 0) or 0)
        if target <= 0:
            continue
        current = batch_counts.get(pool_key, 0)
        if current >= target:
            continue
        batch.append(candidate)
        seen.add(candidate_id)
        batch_counts[pool_key] = current + 1
        if all(batch_counts.get(key, 0) >= value for key, value in targets.items()):
            break
    return batch


def _recommendation_pool_shortages(
    theater_counts: Mapping[str, Mapping[str, int]],
    reserved_counts: Mapping[str, Mapping[str, int]] | None = None,
) -> Dict[Tuple[str, str], int]:
    shortages: Dict[Tuple[str, str], int] = {}
    reserved_counts = reserved_counts or {}
    for language, quotas in config.LANGUAGE_THEATER_QUOTAS.items():
        normalized = rules.normalized_language_quotas(language)
        selected_language_total = sum(int(value or 0) for value in theater_counts.get(language, {}).values())
        language_remaining = max(rules.language_target(language) - selected_language_total, 0)
        if language_remaining <= 0:
            continue
        for theater, limit in normalized.items():
            current = int(theater_counts.get(language, {}).get(theater, 0) or 0) + int(
                reserved_counts.get(language, {}).get(theater, 0) or 0
            )
            shortage = min(max(int(limit or 0) - current, 0), language_remaining)
            if shortage > 0:
                shortages[(language, theater)] = shortage
    return shortages


def _decrement_deferred_pool_count(
    candidate: models.DramaRecord,
    deferred_pool_counts: MutableMapping[str, MutableMapping[str, int]],
) -> None:
    language_counts = deferred_pool_counts.get(candidate.language)
    if not language_counts:
        return
    theater = rules.primary_theater(candidate)
    current = int(language_counts.get(theater, 0) or 0)
    if current <= 1:
        language_counts.pop(theater, None)
    else:
        language_counts[theater] = current - 1
    if not language_counts:
        deferred_pool_counts.pop(candidate.language, None)


def _format_pool_shortages(pool_shortages: Mapping[Tuple[str, str], int]) -> str:
    parts = [
        f"{language}/{theater}:{shortage}"
        for (language, theater), shortage in sorted(pool_shortages.items())
        if int(shortage or 0) > 0
    ]
    return ", ".join(parts)


def _material_pool_key(candidate: models.DramaRecord) -> Tuple[str, str]:
    return candidate.language, rules.primary_theater(candidate)


def _is_material_pool_supported(candidate: models.DramaRecord) -> bool:
    quotas = rules.normalized_language_quotas(candidate.language)
    theater = rules.primary_theater(candidate)
    return int(quotas.get(theater, 0) or 0) > 0


def _prefetch_shortlist_materials(
    candidates: Sequence[models.DramaRecord],
    material_results: MutableMapping[Tuple[str, str, str], int | models.MaterialResult],
    material_client,
    material_start_date: str | None,
    material_end_date: str | None,
    material_cooldown_keys: set[Tuple[str, str, str]] | None = None,
    logger=None,
) -> int:
    if not _should_query_material_client(material_client):
        return 0
    if material_client is None or not material_start_date or not material_end_date:
        return 0
    pending: List[models.DramaRecord] = []
    seen_keys: set[Tuple[str, str, str]] = set()
    cooldown_keys = material_cooldown_keys or set()
    for candidate in candidates:
        theater = _material_lookup_theater(candidate)
        if not candidate.title or not candidate.language or not theater:
            continue
        material_key = rules.material_key(candidate.language, candidate.title, theater)
        if material_key in cooldown_keys:
            continue
        if material_key in material_results or material_key in seen_keys:
            continue
        seen_keys.add(material_key)
        pending.append(replace(candidate, theater=theater))
    if not pending:
        return 0
    prefetch = getattr(material_client, "prefetch_material_results", None)
    if callable(prefetch):
        try:
            _log_info(
                logger,
                f"[处理] 素材批量校验开始: pending={len(pending)} workers={_material_prefetch_workers()} chunk={_material_prefetch_chunk_size()}",
            )
            prefetched = prefetch(pending, material_start_date, material_end_date)
        except Exception as exc:
            _log_info(logger, f"[处理] 素材批量校验降级: {type(exc).__name__}: {exc}")
            return 0
        material_results.update(prefetched)
        return len(prefetched)
    resolved = 0
    for candidate in pending:
        try:
            material_results[_candidate_material_key(candidate)] = material_client.fetch_material_result(
                candidate.language,
                candidate.title,
                candidate.theater,
                material_start_date,
                material_end_date,
            )
            resolved += 1
        except Exception as exc:
            _log_info(logger, f"[处理] 单条素材校验降级: {type(exc).__name__}: {exc}")
    return resolved


def _material_prefetch_chunk_size() -> int:
    return max(int(getattr(config, "MATERIAL_PREFETCH_CHUNK_SIZE", 1) or 1), 1)


def _material_prefetch_workers() -> int:
    return min(max(int(getattr(config, "MATERIAL_PREFETCH_WORKERS", 1) or 1), 1), 5)


def _material_max_expansion_waves_per_pool() -> int:
    return max(int(getattr(config, "MATERIAL_MAX_EXPANSION_WAVES_PER_POOL", 20) or 20), 1)


def _pause_between_material_batches() -> None:
    minimum = max(float(getattr(config, "MATERIAL_PREFETCH_PAUSE_MIN_SECONDS", 0.0) or 0.0), 0.0)
    maximum = max(float(getattr(config, "MATERIAL_PREFETCH_PAUSE_MAX_SECONDS", minimum) or minimum), minimum)
    if maximum <= 0:
        return
    time.sleep(random.uniform(minimum, maximum))

def _material_lookup_theater(candidate: models.DramaRecord) -> str:
    return rules.primary_theater(candidate)


def _candidate_material_key(candidate: models.DramaRecord) -> Tuple[str, str, str]:
    return rules.material_key(candidate.language, candidate.title, _material_lookup_theater(candidate))


def resolve_material_date_range(
    material_start_date: str | None,
    material_end_date: str | None,
    output_date: str | None = None,
) -> Tuple[str, str]:
    end_text = material_end_date or output_date or datetime.now().strftime("%Y-%m-%d")
    end_date = datetime.strptime(end_text, "%Y-%m-%d").date()
    if material_start_date:
        start_date = datetime.strptime(material_start_date, "%Y-%m-%d").date()
    else:
        start_date = end_date - timedelta(days=config.MATERIAL_LOOKBACK_DAYS)
    return start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")


def _filter(candidate: models.DramaRecord, reason: str) -> models.FilterRecord:
    return models.FilterRecord(
        title=candidate.title,
        language=candidate.language,
        theater=candidate.theater,
        reason=reason,
        source=candidate.source,
        score=candidate.score,
        layer=candidate.layer,
        age_bucket=candidate.age_bucket,
        explosion_bucket=candidate.explosion_bucket,
        raw=candidate.raw,
    )


def build_promotion_time(candidate: models.DramaRecord) -> str:
    latest = rules.resolve_promotion_datetime(candidate)
    if latest is not None:
        return latest.strftime("%Y-%m-%d") if latest.time() == datetime.min.time() else latest.strftime("%Y-%m-%d %H:%M:%S")
    for value in (
        (candidate.raw or {}).get("recommend_date"),
        (candidate.raw or {}).get("日期"),
        (candidate.raw or {}).get("更新日期"),
        (candidate.raw or {}).get("更新时间"),
        candidate.publish_at,
        (candidate.raw or {}).get("发布时间"),
        (candidate.raw or {}).get("上架时间"),
        (candidate.raw or {}).get("更新日期"),
        (candidate.raw or {}).get("更新时间"),
        (candidate.raw or {}).get("createTime"),
        (candidate.raw or {}).get("publishTime"),
    ):
        text = _clean_text(value)
        if text:
            return text
    return ""


def build_pre_rank_summary(candidate: models.DramaRecord) -> str:
    details = candidate.source_rank_details or {}
    if not details and candidate.source:
        details = {source.strip(): candidate.rank for source in str(candidate.source).split(",") if source.strip()}
    parts: List[str] = []
    for source, rank in sorted(details.items(), key=lambda item: (_source_sort_key(item[0]), item[1], item[0])):
        if not rank:
            continue
        parts.append(f"{_source_display_name(source)}第{rank}")
    return "、".join(parts)


def build_recommend_reason(
    candidate: models.DramaRecord,
    qualified_count: int,
    total_count: int,
    material_skipped: bool = False,
) -> str:
    parts: List[str] = []
    if candidate.layer:
        layer_text = candidate.layer
        if candidate.age_bucket or candidate.explosion_bucket:
            layer_text += f"层({candidate.age_bucket or '未知'}/{candidate.explosion_bucket or '无'})"
        parts.append(f"推荐分层 {layer_text}")

    promotion_time = build_promotion_time(candidate)
    if promotion_time:
        parts.append(f"推广时间 {promotion_time}")

    rank_summary = build_pre_rank_summary(candidate)
    if rank_summary:
        parts.append(f"推荐前命中 {rank_summary}")

    raw_reason = _pick_reason_text(candidate)
    if raw_reason:
        parts.append(f"外部推荐理由 {raw_reason}")

    content_label = build_content_label(candidate)
    if content_label:
        parts.append(f"内容标签 {content_label}")

    if candidate.matched_tags:
        parts.append(f"命中题材 {'/'.join(candidate.matched_tags[:5])}")

    if candidate.theme_multiplier and abs(candidate.theme_multiplier - 1.0) > 0.0001:
        parts.append(f"题材系数 {candidate.theme_multiplier:.2f}")

    if candidate.content_multiplier < 1.0:
        parts.append(f"内容系数 {candidate.content_multiplier:.1f}")

    source_count = len([source for source in str(candidate.source or "").split(",") if source.strip()])
    if source_count >= 2:
        parts.append(f"多源同时命中 {source_count} 个来源")

    if material_skipped:
        parts.append("素材默认全部达标，未请求素材接口")
    else:
        parts.append(f"素材达标 {qualified_count}/{total_count or qualified_count}")
    return "；".join(parts)


def build_score_breakdown(candidate: models.DramaRecord) -> str:
    parts = [
        f"综合={candidate.score:.4f}",
        f"层内={candidate.layer_score:.4f}",
        f"爆发={candidate.explosion_recommend_score:.4f}",
        f"新鲜={candidate.freshness_score:.4f}",
        f"收入={candidate.revenue_validation_score:.4f}",
        f"来源={candidate.source_signal_score:.4f}",
        f"标签={candidate.tag_score:.4f}",
        f"题材系数={candidate.theme_multiplier:.4f}",
        f"内容系数={candidate.content_multiplier:.4f}",
    ]
    return "；".join(parts)


def build_rule_hits(
    candidate: models.DramaRecord,
    qualified_count: int,
    total_count: int,
    material_skipped: bool = False,
) -> str:
    hits: List[str] = []
    if candidate.layer:
        hits.append(f"分层:{candidate.layer}")
    if candidate.age_bucket:
        hits.append(f"剧龄:{candidate.age_bucket}")
    if candidate.explosion_bucket:
        hits.append(f"爆发:{candidate.explosion_bucket}")
    if candidate.date_signal_hits:
        hits.extend(f"日期信号:{item}" for item in candidate.date_signal_hits[:5])
    if candidate.matched_tags:
        hits.append("题材:" + "/".join(candidate.matched_tags[:5]))
    content_label = build_content_label(candidate)
    if content_label:
        hits.append(f"内容:{content_label}")
    source_count = len([source for source in str(candidate.source or "").split(",") if source.strip()])
    if source_count >= 2:
        hits.append(f"多源:{source_count}")
    if material_skipped:
        hits.append("素材:跳过校验")
    else:
        hits.append(f"素材:{qualified_count}/{total_count or qualified_count}")
    return "；".join(hits)


def build_filter_trace(candidate: models.DramaRecord) -> str:
    return "候选采集 -> 去重合并 -> 元数据回填 -> 已发布过滤通过 -> 本地成人过滤通过 -> 语种校验通过 -> 内容过滤通过 -> 打分排序 -> 素材/配额校验通过 -> 最终推荐"


def build_source_dates(candidate: models.DramaRecord) -> str:
    details = candidate.source_date_details or {}
    if details:
        return "；".join(
            f"{_source_display_name(source)}:{date_text}"
            for source, date_text in sorted(details.items(), key=lambda item: (_source_sort_key(item[0]), item[0]))
            if date_text
        )
    promotion_time = build_promotion_time(candidate)
    publish_at = rules.resolve_publish_at(candidate) or candidate.publish_at
    parts = []
    if promotion_time:
        parts.append(f"推广:{promotion_time}")
    if publish_at:
        parts.append(f"发布:{publish_at}")
    return "；".join(parts)


def build_quality_flags(
    candidate: models.DramaRecord,
    qualified_count: int,
    total_count: int,
    material_skipped: bool = False,
) -> str:
    flags: List[str] = []
    if not candidate.tags:
        flags.append("缺少题材标签")
    if not rules.resolve_publish_at(candidate) and not candidate.publish_at:
        flags.append("缺少发布时间")
    if not candidate.source_rank_details and candidate.rank >= 999:
        flags.append("缺少来源排名")
    if material_skipped:
        flags.append("素材未实时校验")
    elif total_count <= 0:
        flags.append("素材总数为0")
    elif qualified_count < config.VIDEO_THRESHOLD:
        flags.append("素材低于阈值")
    if candidate.tag_status and "兜底" in candidate.tag_status:
        flags.append(candidate.tag_status)
    return "；".join(flags) or "正常"


def build_quality_rows(
    recommendations: Sequence[models.RecommendationResult],
    filters: Sequence[models.FilterRecord],
    candidates: Sequence[models.DramaRecord],
) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []

    def add(metric: str, dimension: str, value: object, note: str = "") -> None:
        rows.append({"指标": metric, "维度": dimension, "值": value, "说明": note})

    recommendation_count = len(recommendations)
    candidate_count = len(candidates)
    filter_count = len(filters)
    add("候选数", "全部", candidate_count, "进入打分后的候选数量")
    add("最终推荐数", "全部", recommendation_count, "最终导出的推荐数量")
    add("过滤数", "全部", filter_count, "所有过滤原因汇总数量")
    add("过滤率", "全部", _ratio_text(filter_count, candidate_count + filter_count), "过滤数 / 过滤前候选估算")
    add("重复推荐数", "全部", _count_duplicate_recommendations(recommendations), "同语种同剧名重复")

    for language in config.LANGUAGE_ORDER:
        language_recommendations = [item for item in recommendations if item.language == language]
        target = rules.language_target(language)
        add("语种达成", language, f"{len(language_recommendations)}/{target}", "最终推荐数 / 目标数")

    theater_counter = Counter(item.theater for item in recommendations if item.theater)
    for theater, count in sorted(theater_counter.items()):
        add("剧场覆盖", theater, count, "最终推荐剧场分布")

    flag_counter = Counter()
    for item in recommendations:
        for flag in str(getattr(item, "quality_flags", "") or "").split("；"):
            flag = flag.strip()
            if flag and flag != "正常":
                flag_counter[flag] += 1
    if not flag_counter:
        add("质量提示", "正常", 0, "未发现推荐解释层面的质量提示")
    else:
        for flag, count in sorted(flag_counter.items()):
            add("质量提示", flag, count, "最终推荐中的提示次数")

    score_values = [float(item.score or 0.0) for item in recommendations]
    if score_values:
        add("分数分布", "最高", round(max(score_values), 4), "最终推荐综合得分")
        add("分数分布", "最低", round(min(score_values), 4), "最终推荐综合得分")
        add("分数分布", "平均", round(sum(score_values) / len(score_values), 4), "最终推荐综合得分")

    return rows


def build_language_theater_coverage_rows(
    candidates: Sequence[models.DramaRecord],
    deduped: Sequence[models.DramaRecord],
    scored: Sequence[models.DramaRecord],
    material_qualified: Sequence[models.DramaRecord],
    recommendations: Sequence[models.RecommendationResult],
    filters: Sequence[models.FilterRecord],
) -> List[Dict[str, object]]:
    candidate_counts = _count_target_pool_rows(candidates)
    deduped_counts = _count_target_pool_rows(deduped)
    scored_counts = _count_target_pool_rows(scored)
    material_counts = _count_target_pool_rows(material_qualified)
    recommendation_counts = _count_target_pool_rows(recommendations)
    filter_counts, filter_reasons = _count_target_pool_filters(filters)

    rows: List[Dict[str, object]] = []
    for language in config.LANGUAGE_ORDER:
        quotas = rules.normalized_language_quotas(language)
        for theater, quota in sorted(quotas.items()):
            target = int(quota or 0)
            if target <= 0:
                continue
            key = (language, theater)
            final_count = recommendation_counts.get(key, 0)
            shortage = max(target - final_count, 0)
            status = "达标" if shortage <= 0 else "缺口"
            if candidate_counts.get(key, 0) <= 0:
                status = "无候选"
            elif scored_counts.get(key, 0) <= 0:
                status = "过滤后无候选"
            elif material_counts.get(key, 0) <= 0 and _is_realtime_material_check_enabled():
                status = "素材未达标"
            rows.append(
                {
                    "语言": language,
                    "剧场": theater,
                    "目标配额": target,
                    "原始候选": candidate_counts.get(key, 0),
                    "去重后": deduped_counts.get(key, 0),
                    "进入打分": scored_counts.get(key, 0),
                    "素材达标": material_counts.get(key, 0),
                    "最终推荐": final_count,
                    "缺口": shortage,
                    "过滤数": filter_counts.get(key, 0),
                    "主要过滤原因": _format_reason_counts(filter_reasons.get(key, Counter())),
                    "状态": status,
                }
            )
    return rows


def _count_target_pool_rows(rows: Sequence[object]) -> Dict[Tuple[str, str], int]:
    counter: Counter[Tuple[str, str]] = Counter()
    for row in rows:
        language = str(getattr(row, "language", "") or "").strip()
        theater_text = str(getattr(row, "theater", "") or "").strip()
        if not language or not theater_text:
            continue
        theater = rules.normalize_theater(str(theater_text).split(",", 1)[0])
        if theater and theater in set(rules.normalized_language_quotas(language)):
            counter[(language, theater)] += 1
    return dict(counter)


def _count_target_pool_filters(
    filters: Sequence[models.FilterRecord],
) -> Tuple[Dict[Tuple[str, str], int], Dict[Tuple[str, str], Counter[str]]]:
    counts: Counter[Tuple[str, str]] = Counter()
    reasons: Dict[Tuple[str, str], Counter[str]] = {}
    for row in filters:
        language = str(getattr(row, "language", "") or "").strip()
        theater_text = str(getattr(row, "theater", "") or "").strip()
        reason = str(getattr(row, "reason", "") or "").strip() or "未知"
        if not language or not theater_text:
            continue
        theater = rules.normalize_theater(str(theater_text).split(",", 1)[0])
        if theater and theater in set(rules.normalized_language_quotas(language)):
            key = (language, theater)
            counts[key] += 1
            reasons.setdefault(key, Counter())[reason] += 1
    return dict(counts), reasons


def _format_reason_counts(counter: Counter[str]) -> str:
    if not counter:
        return ""
    return "；".join(f"{reason}:{count}" for reason, count in counter.most_common(3))


def _ratio_text(numerator: int, denominator: int) -> str:
    if denominator <= 0:
        return "0.00%"
    return f"{(numerator / denominator) * 100:.2f}%"


def _is_realtime_material_check_enabled() -> bool:
    return bool(getattr(config, "MATERIAL_CHECK_ENABLED", False)) and not bool(
        getattr(config, "MATERIAL_POST_VALIDATION_ENABLED", False)
    )


def _should_query_material_client(material_client) -> bool:
    return material_client is not None


def build_content_label(candidate: models.DramaRecord) -> str:
    return "/".join(rules.content_labels(candidate))


def _pick_reason_text(candidate: models.DramaRecord) -> str:
    raw = candidate.raw or {}
    for key in ("recommend_reason", "推荐理由", "理由", "备注"):
        text = _clean_text(raw.get(key))
        if text:
            return text[:80]
    return ""


def _source_display_name(source: str) -> str:
    return SOURCE_DISPLAY_NAMES.get(source, source)


def _source_sort_key(source: str) -> int:
    order = ["mobo_new", "beidou_new", "mobo_recommend", "beidou_hot", "beidou_income", "duole_recommend"]
    try:
        return order.index(source)
    except ValueError:
        return len(order)


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def main() -> None:
    args = parse_args()
    date_value = args.date or datetime.now().strftime("%Y-%m-%d")
    logger = runtime.PipelineLogger("today_recommend", run_date=date_value)
    exit_code = 0
    try:
        runtime.bootstrap_runtime(config_module=config, logger=logger)
        paths = run_today_recommendation(
            selected_languages=args.languages or None,
            date_text=date_value,
            material_start_date=args.material_start or None,
            material_end_date=args.material_end or None,
            logger=logger,
        )
        print("今日剧集推荐输出完成:", flush=True)
        for key, path in paths.items():
            print(f"  {key}: {path}", flush=True)
    except Exception as exc:
        logger.exception(exc, "今日推荐执行失败")
        log_path = logger.flush_if_needed()
        if log_path is not None:
            print(f"  debug_log: {log_path}", flush=True)
        if runtime.is_frozen_exe():
            import traceback

            traceback.print_exc()
            exit_code = 1
        else:
            raise
    else:
        log_path = logger.flush_if_needed()
        if log_path is not None:
            print(f"  debug_log: {log_path}", flush=True)
    finally:
        runtime.pause_before_exit()
    if exit_code:
        raise SystemExit(exit_code)


if __name__ == "__main__":
    main()
