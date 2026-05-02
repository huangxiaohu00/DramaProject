import importlib
import tempfile
import unittest
from pathlib import Path

import openpyxl


TEST_TMP_ROOT = Path("drama_pipeline") / ".tmp_tests"
TEST_TMP_ROOT.mkdir(parents=True, exist_ok=True)


class ExcelIoTests(unittest.TestCase):
    def test_write_workbook_creates_headers_and_rows(self):
        excel_io = importlib.import_module("drama_pipeline.5_excel_io")

        with tempfile.TemporaryDirectory(dir=TEST_TMP_ROOT) as tmp:
            path = Path(tmp) / "out.xlsx"
            excel_io.write_workbook(path, {"Sheet": [{"语言": "英语", "剧名": "A"}]})
            workbook = openpyxl.load_workbook(path)
            try:
                worksheet = workbook["Sheet"]
                self.assertEqual(worksheet.cell(1, 1).value, "剧名")
                self.assertEqual(worksheet.cell(1, 2).value, "语言")
                self.assertEqual(worksheet.cell(2, 1).value, "A")
                self.assertEqual(worksheet.cell(2, 2).value, "英语")
            finally:
                workbook.close()

    def test_prefixed_output_path_uses_stage_number_and_name(self):
        excel_io = importlib.import_module("drama_pipeline.5_excel_io")

        path = excel_io.build_output_path(Path("root"), "2026-04-19", 9, "最终达标推荐")

        self.assertEqual(path.parts[-2:], ("2026-04-19", "9_最终达标推荐.xlsx"))

    def test_write_workbook_can_write_headers_without_rows(self):
        excel_io = importlib.import_module("drama_pipeline.5_excel_io")

        with tempfile.TemporaryDirectory(dir=TEST_TMP_ROOT) as tmp:
            path = Path(tmp) / "out.xlsx"
            excel_io.write_workbook(path, {"Errors": []}, sheet_headers={"Errors": ["平台", "账号"]})
            workbook = openpyxl.load_workbook(path)
            try:
                worksheet = workbook["Errors"]
                self.assertEqual(worksheet.cell(1, 1).value, "平台")
                self.assertEqual(worksheet.cell(1, 2).value, "账号")
                self.assertEqual(worksheet.max_row, 1)
            finally:
                workbook.close()

    def test_write_workbook_respects_explicit_header_order_with_rows(self):
        excel_io = importlib.import_module("drama_pipeline.5_excel_io")

        with tempfile.TemporaryDirectory(dir=TEST_TMP_ROOT) as tmp:
            path = Path(tmp) / "ordered.xlsx"
            excel_io.write_workbook(
                path,
                {"Stats": [{"条数": 2, "分类": "fetched", "指标": "all_candidates"}]},
                sheet_headers={"Stats": ["分类", "指标", "条数"]},
            )
            workbook = openpyxl.load_workbook(path)
            try:
                worksheet = workbook["Stats"]
                self.assertEqual(worksheet.cell(1, 1).value, "分类")
                self.assertEqual(worksheet.cell(1, 2).value, "指标")
                self.assertEqual(worksheet.cell(1, 3).value, "条数")
                self.assertEqual(worksheet.cell(2, 1).value, "fetched")
                self.assertEqual(worksheet.cell(2, 2).value, "all_candidates")
                self.assertEqual(worksheet.cell(2, 3).value, 2)
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
