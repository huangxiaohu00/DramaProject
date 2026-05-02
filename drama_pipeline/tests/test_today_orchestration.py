import importlib
import os
import tempfile
import time
import unittest
from pathlib import Path

import openpyxl


models = importlib.import_module("drama_pipeline.3_models")
TEST_TMP_ROOT = Path("drama_pipeline") / ".tmp_tests"
TEST_TMP_ROOT.mkdir(parents=True, exist_ok=True)
WRITABLE_TMP_ROOT = Path(os.environ.get("TEMP") or os.environ.get("TMP") or ".")


class FakeMoboClient:
    def __init__(self):
        self.new_languages = []
        self.recommend_languages = []
        self.new_calls = []
        self.recommend_calls = []

    def fetch_new_dramas(self, language, theater=""):
        self.new_languages.append(language)
        self.new_calls.append((language, theater))
        theater_text = theater or "MoboReels"
        return [models.DramaRecord(title=f"{language} {theater_text} Mobo New", language=language, theater=theater_text, source="mobo_new", rank=1)]

    def fetch_recommend_dramas(self, language, theater=""):
        self.recommend_languages.append(language)
        self.recommend_calls.append((language, theater))
        theater_text = theater or "ShortMax"
        return [models.DramaRecord(title=f"{language} {theater_text} Mobo Rec", language=language, theater=theater_text, source="mobo_recommend", rank=1)]


class FakeBeidouClient:
    def __init__(self):
        self.new_languages = []
        self.new_calls = []
        self.income_languages = []

    def fetch_new_dramas(self, language, theater=""):
        self.new_languages.append(language)
        self.new_calls.append((language, theater))
        if theater:
            return [models.DramaRecord(title=f"{language} {theater} New", language=language, theater=theater, source="beidou_new", rank=1)]
        return [
            models.DramaRecord(title=f"{language} Beidou New", language=language, theater="ShortMax", source="beidou_new", rank=1),
            models.DramaRecord(title=f"{language} Mobo Duplicate", language=language, theater="MoboReels", source="beidou_new", rank=2),
        ]

    def fetch_income_dramas(self, language):
        self.income_languages.append(language)
        return [models.DramaRecord(title=f"{language} Income Hot", language=language, theater="ShortMax", source="beidou_income", rank=1)]


class FakeFeishuClient:
    def __init__(self):
        self.languages = []
        self.hot_calls = 0

    def fetch_published(self, language):
        self.languages.append(language)
        return [models.PublishedRecord(title=f"{language} Published", language=language, theater="ShortMax")]

    def fetch_beidou_hot_dramas(self):
        self.hot_calls += 1
        return [models.DramaRecord(title="English Hot", language="英语", theater="ShortMax", source="beidou_hot", rank=1)]


class FakeDuoleClient:
    def fetch_recommend_dramas(self):
        return [models.DramaRecord(title="Duole Rec", language="法语", theater="DramaBox", source="duole_recommend", rank=1)]


class FakeMaterialClient:
    def __init__(self, counts=None):
        self.calls = []
        self.counts = counts or {}

    def fetch_material_result(self, language, title, theater, start_date, end_date):
        self.calls.append((language, title, theater, start_date, end_date))
        count = self.counts.get((language, title, theater), 0)
        return models.MaterialResult(language=language, title=title, theater=theater, qualified_count=count, total_count=count)

    def prefetch_material_results(self, candidates, start_date, end_date):
        results = {}
        for candidate in candidates:
            result = self.fetch_material_result(candidate.language, candidate.title, candidate.theater, start_date, end_date)
            results[(candidate.language, candidate.title_norm, candidate.theater)] = result
        return results


class FailingMaterialClient:
    def fetch_material_result(self, language, title, theater, start_date, end_date):
        raise RuntimeError("429")

    def prefetch_material_results(self, candidates, start_date, end_date):
        raise RuntimeError("429")


class FakeMetadataClient:
    def __init__(self, mapping=None):
        self.mapping = mapping or {}
        self.calls = []

    def lookup_drama_metadata(self, title, language=""):
        self.calls.append(("mobo", title, language))
        return dict(self.mapping.get((language, title), {}))

    def lookup_task_metadata(self, title, language=""):
        self.calls.append(("beidou", title, language))
        return dict(self.mapping.get((language, title), {}))


class FakeLogger:
    def __init__(self):
        self.messages = []

    def info(self, message):
        self.messages.append(("info", message))

    def warning(self, message):
        self.messages.append(("warning", message))

    def step(self, message):
        self.messages.append(("step", message))


class TodayOrchestrationTests(unittest.TestCase):
    def test_build_offline_recommendation_skips_material_check_by_default(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")

        result = today.build_offline_recommendation(
            candidates=[models.DramaRecord(title="A", language="英语", theater="ShortMax", source="mobo_recommend", rank=1)],
            published=[],
            material_results={},
            material_client=None,
        )

        self.assertEqual(len(result.recommendations), 1)
        self.assertIn("素材默认全部达标，未请求素材接口", result.recommendations[0].recommend_reason)
        self.assertEqual(result.stats["materials"]["checked"], 0)
        self.assertEqual(result.stats["materials"]["skipped"], 1)

    def test_build_offline_recommendation_continues_when_material_api_fails(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")

        result = today.build_offline_recommendation(
            candidates=[models.DramaRecord(title="A", language="英语", theater="ShortMax", source="mobo_recommend", rank=1)],
            published=[],
            material_results={},
            material_client=FailingMaterialClient(),
            material_start_date="2024-04-20",
            material_end_date="2026-04-19",
        )

        self.assertEqual(len(result.recommendations), 1)
        self.assertIn("素材默认全部达标，未请求素材接口", result.recommendations[0].recommend_reason)
        self.assertEqual(result.stats["materials"]["checked"], 0)
        self.assertEqual(result.stats["materials"]["skipped"], 1)

    def test_collect_today_inputs_only_queries_selected_languages_and_filters_beidou_moboreels(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")
        mobo = FakeMoboClient()
        beidou = FakeBeidouClient()
        feishu = FakeFeishuClient()

        bundle = today.collect_today_inputs(
            mobo_client=mobo,
            beidou_client=beidou,
            feishu_client=feishu,
            duole_client=FakeDuoleClient(),
            selected_languages=["德语", "法语"],
        )

        self.assertEqual(
            sorted(mobo.new_calls),
            sorted(
                [
                    ("德语", "MoboReels"),
                    ("德语", "ShortMax"),
                    ("法语", "MoboReels"),
                    ("法语", "ShortMax"),
                    ("法语", "HoneyReels"),
                ]
            ),
        )
        self.assertEqual(sorted(mobo.recommend_calls), sorted(mobo.new_calls))
        self.assertEqual(
            sorted(beidou.new_calls),
            sorted(
                [
                    ("德语", "ShortMax"),
                    ("德语", "FlareFlow"),
                    ("德语", "ReelShort"),
                    ("德语", "DramaBox"),
                    ("法语", "ShortMax"),
                    ("法语", "HoneyReels"),
                    ("法语", "ReelShort"),
                    ("法语", "DramaBox"),
                ]
            ),
        )
        self.assertEqual(beidou.income_languages, ["德语", "法语"])
        self.assertEqual(feishu.languages, ["德语", "法语"])
        self.assertEqual(feishu.hot_calls, 0)
        self.assertEqual(
            sorted(row.theater for row in bundle["beidou_new"]),
            sorted(["ShortMax", "FlareFlow", "ReelShort", "DramaBox", "ShortMax", "HoneyReels", "ReelShort", "DramaBox"]),
        )
        self.assertEqual([row.title for row in bundle["beidou_hot"]], ["德语 Income Hot", "法语 Income Hot"])
        self.assertEqual(len(bundle["duole_recommend"]), 1)

    def test_collect_today_inputs_fetches_english_hot_only_when_english_selected(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")
        beidou = FakeBeidouClient()
        feishu = FakeFeishuClient()

        bundle = today.collect_today_inputs(
            mobo_client=FakeMoboClient(),
            beidou_client=beidou,
            feishu_client=feishu,
            duole_client=FakeDuoleClient(),
            selected_languages=["英语"],
        )

        self.assertEqual([row.title for row in bundle["beidou_hot"]], ["English Hot"])
        self.assertEqual(beidou.income_languages, [])
        self.assertEqual(feishu.hot_calls, 1)

    def test_collect_today_inputs_combines_feishu_english_and_beidou_non_english_hot_sources(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")
        beidou = FakeBeidouClient()
        feishu = FakeFeishuClient()

        bundle = today.collect_today_inputs(
            mobo_client=FakeMoboClient(),
            beidou_client=beidou,
            feishu_client=feishu,
            duole_client=FakeDuoleClient(),
            selected_languages=["英语", "法语"],
        )

        self.assertEqual(beidou.income_languages, ["法语"])
        self.assertEqual(feishu.hot_calls, 1)
        self.assertEqual([row.title for row in bundle["beidou_hot"]], ["法语 Income Hot", "English Hot"])

    def test_export_today_stage_files_writes_prefixed_workbooks(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")
        bundle = {
            "mobo_new": [models.DramaRecord(title="A", language="英语", theater="MoboReels")],
            "beidou_new": [],
            "mobo_recommend": [],
            "beidou_hot": [],
            "duole_recommend": [],
            "published": [models.PublishedRecord(title="P", language="英语", theater="ShortMax")],
        }

        with tempfile.TemporaryDirectory(dir=TEST_TMP_ROOT) as tmp:
            paths = today.export_today_stage_files(bundle, Path(tmp), "2026-04-19")
            self.assertTrue(paths["mobo_new"].name.startswith("1_"))
            self.assertTrue(paths["published"].name.startswith("6_"))
            workbook = openpyxl.load_workbook(paths["mobo_new"])
            try:
                values = [workbook.active.cell(2, column).value for column in range(1, workbook.active.max_column + 1)]
                self.assertIn("英语", values)
            finally:
                workbook.close()

    def test_run_today_recommendation_writes_stage_and_final_files_with_injected_clients(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")

        with tempfile.TemporaryDirectory(dir=TEST_TMP_ROOT) as tmp:
            paths = today.run_today_recommendation(
                selected_languages=["英语"],
                date_text="2026-04-19",
                output_root=Path(tmp),
                material_results={("英语", "英语 shortmax mobo rec", "ShortMax"): 12},
                clients={
                    "mobo_client": FakeMoboClient(),
                    "beidou_client": FakeBeidouClient(),
                    "feishu_client": FakeFeishuClient(),
                    "duole_client": FakeDuoleClient(),
                    "material_client": FakeMaterialClient(),
                },
            )

            self.assertTrue(paths["mobo_new"].exists())
            self.assertTrue(paths["stats"].exists())
            self.assertTrue(paths["recommendations"].exists())
            workbook = openpyxl.load_workbook(paths["recommendations"])
            try:
                headers = [workbook.active.cell(1, column).value for column in range(1, workbook.active.max_column + 1)]
                values = [workbook.active.cell(2, column).value for column in range(1, workbook.active.max_column + 1)]
                self.assertIn("英语 ShortMax Mobo Rec", values)
                self.assertEqual(
                    headers,
                    [
                        "语言",
                        "剧场",
                        "剧名",
                        "排名",
                        "综合得分",
                        "达标视频",
                        "内容标签",
                        "推荐语",
                        "分数明细",
                        "命中规则",
                        "过滤链路",
                        "来源时间",
                        "质量提示",
                    ],
                )
                self.assertTrue(any(isinstance(value, str) and "素材达标" in value for value in values))
            finally:
                workbook.close()
            stats_workbook = openpyxl.load_workbook(paths["stats"])
            try:
                summary = stats_workbook["汇总统计"]
                values = [summary.cell(2, column).value for column in range(1, summary.max_column + 1)]
                self.assertIn("fetched", values)
                self.assertIn("all_candidates", values)
                summary_rows = [
                    [summary.cell(row, column).value for column in range(1, summary.max_column + 1)]
                    for row in range(2, summary.max_row + 1)
                ]
                self.assertTrue(any(row[0] == "processed" and str(row[1]).startswith("layer:") for row in summary_rows))
                source_sheet = stats_workbook["来源统计"]
                source_rows = [
                    [source_sheet.cell(row, column).value for column in range(1, source_sheet.max_column + 1)]
                    for row in range(2, source_sheet.max_row + 1)
                ]
                self.assertTrue(any(row[0] == "入编排" and row[1] == "mobo_recommend" and row[2] == "剧场" and row[3] == "ShortMax" for row in source_rows))
                self.assertTrue(any(row[0] == "入编排" and row[1] == "mobo_recommend" and row[2] == "语言+剧场" and row[3] == "英语 / ShortMax" for row in source_rows))
            finally:
                stats_workbook.close()

    def test_run_today_recommendation_fetches_material_when_no_material_results_are_injected(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")
        material = FakeMaterialClient({("英语", "英语 ShortMax Mobo Rec", "ShortMax"): 12})

        with tempfile.TemporaryDirectory(dir=TEST_TMP_ROOT) as tmp:
            paths = today.run_today_recommendation(
                selected_languages=["英语"],
                date_text="2026-04-19",
                output_root=Path(tmp),
                material_start_date="2024-04-20",
                material_end_date="2026-04-19",
                clients={
                    "mobo_client": FakeMoboClient(),
                    "beidou_client": FakeBeidouClient(),
                    "feishu_client": FakeFeishuClient(),
                    "duole_client": FakeDuoleClient(),
                    "material_client": material,
                },
            )

            self.assertIn(("英语", "英语 ShortMax Mobo Rec", "ShortMax", "2024-04-20", "2026-04-19"), material.calls)
            workbook = openpyxl.load_workbook(paths["recommendations"])
            try:
                values = [workbook.active.cell(2, column).value for column in range(1, workbook.active.max_column + 1)]
                self.assertIn("英语 ShortMax Mobo Rec", values)
            finally:
                workbook.close()

    
    def test_run_today_recommendation_uses_local_material_results_then_fetches_missing_online(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")
        english = today.config.LANGUAGE_CONFIG[2]
        material = FakeMaterialClient(
            {
                (english, f"{english} MoboReels Mobo New", "MoboReels"): 12,
                (english, f"{english} ShortMax Mobo Rec", "ShortMax"): 12,
            }
        )

        with tempfile.TemporaryDirectory(dir=TEST_TMP_ROOT) as tmp:
            today.run_today_recommendation(
                selected_languages=[english],
                date_text="2026-04-19",
                output_root=Path(tmp),
                material_results={(english, f"{english} MoboReels Mobo New".lower(), "MoboReels"): 12},
                material_start_date="2024-04-20",
                material_end_date="2026-04-19",
                clients={
                    "mobo_client": FakeMoboClient(),
                    "beidou_client": FakeBeidouClient(),
                    "feishu_client": FakeFeishuClient(),
                    "duole_client": FakeDuoleClient(),
                    "material_client": material,
                },
            )

        self.assertIn((english, f"{english} ShortMax Mobo Rec", "ShortMax", "2024-04-20", "2026-04-19"), material.calls)
    def test_build_offline_recommendation_marks_unsupported_theater_separately(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")

        result = today.build_offline_recommendation(
            candidates=[models.DramaRecord(title="A", language="英语", theater="DramaBox", source="duole_recommend", rank=1)],
            published=[],
            material_results={("英语", "a", "DramaBox"): 12},
        )

        self.assertEqual(result.recommendations, [])
        self.assertEqual(len(result.filters), 1)
        self.assertEqual(result.filters[0].reason, "剧场不在语言配额内")

    def test_build_offline_recommendation_filters_local_adult_block_by_language_and_title(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")

        result = today.build_offline_recommendation(
            candidates=[
                models.DramaRecord(title="A", language="英语", theater="ShortMax", source="mobo_recommend", rank=1),
                models.DramaRecord(title="B", language="英语", theater="ShortMax", source="mobo_recommend", rank=2),
            ],
            published=[],
            material_results={("英语", "a", "ShortMax"): 12, ("英语", "b", "ShortMax"): 12},
            adult_blocks=[models.TitleBlockRecord(title="A", language="英语", theater="DramaBox", source="adult_filter")],
        )

        self.assertEqual([item.title for item in result.recommendations], ["B"])
        self.assertTrue(any(item.reason == "本地成人过滤" for item in result.filters))

    def test_material_prefetch_is_limited_to_shortlist_candidates(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")
        english = today.config.LANGUAGE_CONFIG[2]
        old_quotas = dict(today.config.LANGUAGE_THEATER_QUOTAS.get(english, {}))
        old_post_validation = getattr(today.config, "MATERIAL_POST_VALIDATION_ENABLED", True)
        material = FakeMaterialClient(
            {
                (english, "A", "ShortMax"): 12,
                (english, "B", "ShortMax"): 12,
                (english, "C", "ShortMax"): 12,
            }
        )

        try:
            today.config.LANGUAGE_THEATER_QUOTAS[english] = {"ShortMax": 1}
            today.config.MATERIAL_POST_VALIDATION_ENABLED = False
            result = today.build_offline_recommendation(
                candidates=[
                    models.DramaRecord(title="A", language=english, theater="ShortMax", source="mobo_recommend", rank=1),
                    models.DramaRecord(title="B", language=english, theater="ShortMax", source="mobo_recommend", rank=2),
                    models.DramaRecord(title="C", language=english, theater="ShortMax", source="mobo_recommend", rank=3),
                    models.DramaRecord(title="D", language=english, theater="ShortMax", source="mobo_recommend", rank=4),
                ],
                published=[],
                material_results={},
                material_client=material,
                material_start_date="2024-04-20",
                material_end_date="2026-04-19",
            )
        finally:
            today.config.LANGUAGE_THEATER_QUOTAS[english] = old_quotas
            today.config.MATERIAL_POST_VALIDATION_ENABLED = old_post_validation

        self.assertEqual([item.title for item in result.recommendations], ["A"])
        self.assertEqual(
            material.calls,
            [
                (english, "A", "ShortMax", "2024-04-20", "2026-04-19"),
                (english, "B", "ShortMax", "2024-04-20", "2026-04-19"),
                (english, "C", "ShortMax", "2024-04-20", "2026-04-19"),
                (english, "D", "ShortMax", "2024-04-20", "2026-04-19"),
            ],
        )

    def test_shortlist_expands_when_material_results_do_not_fill_quota(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")
        english = today.config.LANGUAGE_CONFIG[2]
        old_quotas = dict(today.config.LANGUAGE_THEATER_QUOTAS.get(english, {}))
        old_multiplier = today.config.MATERIAL_SHORTLIST_MULTIPLIER
        old_post_validation = getattr(today.config, "MATERIAL_POST_VALIDATION_ENABLED", True)
        material = FakeMaterialClient(
            {
                (english, "B", "ShortMax"): 12,
            }
        )

        try:
            today.config.LANGUAGE_THEATER_QUOTAS[english] = {"ShortMax": 1}
            today.config.MATERIAL_SHORTLIST_MULTIPLIER = 1
            today.config.MATERIAL_POST_VALIDATION_ENABLED = False
            result = today.build_offline_recommendation(
                candidates=[
                    models.DramaRecord(title="A", language=english, theater="ShortMax", source="mobo_recommend", rank=1),
                    models.DramaRecord(title="B", language=english, theater="ShortMax", source="mobo_recommend", rank=2),
                    models.DramaRecord(title="C", language=english, theater="ShortMax", source="mobo_recommend", rank=3),
                ],
                published=[],
                material_results={},
                material_client=material,
                material_start_date="2024-04-20",
                material_end_date="2026-04-19",
            )
        finally:
            today.config.LANGUAGE_THEATER_QUOTAS[english] = old_quotas
            today.config.MATERIAL_SHORTLIST_MULTIPLIER = old_multiplier
            today.config.MATERIAL_POST_VALIDATION_ENABLED = old_post_validation

        self.assertEqual([item.title for item in result.recommendations], ["B"])
        self.assertEqual(
            material.calls,
            [
                (english, "A", "ShortMax", "2024-04-20", "2026-04-19"),
                (english, "B", "ShortMax", "2024-04-20", "2026-04-19"),
            ],
        )

    def test_shortlist_expands_by_language_theater_pool_before_other_pools(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")
        english = today.config.LANGUAGE_CONFIG[2]
        french = today.config.LANGUAGE_CONFIG[5]
        old_english_quotas = dict(today.config.LANGUAGE_THEATER_QUOTAS.get(english, {}))
        old_french_quotas = dict(today.config.LANGUAGE_THEATER_QUOTAS.get(french, {}))
        old_multiplier = today.config.MATERIAL_SHORTLIST_MULTIPLIER
        material = FakeMaterialClient(
            {
                (english, "B", "ShortMax"): 12,
                (french, "C", "DramaBox"): 12,
            }
        )

        try:
            today.config.LANGUAGE_THEATER_QUOTAS[english] = {"ShortMax": 1}
            today.config.LANGUAGE_THEATER_QUOTAS[french] = {"DramaBox": 1}
            today.config.MATERIAL_SHORTLIST_MULTIPLIER = 1
            result = today.build_offline_recommendation(
                candidates=[
                    models.DramaRecord(title="A", language=english, theater="ShortMax", source="mobo_recommend", rank=1),
                    models.DramaRecord(title="C", language=french, theater="DramaBox", source="duole_recommend", rank=2),
                    models.DramaRecord(title="B", language=english, theater="ShortMax", source="mobo_recommend", rank=3),
                ],
                published=[],
                material_results={},
                material_client=material,
                material_start_date="2024-04-20",
                material_end_date="2026-04-19",
            )
        finally:
            today.config.LANGUAGE_THEATER_QUOTAS[english] = old_english_quotas
            today.config.LANGUAGE_THEATER_QUOTAS[french] = old_french_quotas
            today.config.MATERIAL_SHORTLIST_MULTIPLIER = old_multiplier

        self.assertEqual(
            {(item.language, item.title) for item in result.recommendations},
            {(english, "B"), (french, "C")},
        )
        self.assertEqual(
            material.calls,
            [
                (english, "A", "ShortMax", "2024-04-20", "2026-04-19"),
                (english, "B", "ShortMax", "2024-04-20", "2026-04-19"),
                (french, "C", "DramaBox", "2024-04-20", "2026-04-19"),
            ],
        )

    def test_build_offline_recommendation_uses_primary_theater_for_material_lookup_after_cross_theater_merge(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")
        material = FakeMaterialClient({("法语", "A", "ShortMax"): 12})

        result = today.build_offline_recommendation(
            candidates=[
                models.DramaRecord(
                    title="A",
                    language="法语",
                    theater="ShortMax",
                    source="beidou_new",
                    rank=5,
                    publish_at="2026-04-01 10:00:00",
                ),
                models.DramaRecord(
                    title="A",
                    language="法语",
                    theater="DramaBox",
                    source="duole_recommend",
                    rank=30,
                    raw={"上架时间": "2026-04-10"},
                ),
            ],
            published=[],
            material_results={},
            material_client=material,
            material_start_date="2024-04-20",
            material_end_date="2026-04-19",
        )

        self.assertEqual(len(result.recommendations), 1)
        self.assertEqual(result.recommendations[0].theater, "ShortMax,DramaBox")
        self.assertIn(("法语", "A", "ShortMax", "2024-04-20", "2026-04-19"), material.calls)

    def test_build_offline_recommendation_with_material_client_and_logger_does_not_raise(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")
        logger = FakeLogger()
        material = FakeMaterialClient({("英语", "A", "ShortMax"): 12})

        result = today.build_offline_recommendation(
            candidates=[models.DramaRecord(title="A", language="英语", theater="ShortMax", source="mobo_recommend", rank=1)],
            published=[],
            material_results={},
            material_client=material,
            material_start_date="2024-04-20",
            material_end_date="2026-04-19",
            logger=logger,
        )

        self.assertEqual(len(result.recommendations), 1)
        self.assertTrue(any(level == "info" for level, _ in logger.messages))

    def test_build_offline_recommendation_marks_ai_and_anime_in_final_output(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")

        result = today.build_offline_recommendation(
            candidates=[
                models.DramaRecord(
                    title="A",
                    language="法语",
                    theater="DramaBox",
                    source="duole_recommend",
                    rank=1,
                    tags=["AI转绘", "动漫"],
                )
            ],
            published=[],
            material_results={("法语", "a", "DramaBox"): 12},
        )

        self.assertEqual(len(result.recommendations), 1)
        self.assertEqual(result.recommendations[0].content_label, "AI剧/漫剧")
        self.assertIn("内容标签 AI剧/漫剧", result.recommendations[0].recommend_reason)

    def test_build_metadata_index_backfills_publish_time_and_tags(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")
        metadata_index = today.build_metadata_index(
            {
                "mobo_new": [
                    models.DramaRecord(
                        title="A",
                        language="法语",
                        theater="DramaBox",
                        source="mobo_new",
                        publish_at="2026-04-18",
                        tags=["复仇"],
                    )
                ],
                "beidou_new": [],
                "mobo_recommend": [],
                "beidou_hot": [],
                "duole_recommend": [],
                "published": [],
            }
        )

        result = today.build_offline_recommendation(
            candidates=[
                models.DramaRecord(
                    title="A",
                    language="法语",
                    theater="DramaBox",
                    source="duole_recommend",
                    rank=1,
                    raw={"recommend_date": "2026-04-20"},
                )
            ],
            published=[],
            material_results={("法语", "a", "DramaBox"): 12},
            metadata_index=metadata_index,
        )

        self.assertEqual(result.candidates[0].publish_at, "2026-04-18")
        self.assertEqual(result.candidates[0].age_bucket, "新")
        self.assertEqual(result.candidates[0].tags, ["复仇"])

    def test_build_offline_recommendation_reorders_primary_theater_from_metadata_index(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")
        material = FakeMaterialClient({("法语", "A", "DramaBox"): 12})
        metadata_index = {
            ("法语", "a"): {"publish_at": "2026-04-18", "tags": ["爱情"], "theater": "DramaBox", "sources": ["beidou_new"]}
        }

        result = today.build_offline_recommendation(
            candidates=[
                models.DramaRecord(
                    title="A",
                    language="法语",
                    theater="ShortMax",
                    source="beidou_hot",
                    rank=1,
                    raw={"recommend_date": "2026-04-20"},
                ),
                models.DramaRecord(
                    title="A",
                    language="法语",
                    theater="DramaBox",
                    source="duole_recommend",
                    rank=2,
                    raw={"recommend_date": "2026-04-19"},
                ),
            ],
            published=[],
            material_results={},
            material_client=material,
            material_start_date="2024-04-20",
            material_end_date="2026-04-19",
            metadata_index=metadata_index,
        )

        self.assertEqual(result.recommendations[0].theater, "DramaBox,ShortMax")
        self.assertIn(("法语", "A", "DramaBox", "2024-04-20", "2026-04-19"), material.calls)

    def test_build_offline_recommendation_limits_ai_anime_per_language(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")
        old_limit = today.config.MAX_AI_ANIME_PER_LANGUAGE
        try:
            today.config.MAX_AI_ANIME_PER_LANGUAGE = 3
            result = today.build_offline_recommendation(
                candidates=[
                    models.DramaRecord(
                        title=f"AI {idx}",
                        language="法语",
                        theater="DramaBox",
                        source="duole_recommend",
                        rank=idx,
                        tags=["AI转绘"],
                    )
                    for idx in range(1, 5)
                ],
                published=[],
                material_results={("法语", f"ai {idx}", "DramaBox"): 12 for idx in range(1, 5)},
            )
        finally:
            today.config.MAX_AI_ANIME_PER_LANGUAGE = old_limit

        self.assertEqual(len(result.recommendations), 3)
        self.assertEqual(result.stats["filtered"]["AI/漫剧语种上限"], 1)

    def test_build_offline_recommendation_balances_english_local_and_translated(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")
        old_ratio = today.config.ENGLISH_LOCAL_TRANSLATED_RATIO
        old_quotas = dict(today.config.LANGUAGE_THEATER_QUOTAS)
        try:
            today.config.ENGLISH_LOCAL_TRANSLATED_RATIO = 0.5
            today.config.LANGUAGE_THEATER_QUOTAS = {
                **today.config.LANGUAGE_THEATER_QUOTAS,
                "英语": {"ShortMax": 10},
            }
            candidates = []
            material_results = {}
            for idx in range(1, 9):
                title = f"Local {idx}"
                candidates.append(
                    models.DramaRecord(
                        title=title,
                        language="英语",
                        theater="ShortMax",
                        source="mobo_recommend",
                        rank=idx,
                        publish_at="2026-04-20",
                        source_rank_details={"mobo_recommend": idx},
                        tags=["本土剧"],
                    )
                )
                material_results[("英语", title.lower(), "ShortMax")] = 12
            for idx in range(1, 5):
                title = f"Translated {idx}"
                candidates.append(
                    models.DramaRecord(
                        title=title,
                        language="英语",
                        theater="ShortMax",
                        source="mobo_recommend",
                        rank=20 + idx,
                        publish_at="2026-04-20",
                        source_rank_details={"mobo_recommend": 20 + idx},
                        tags=["版权剧"],
                    )
                )
                material_results[("英语", title.lower(), "ShortMax")] = 12

            result = today.build_offline_recommendation(
                candidates=candidates,
                published=[],
                material_results=material_results,
            )
        finally:
            today.config.ENGLISH_LOCAL_TRANSLATED_RATIO = old_ratio
            today.config.LANGUAGE_THEATER_QUOTAS = old_quotas

        local_count = sum(1 for item in result.recommendations if item.title.startswith("Local"))
        translated_count = sum(1 for item in result.recommendations if item.title.startswith("Translated"))
        self.assertEqual(translated_count, 4)
        self.assertEqual(local_count, 6)

    def test_deferred_english_balance_counts_as_pool_capacity(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")
        english = today.config.LANGUAGE_CONFIG[2]
        old_ratio = today.config.ENGLISH_LOCAL_TRANSLATED_RATIO
        old_quotas = dict(today.config.LANGUAGE_THEATER_QUOTAS)
        old_multiplier = today.config.MATERIAL_SHORTLIST_MULTIPLIER
        old_post_validation = getattr(today.config, "MATERIAL_POST_VALIDATION_ENABLED", True)
        logger = FakeLogger()

        try:
            today.config.ENGLISH_LOCAL_TRANSLATED_RATIO = 0.5
            today.config.LANGUAGE_THEATER_QUOTAS = {
                **today.config.LANGUAGE_THEATER_QUOTAS,
                english: {"ShortMax": 1, "SnackShort": 1},
            }
            today.config.MATERIAL_SHORTLIST_MULTIPLIER = 1
            today.config.MATERIAL_POST_VALIDATION_ENABLED = True
            result = today.build_offline_recommendation(
                candidates=[
                    models.DramaRecord(
                        title="Local ShortMax",
                        language=today.config.LANGUAGE_CONFIG[2],
                        theater="ShortMax",
                        source="mobo_recommend",
                        rank=1,
                        tags=["本土剧"],
                    ),
                    models.DramaRecord(
                        title="Translated SnackShort",
                        language=today.config.LANGUAGE_CONFIG[2],
                        theater="SnackShort",
                        source="mobo_recommend",
                        rank=2,
                        tags=["版权剧"],
                    ),
                    models.DramaRecord(
                        title="Extra SnackShort",
                        language=today.config.LANGUAGE_CONFIG[2],
                        theater="SnackShort",
                        source="mobo_recommend",
                        rank=3,
                        tags=["本土剧"],
                    ),
                ],
                published=[],
                material_results={},
                logger=logger,
            )
        finally:
            today.config.ENGLISH_LOCAL_TRANSLATED_RATIO = old_ratio
            today.config.LANGUAGE_THEATER_QUOTAS = old_quotas
            today.config.MATERIAL_SHORTLIST_MULTIPLIER = old_multiplier
            today.config.MATERIAL_POST_VALIDATION_ENABLED = old_post_validation

        self.assertEqual([item.title for item in result.recommendations], ["Local ShortMax", "Translated SnackShort"])
        self.assertFalse(any("扩容 英语/SnackShort" in message for _, message in logger.messages))

    def test_material_failed_registry_skips_recent_failed_titles(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")
        english = today.config.LANGUAGE_CONFIG[2]
        material = FakeMaterialClient({(english, "Recent Failed", "ShortMax"): 12})
        registry_path = TEST_TMP_ROOT / f"material_failed_records_skip_{time.time_ns()}.xlsx"

        today.append_material_failed_records(
            registry_path,
            [
                {
                    "date": "2026-04-24",
                    "language": english,
                    "theater": "ShortMax",
                    "title": "Recent Failed",
                    "qualified_count": 0,
                    "total_count": 0,
                }
            ],
        )
        result = today.build_offline_recommendation(
            candidates=[
                models.DramaRecord(
                    title="Recent Failed",
                    language=english,
                    theater="ShortMax",
                    source="mobo_recommend",
                    rank=1,
                )
            ],
            published=[],
            material_results={},
            material_client=material,
            material_start_date="2024-04-25",
            material_end_date="2026-04-25",
            reference_date="2026-04-25",
            material_failed_registry_path=registry_path,
        )

        self.assertEqual(len(result.recommendations), 0)
        self.assertEqual(material.calls, [])
        self.assertEqual(result.stats["filtered"]["素材冷却期未达标"], 1)

    def test_material_failed_registry_records_checked_failures(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")
        english = today.config.LANGUAGE_CONFIG[2]
        material = FakeMaterialClient({(english, "No Material", "ShortMax"): 0})
        registry_path = TEST_TMP_ROOT / f"material_failed_records_write_{time.time_ns()}.xlsx"

        result = today.build_offline_recommendation(
            candidates=[
                models.DramaRecord(
                    title="No Material",
                    language=english,
                    theater="ShortMax",
                    source="mobo_recommend",
                    rank=1,
                )
            ],
            published=[],
            material_results={},
            material_client=material,
            material_start_date="2024-04-25",
            material_end_date="2026-04-25",
            reference_date="2026-04-25",
            material_failed_registry_path=registry_path,
        )
        keys = today.load_material_failed_cooldown_keys(registry_path, "2026-04-26", 7)

        self.assertEqual(len(result.recommendations), 0)
        self.assertIn((english, "no material", "ShortMax"), keys)
        self.assertEqual(result.stats["materials"]["failed_recorded"], 1)

    def test_collect_today_inputs_limits_duole_rows_per_sheet_after_language_filter(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")

        class RichDuoleClient:
            def fetch_recommend_dramas(self):
                rows = []
                for index in range(1, 651):
                    rows.append(
                        models.DramaRecord(
                            title=f"English {index}",
                            language="英语",
                            theater="DramaBox",
                            source="duole_recommend",
                            rank=index,
                            raw={"sheet_name": "2.推荐剧单"},
                        )
                    )
                for index in range(1, 451):
                    rows.append(
                        models.DramaRecord(
                            title=f"French {index}",
                            language="法语",
                            theater="DramaBox",
                            source="duole_recommend",
                            rank=index,
                            raw={"sheet_name": "13.DramaBox小语种"},
                        )
                    )
                for index in range(1, 51):
                    rows.append(
                        models.DramaRecord(
                            title=f"Japanese {index}",
                            language="日语",
                            theater="DramaBox",
                            source="duole_recommend",
                            rank=index,
                            raw={"sheet_name": "13.DramaBox小语种"},
                        )
                    )
                return rows

        bundle = today.collect_today_inputs(
            mobo_client=FakeMoboClient(),
            beidou_client=FakeBeidouClient(),
            feishu_client=FakeFeishuClient(),
            duole_client=RichDuoleClient(),
            selected_languages=["英语", "法语"],
        )

        duole_rows = bundle["duole_recommend"]
        self.assertEqual(len([row for row in duole_rows if row.raw.get("sheet_name") == "2.推荐剧单"]), 600)
        self.assertEqual(len([row for row in duole_rows if row.raw.get("sheet_name") == "13.DramaBox小语种"]), 400)
        self.assertEqual(duole_rows[0].rank, 1)
        self.assertEqual(duole_rows[599].rank, 600)
        self.assertEqual(duole_rows[600].rank, 1)
        self.assertEqual(duole_rows[-1].rank, 400)

    def test_load_local_adult_filter_reads_same_folder_excel(self):
        today = importlib.import_module("drama_pipeline.6_today_recommend")
        runtime = importlib.import_module("drama_pipeline.10_runtime")

        with tempfile.TemporaryDirectory(dir=TEST_TMP_ROOT) as tmp:
            path = runtime.ensure_adult_filter_workbook(Path(tmp) / runtime.ADULT_FILTER_WORKBOOK_NAME)
            workbook = openpyxl.load_workbook(path)
            worksheet = workbook[runtime.ADULT_FILTER_SHEET]
            worksheet.append(["英语", "DramaBox", "Adult A"])
            worksheet.append(["法语", "", "Adult B"])
            workbook.save(path)
            workbook.close()

            rows = today.load_local_adult_filter(app_dir=Path(tmp))

        self.assertEqual(
            [(row.language, row.theater, row.title) for row in rows],
            [("英语", "DramaBox", "Adult A"), ("法语", "", "Adult B")],
        )


if __name__ == "__main__":
    unittest.main()
