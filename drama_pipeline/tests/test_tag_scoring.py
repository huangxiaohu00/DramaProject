import importlib
import math
import tempfile
import unittest
from pathlib import Path

import openpyxl


def _build_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    language = workbook.active
    language.title = "按语言偏好标签"
    language.append(["语言", "标签", "金额", "订单数", "排名", "推荐权重"])
    language.append(["英语", "总裁", 1000, 100, 1, 1.0])
    language.append(["英语", "复仇", 200, 10, 2, 0.4])
    language.append(["德语", "复仇", 100, 5, 3, 0.3])

    theater = workbook.create_sheet("按剧场偏好标签")
    theater.append(["剧场", "标签", "金额", "订单数", "排名", "推荐权重"])
    theater.append(["ShortMax", "复仇", 300, 30, 1, 0.2])

    language_theater = workbook.create_sheet("按语言剧场偏好标签")
    language_theater.append(["语言", "剧场", "标签", "金额", "订单数", "排名", "推荐权重"])
    language_theater.append(["英语", "ShortMax", "复仇", 50, 2, 1, 1.0])

    normalization = workbook.create_sheet("标签归一化检查")
    normalization.append(["raw_tag", "normalized_tag"])
    normalization.append(["revenge", "复仇"])
    normalization.append(["period", "古代"])

    workbook.save(path)


class TagScoringTests(unittest.TestCase):
    def _scorer(self):
        scoring = importlib.import_module("drama_pipeline.8_tag_scoring")
        tempdir = tempfile.TemporaryDirectory()
        self.addCleanup(tempdir.cleanup)
        workbook_path = Path(tempdir.name) / "tags.xlsx"
        _build_workbook(workbook_path)
        return scoring.TagWeightScorer(workbook_path)

    def test_global_score_uses_language_sheet_aggregate_only(self):
        scorer = self._scorer()

        boss = scorer.global_weight("总裁")
        revenge = scorer.global_weight("复仇")

        self.assertAlmostEqual(boss.score, 1.2, places=4)
        self.assertGreaterEqual(revenge.score, 0.8)
        self.assertLess(revenge.score, 1.0)
        self.assertEqual(boss.orders, 100)
        self.assertEqual(revenge.orders, 15)

    def test_small_sample_language_theater_score_shrinks_toward_global(self):
        scorer = self._scorer()

        result = scorer.tag_weight("英语", "ShortMax", "复仇")
        language_theater = next(item for item in result.evidence if item.level == "language_theater")
        global_score = scorer.global_weight("复仇").score

        expected_confidence = math.log1p(2) / math.log1p(30)
        expected_calibrated = global_score + expected_confidence * (1.2 - global_score)
        self.assertAlmostEqual(language_theater.final_confidence, expected_confidence, places=4)
        self.assertAlmostEqual(language_theater.calibrated_score, expected_calibrated, places=4)
        self.assertLess(language_theater.calibrated_score, 1.2)

    def test_multiple_levels_are_confidence_weighted(self):
        scorer = self._scorer()

        result = scorer.tag_weight("英语", "ShortMax", "复仇")
        levels = {item.level for item in result.evidence}

        self.assertEqual(levels, {"language_theater", "language", "theater", "global"})
        self.assertGreaterEqual(result.weight, 0.8)
        self.assertLessEqual(result.weight, 1.2)
        self.assertEqual(result.status, "标签命中")

    def test_score_tags_uses_top_three_decay_and_default(self):
        scorer = self._scorer()

        hit = scorer.score_tags("英语", "ShortMax", ["revenge"])
        fallback = scorer.score_tags("英语", "ShortMax", [])

        self.assertEqual(hit.status, "标签命中")
        self.assertEqual(hit.matched_tags, ["复仇"])
        self.assertGreaterEqual(hit.score, 0.8)
        self.assertLessEqual(hit.score, 1.2)
        self.assertEqual(fallback.score, 1.0)
        self.assertEqual(fallback.status, "无标签兜底")

    def test_score_record_uses_merged_tags_and_best_matching_theater(self):
        scorer = self._scorer()
        models = importlib.import_module("drama_pipeline.3_models")

        record = models.DramaRecord(
            title="A",
            language="英语",
            theater="DramaBox,ShortMax",
            tags=["revenge", "revenge"],
            raw={"merged_theaters": ["DramaBox", "ShortMax"]},
        )

        merged = scorer.score_record(record)
        shortmax = scorer.score_tags("英语", "ShortMax", ["revenge"])
        dramabox = scorer.score_tags("英语", "DramaBox", ["revenge"])

        self.assertEqual(merged.matched_tags, ["复仇"])
        self.assertAlmostEqual(merged.score, max(shortmax.score, dramabox.score), places=4)

    def test_excluded_broad_tags_do_not_participate_in_theme_scoring(self):
        scorer = self._scorer()

        fallback = scorer.score_tags("英语", "ShortMax", ["古代", "女频", "都市"])
        mixed = scorer.score_tags("英语", "ShortMax", ["revenge", "现代", "女频"])
        semi_broad = scorer.score_tags("英语", "ShortMax", ["现代言情", "古代言情", "都市情感"])

        self.assertEqual(fallback.score, 1.0)
        self.assertEqual(fallback.matched_tags, [])
        self.assertEqual(mixed.matched_tags, ["复仇"])
        self.assertEqual(semi_broad.score, 1.0)
        self.assertEqual(semi_broad.matched_tags, [])


if __name__ == "__main__":
    unittest.main()
