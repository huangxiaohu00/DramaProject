import importlib
import unittest
import shutil
import uuid
from pathlib import Path

import openpyxl


models = importlib.import_module("drama_pipeline.3_models")


class FakeLogger:
    def __init__(self):
        self.infos = []
        self.warnings = []

    def info(self, message):
        self.infos.append(str(message))

    def warning(self, message):
        self.warnings.append(str(message))


def workspace_tempdir():
    base = Path("D:/project/DramaProject/drama_pipeline/.tmp_tests")
    base.mkdir(parents=True, exist_ok=True)
    path = base / f"tmp_{uuid.uuid4().hex}"
    path.mkdir(parents=True, exist_ok=True)
    return path


class FakeOrderClient:
    def __init__(self, platform, account, rows, stats=None):
        self.platform = platform
        self.account = account
        self.rows = rows
        self.last_order_fetch_stats = stats or {}
        self.calls = []

    def fetch_orders(self, begin_date, end_date, account=""):
        self.calls.append((begin_date, end_date, account))
        output = []
        for row in self.rows:
            output.append(
                models.OrderRecord(
                    date=begin_date,
                    title=row["title"],
                    platform=self.platform,
                    language=row["language"],
                    theater=row["theater"],
                    order_count=row["order_count"],
                    amount=row["amount"],
                    account=account or self.account,
                    task_id=row.get("task_id", ""),
                    order_type=row.get("order_type", "订单金额"),
                )
            )
        return output


class FailingOrderClient:
    account = "bad"

    def fetch_orders(self, begin_date, end_date, account=""):
        raise RuntimeError("unauthorized")


class OrderOrchestrationTests(unittest.TestCase):
    def test_order_record_outputs_only_business_columns_in_required_order(self):
        row = models.OrderRecord(
            date="2026-04-18",
            title="A",
            platform="Mobo",
            language="英语",
            theater="MoboReels",
            order_count=2,
            amount=10.5,
            account="m1",
            task_id="task-1",
            order_type="订单金额",
        )

        self.assertEqual(
            list(row.to_dict().keys()),
            ["语言", "剧场", "剧名", "订单数", "金额", "平台", "类型"],
        )
        self.assertEqual(row.to_dict()["类型"], "订单金额")

    def test_collect_order_inputs_fetches_all_accounts_and_aggregates_summary(self):
        orders = importlib.import_module("drama_pipeline.7_yesterday_orders")
        mobo_clients = [
            FakeOrderClient("Mobo", "m1", [{"title": "A", "language": "英语", "theater": "MoboReels", "order_count": 1, "amount": 2.0}]),
            FakeOrderClient("Mobo", "m2", [{"title": "A", "language": "英语", "theater": "MoboReels", "order_count": 2, "amount": 3.0}]),
        ]
        beidou_clients = [
            FakeOrderClient("北斗", "b1", [{"title": "B", "language": "德语", "theater": "DramaBox", "order_count": 3, "amount": 4.0}])
        ]

        bundle = orders.collect_order_inputs(mobo_clients, beidou_clients, "2026-04-18", "2026-04-18")

        self.assertEqual(len(bundle["mobo_orders"]), 2)
        self.assertEqual(len(bundle["beidou_orders"]), 1)
        self.assertEqual(len(bundle["summary"]), 2)
        self.assertEqual(bundle["summary"][0].order_count, 3)
        self.assertEqual(mobo_clients[0].calls, [("2026-04-18", "2026-04-18", "m1")])

    def test_collect_order_inputs_records_account_errors_and_continues(self):
        orders = importlib.import_module("drama_pipeline.7_yesterday_orders")
        mobo_clients = [
            FailingOrderClient(),
            FakeOrderClient("Mobo", "m1", [{"title": "A", "language": "英语", "theater": "MoboReels", "order_count": 1, "amount": 2.0}]),
        ]

        bundle = orders.collect_order_inputs(mobo_clients, [], "2026-04-18", "2026-04-18")

        self.assertEqual(len(bundle["mobo_orders"]), 1)
        self.assertEqual(len(bundle["errors"]), 1)
        self.assertEqual(bundle["errors"][0]["平台"], "Mobo")
        self.assertEqual(bundle["errors"][0]["账号"], "bad")

    def test_collect_order_inputs_logs_account_fetch_stats(self):
        orders = importlib.import_module("drama_pipeline.7_yesterday_orders")
        logger = FakeLogger()
        mobo_clients = [
            FakeOrderClient(
                "Mobo",
                "m1",
                [{"title": "A", "language": "英语", "theater": "MoboReels", "order_count": 2, "amount": 3.0}],
                stats={"raw_count": 3, "parsed_count": 1, "total_count": 3, "amount_sum": 3.0},
            )
        ]

        orders.collect_order_inputs(mobo_clients, [], "2026-04-18", "2026-04-18", logger=logger)

        self.assertTrue(
            any(
                "raw_count=3" in message
                and "parsed_count=1" in message
                and "page_total=3" in message
                and "amount_sum=3.0" in message
                for message in logger.infos
            )
        )

    def test_summarize_order_errors_groups_accounts_by_platform(self):
        orders = importlib.import_module("drama_pipeline.7_yesterday_orders")
        messages = orders.summarize_order_errors(
            [
                {"平台": "Mobo", "账号": "m1"},
                {"平台": "Mobo", "账号": "m2"},
                {"平台": "Beidou", "账号": "b1"},
            ]
        )

        self.assertEqual(messages[0], "账号错误汇总: 平台=Mobo count=2 accounts=m1,m2")
        self.assertEqual(messages[1], "账号错误汇总: 平台=Beidou count=1 accounts=b1")

    def test_build_order_category_summary_groups_platform_and_type_with_total(self):
        orders = importlib.import_module("drama_pipeline.7_yesterday_orders")
        rows = [
            models.OrderRecord("2026-04-18", "A", "Mobo", "英语", "MoboReels", 2, 10.0, order_type="订单金额"),
            models.OrderRecord("2026-04-18", "B", "Mobo", "英语", "MoboReels", 1, 3.5, order_type="广告金额"),
            models.OrderRecord("2026-04-18", "C", "北斗", "德语", "DramaBox", 1, 8.0, order_type="订单金额"),
        ]

        summary = orders.build_order_category_summary(rows)

        self.assertEqual(
            summary,
            [
                {"平台": "Mobo", "类型": "订单金额", "订单数": 2, "金额": 10.0},
                {"平台": "Mobo", "类型": "广告金额", "订单数": 1, "金额": 3.5},
                {"平台": "北斗", "类型": "订单金额", "订单数": 1, "金额": 8.0},
                {"平台": "全部", "类型": "全部", "订单数": 4, "金额": 21.5},
            ],
        )

    def test_aggregate_orders_keeps_large_ads_and_groups_small_ads_by_platform(self):
        orders = importlib.import_module("drama_pipeline.7_yesterday_orders")
        rows = [
            models.OrderRecord("2026-04-18", "Normal", "Mobo", "英语", "MoboReels", 2, 10.0, order_type="订单金额"),
            models.OrderRecord("2026-04-18", "Big Ad", "Mobo", "英语", "其它", 1, 12.0, order_type="广告金额"),
            models.OrderRecord("2026-04-18", "Small A", "Mobo", "英语", "MoboReels", 1, 3.0, order_type="广告金额"),
            models.OrderRecord("2026-04-18", "Small B", "Mobo", "德语", "DramaBox", 1, 4.0, order_type="广告金额"),
            models.OrderRecord("2026-04-18", "Small C", "北斗", "法语", "ShortMax", 1, 6.0, order_type="广告金额"),
        ]

        summary = [row.to_dict() for row in orders.aggregate_orders(rows)]

        self.assertEqual(
            summary,
            [
                {"语言": "英语", "剧场": "MoboReels", "剧名": "Normal", "订单数": 2, "金额": 10.0, "平台": "Mobo", "类型": "订单金额"},
                {"语言": "英语", "剧场": "其它", "剧名": "Big Ad", "订单数": 1, "金额": 12.0, "平台": "Mobo", "类型": "广告金额"},
                {"语言": "全部", "剧场": "全部", "剧名": "其它", "订单数": 2, "金额": 7.0, "平台": "Mobo", "类型": "广告金额"},
                {"语言": "全部", "剧场": "全部", "剧名": "其它", "订单数": 1, "金额": 6.0, "平台": "北斗", "类型": "广告金额"},
            ],
        )

    def test_export_order_stage_files_writes_single_merged_workbook(self):
        orders = importlib.import_module("drama_pipeline.7_yesterday_orders")
        bundle = {
            "mobo_orders": [models.OrderRecord("2026-04-18", "A", "Mobo", "英语", "MoboReels", 1, 2.0, "m1")],
            "beidou_orders": [models.OrderRecord("2026-04-18", "B", "北斗", "德语", "DramaBox", 1, 3.0, "b1")],
            "summary": [models.OrderRecord("2026-04-18", "A", "Mobo", "英语", "MoboReels", 1, 2.0, "m1")],
        }

        tmp = workspace_tempdir()
        try:
            paths = orders.export_order_stage_files(bundle, tmp, "2026-04-18")
            self.assertEqual(list(paths), ["summary"])
            self.assertEqual(paths["summary"].parent, tmp)
            self.assertEqual(paths["summary"].name, "yesterday_orders_summary_2026-04-18.xlsx")
            workbook = openpyxl.load_workbook(paths["summary"])
            try:
                headers = [workbook.active.cell(1, column).value for column in range(1, workbook.active.max_column + 1)]
                self.assertEqual(headers, ["语言", "剧场", "剧名", "订单数", "金额", "平台", "类型"])
                values = [workbook.active.cell(2, column).value for column in range(1, workbook.active.max_column + 1)]
                self.assertIn("A", values)
                self.assertEqual(workbook.sheetnames, ["yesterday_orders_summary"])
            finally:
                workbook.close()
        finally:
            shutil.rmtree(tmp, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
