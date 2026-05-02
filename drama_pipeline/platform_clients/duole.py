from __future__ import annotations

import json
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import importlib
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

try:
    from selenium import webdriver
    from selenium.common.exceptions import TimeoutException
    from selenium.webdriver.edge.options import Options as EdgeOptions
    from selenium.webdriver.support.ui import WebDriverWait
except Exception:
    webdriver = None
    TimeoutException = None
    EdgeOptions = None
    WebDriverWait = None

try:
    from playwright.sync_api import sync_playwright
except Exception:
    sync_playwright = None

config = importlib.import_module("drama_pipeline.2_config")
models = importlib.import_module("drama_pipeline.3_models")
runtime = importlib.import_module("drama_pipeline.10_runtime")

from .common import (
    FetchStatsMixin,
    _log_info,
    _log_warning,
    _safe_int,
    normalize_theater_name,
)

class DuoleClient(FetchStatsMixin):
    def __init__(
        self,
        cookie: str = config.DUOLE_COOKIE,
        local_workbook: Optional[Path | str] = None,
        logger: Any = None,
    ):
        super().__init__()
        self.cookie = cookie
        self.share_url = config.DUOLE_SHARE_URL
        self.target_sheets = list(config.DUOLE_TARGET_SHEETS)
        self.sheet_limits = dict(config.DUOLE_SHEET_LIMITS)
        self.local_workbook_provided = local_workbook is not None
        self.local_workbook = Path(local_workbook) if local_workbook else None
        self.logger = logger

    def fetch_recommend_dramas(self) -> List[models.DramaRecord]:
        if self.local_workbook_provided:
            return self.fetch_local_recommend_dramas()

        if config.DUOLE_USE_LOCAL_FIRST:
            workbook_path = self._find_local_workbook()
            if workbook_path is not None:
                local_rows = self.fetch_local_recommend_dramas()
                if local_rows:
                    return local_rows

        web_rows = self.fetch_web_sheet_rows()
        if web_rows:
            parsed = self.parse_web_sheet_rows(web_rows)
            if parsed:
                self._record_duole_web_stats(web_rows, parsed)
                return parsed

        return []

    def fetch_web_sheet_rows(self) -> Dict[str, List[Dict[str, Any]]]:
        if webdriver is None or EdgeOptions is None or WebDriverWait is None:
            return {}
        if not self.cookie:
            _log_warning(self.logger, "[Duole] cookie missing, start local login refresh")
            refreshed = _refresh_duole_cookie_via_playwright(self.share_url, logger=self.logger)
            if not refreshed:
                return {}
            self.cookie = refreshed
            _save_duole_cookie_to_config(refreshed)
            _log_info(self.logger, "[Duole] cookie refreshed and saved, fetching again")
        try:
            result = _fetch_duole_records_from_web(self.share_url, self.cookie, self.target_sheets, logger=self.logger)
            if _has_any_duole_sheet_rows(result, self.target_sheets):
                return result
            _log_warning(self.logger, "[Duole] all target sheets are empty, cookie may be expired")
            raise RuntimeError("duole web fetch returned empty data for all sheets")
        except Exception as exc:
            _log_warning(self.logger, f"[Duole] web read failed: {type(exc).__name__}: {exc}")
            refreshed = _refresh_duole_cookie_via_playwright(self.share_url, logger=self.logger)
            if refreshed:
                self.cookie = refreshed
                _save_duole_cookie_to_config(refreshed)
                _log_info(self.logger, "[Duole] cookie refreshed and saved, retrying web fetch")
                try:
                    result = _fetch_duole_records_from_web(self.share_url, self.cookie, self.target_sheets, logger=self.logger)
                    if _has_any_duole_sheet_rows(result, self.target_sheets):
                        return result
                    _log_warning(self.logger, "[Duole] all target sheets are still empty after refresh")
                except Exception as retry_exc:
                    _log_warning(self.logger, f"[Duole] retry after refresh failed: {type(retry_exc).__name__}: {retry_exc}")
            return {}

    def parse_web_sheet_rows(self, sheet_rows: Dict[str, List[Dict[str, Any]]]) -> List[models.DramaRecord]:
        output: List[models.DramaRecord] = []
        for sheet_name in self.target_sheets:
            rank = 0
            for row in sheet_rows.get(sheet_name) or []:
                record = self._build_record_from_row(sheet_name, row, rank + 1)
                if record is None:
                    continue
                rank += 1
                output.append(record)
        return output

    def fetch_local_recommend_dramas(self) -> List[models.DramaRecord]:
        return self._fetch_local_recommend_dramas_for_sheets(self.target_sheets)

    def _fetch_local_recommend_dramas_for_sheets(self, sheet_names: List[str]) -> List[models.DramaRecord]:
        workbook_path = self._find_local_workbook()
        if workbook_path is None or not workbook_path.exists():
            return []
        workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            rows: List[models.DramaRecord] = []
            for sheet_name in sheet_names:
                if sheet_name not in workbook.sheetnames:
                    continue
                rows.extend(self._parse_sheet(workbook[sheet_name], sheet_name))
            return rows
        finally:
            workbook.close()

    def _merge_with_local_fallback(
        self,
        web_rows: Dict[str, List[Dict[str, Any]]],
        parsed: List[models.DramaRecord],
    ) -> List[models.DramaRecord]:
        if not parsed:
            return parsed
        missing_sheets = [
            sheet_name for sheet_name in self.target_sheets if not list(web_rows.get(sheet_name) or [])
        ]
        if not missing_sheets:
            return parsed
        local_rows = self._fetch_local_recommend_dramas_for_sheets(missing_sheets)
        if not local_rows:
            return parsed
        _log_warning(self.logger, f"[多乐] Web 缺失 sheet {','.join(missing_sheets)}，已用本地文件补齐 {len(local_rows)} 条")
        merged = list(parsed)
        merged.extend(local_rows)
        return merged

    def _parse_sheet(self, worksheet, sheet_name: str) -> List[models.DramaRecord]:
        header_row = self._resolve_header_row(worksheet, sheet_name)
        layout = config.DUOLE_SHEET_LAYOUTS.get(sheet_name, {})
        max_data_rows = int(layout.get("max_data_rows") or self.sheet_limits.get(sheet_name, max(worksheet.max_row - header_row, 0)))
        max_rows = min(worksheet.max_row, header_row + max_data_rows)
        if max_rows < header_row:
            return []

        header_values = next(
            worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True),
            (),
        )
        headers = _make_unique_headers(list(header_values))
        output: List[models.DramaRecord] = []
        raw_count = 0
        for values in worksheet.iter_rows(min_row=header_row + 1, max_row=max_rows, values_only=True):
            row = _build_duole_raw_row(headers, list(values))
            row = _normalize_duole_sheet_row(sheet_name, row, list(values))
            if not any(str(value).strip() for value in row.values()):
                continue
            raw_count += 1
            record = self._build_record_from_row(sheet_name, row, len(output) + 1)
            if record is not None:
                output.append(record)
        self._record_fetch_total("duole_recommend", "原始返回", raw_count, sheet_name=sheet_name)
        if layout.get("note_filter_required"):
            self._record_fetch_total("duole_recommend", "备注命中", len(output), sheet_name=sheet_name)
        self._record_parsed_rows("duole_recommend", "解析成功", output)
        return output

    def _build_record_from_row(
        self,
        sheet_name: str,
        row: Dict[str, Any],
        rank: int,
    ) -> Optional[models.DramaRecord]:
        layout = config.DUOLE_SHEET_LAYOUTS.get(sheet_name, {})
        title = self._pick_first(
            self._layout_value(row, str(layout.get("title_column") or "")),
            row.get("title_raw"),
            row.get("剧名"),
            row.get("外语名"),
            row.get("中文名"),
            row.get("title"),
            row.get("name"),
        )
        if not title or str(title).isdigit():
            title = _choose_duole_title(row)
        language = self._pick_first(
            layout.get("language_name"),
            self._layout_value(row, str(layout.get("language_column") or "")),
            row.get("language_name"),
            row.get("语言"),
            row.get("语种"),
        )
        language = runtime.normalize_language_name(language)
        if not language:
            language = _infer_duole_language(sheet_name, row)
        theater = self._pick_first(
            layout.get("theater_name"),
            self._layout_value(row, str(layout.get("theater_column") or "")),
            row.get("theater_name"),
            row.get("剧场"),
            row.get("theater"),
        )
        if not theater:
            theater = _infer_duole_theater(sheet_name, row)
        tags = [value for value in [self._pick_first(row.get("类型"), row.get("受众"))] if value]
        title_text = "" if title is None else str(title).strip()
        language_text = runtime.normalize_language_name(language)
        theater_text = normalize_theater_name(theater)
        if not title_text or not language_text:
            return None
        note_text = self._pick_first(
            self._layout_value(row, str(layout.get("note_column") or "")),
            row.get("备注"),
            row.get("note"),
        )
        if layout.get("note_filter_required") and not _duole_note_matches(note_text):
            return None
        publish_at = self._pick_first(
            self._layout_value(row, str(layout.get("publish_column") or "")),
            row.get("上架时间"),
            row.get("publish_at"),
            row.get("publishTime"),
        )
        recommend_date = self._pick_first(
            self._layout_value(row, str(layout.get("recommend_column") or "")),
            row.get("recommend_date"),
            row.get("日期"),
            row.get("更新日期"),
            row.get("更新时间"),
        )
        raw = dict(row)
        raw.setdefault("sheet_name", sheet_name)
        if recommend_date:
            raw.setdefault("recommend_date", recommend_date)
        if note_text:
            raw.setdefault("备注", note_text)
        return models.DramaRecord(
            title=title_text,
            language=language_text,
            theater=theater_text,
            source="duole_recommend",
            rank=rank,
            publish_at=str(publish_at or "").strip(),
            task_id=str(row.get("剧目ID") or row.get("record_key") or ""),
            tags=[str(tag).strip() for tag in tags if str(tag).strip()],
            raw=raw,
        )

    def _resolve_header_row(self, worksheet, sheet_name: str) -> int:
        layout = config.DUOLE_SHEET_LAYOUTS.get(sheet_name, {})
        expected = {str(header).strip() for header in layout.get("headers") or [] if str(header).strip()}
        configured = int(layout.get("header_row") or 1)
        scan_rows: List[int] = []
        if configured:
            scan_rows.append(configured)
        scan_rows.extend(range(1, min(worksheet.max_row, 6) + 1))
        seen = set()
        for row_number in scan_rows:
            if row_number in seen or row_number > worksheet.max_row:
                continue
            seen.add(row_number)
            values = next(
                worksheet.iter_rows(min_row=row_number, max_row=row_number, values_only=True),
                (),
            )
            texts = {str(value).strip() for value in values if value is not None}
            if expected and texts.intersection(expected):
                return row_number
        return configured or 1

    def _find_local_workbook(self) -> Optional[Path]:
        candidates: List[Path] = []
        if self.local_workbook is not None:
            candidates.append(Path(self.local_workbook))
        for directory in (config.PIPELINE_DIR.parent / "platform_catalog_output", config.PIPELINE_DIR.parent):
            candidates.extend(directory / name for name in config.DUOLE_LOCAL_CANDIDATES)
        seen = set()
        for candidate in candidates:
            resolved = str(candidate)
            if resolved in seen:
                continue
            seen.add(resolved)
            if candidate.exists() and candidate.is_file():
                return candidate
        return None

    @staticmethod
    def _layout_value(row: Dict[str, Any], column: str) -> Any:
        if not column:
            return ""
        return row.get(column, "")

    @staticmethod
    def _pick_first(*values: Any) -> str:
        for value in values:
            text = "" if value is None else str(value).strip()
            if text:
                return text
        return ""

    def _record_duole_web_stats(self, sheet_rows: Dict[str, List[Dict[str, Any]]], parsed: List[models.DramaRecord]) -> None:
        parsed_by_sheet: Dict[str, List[models.DramaRecord]] = {}
        for record in parsed:
            sheet_name = str((record.raw or {}).get("sheet_name") or "")
            parsed_by_sheet.setdefault(sheet_name, []).append(record)
        for sheet_name in self.target_sheets:
            rows = list(sheet_rows.get(sheet_name) or [])
            self._record_fetch_total("duole_recommend", "原始返回", len(rows), sheet_name=sheet_name)
            layout = config.DUOLE_SHEET_LAYOUTS.get(sheet_name, {})
            if layout.get("note_filter_required"):
                self._record_fetch_total(
                    "duole_recommend",
                    "备注命中",
                    len(parsed_by_sheet.get(sheet_name, [])),
                    sheet_name=sheet_name,
                )
        self._record_parsed_rows("duole_recommend", "解析成功", parsed)


def _refresh_duole_cookie_via_playwright(share_url: str, logger: Any = None) -> str:
    """Open a local browser for manual Duole login and capture fresh auth cookies."""
    if sync_playwright is None:
        _log_warning(logger, "[Duole] Playwright unavailable, cannot refresh cookie")
        return ""
    _log_info(logger, "[Duole] cookie may be expired, opening browser for manual login")
    from playwright.sync_api import sync_playwright as _sp

    with _sp() as p:
        browser = p.chromium.launch(headless=False)
        try:
            context = browser.new_context(viewport={"width": 1280, "height": 900})
            page = context.new_page()
            page.goto(share_url, wait_until="domcontentloaded", timeout=30000)

            current_url = page.url
            if _is_duole_login_url(current_url):
                _log_info(logger, "[Duole] login page detected, please scan or sign in within 3 minutes")
                for _ in range(180):
                    time.sleep(1)
                    current_url = page.url
                    if not _is_duole_login_url(current_url):
                        break
                else:
                    _log_warning(logger, "[Duole] login wait timed out")
                    return ""

            try:
                page.wait_for_load_state("domcontentloaded", timeout=10000)
            except Exception:
                pass
            time.sleep(2)

            cookie_text = _serialize_duole_cookies(context.cookies())
            if not cookie_text:
                _log_warning(logger, "[Duole] no usable auth cookies captured")
                return ""
            _log_info(logger, "[Duole] refreshed auth cookies captured")
            return cookie_text
        finally:
            browser.close()

def _save_duole_cookie_to_config(cookie_value: str) -> None:
    """灏嗘柊 cookie 鍐欏叆 drama_pipeline_config.xlsx 鐨?settings sheet銆?"""
    try:
        runtime.update_duole_cookie_in_config(cookie_value)
    except Exception as exc:
        print(f"[澶氫箰] 淇濆瓨 cookie 鍒伴厤缃枃浠跺け璐? {type(exc).__name__}: {exc}")


def _has_any_duole_sheet_rows(result: Dict[str, List[Dict[str, Any]]], target_sheets: List[str]) -> bool:
    if not result:
        return False
    return any(result.get(sheet_name) for sheet_name in target_sheets)


def _is_duole_login_url(url: str) -> bool:
    text = str(url or "").lower()
    return "account.kdocs.cn" in text or "account.wps.cn" in text or "passport" in text


def _serialize_duole_cookies(cookies: List[Dict[str, Any]]) -> str:
    filtered = []
    for cookie in cookies or []:
        domain = str(cookie.get("domain") or "").lower()
        if "kdocs.cn" not in domain and "wps.cn" not in domain:
            continue
        filtered.append(
            {
                "name": str(cookie.get("name") or ""),
                "value": str(cookie.get("value") or ""),
                "domain": str(cookie.get("domain") or ""),
                "path": str(cookie.get("path") or "/"),
                "secure": bool(cookie.get("secure", False)),
                "httpOnly": bool(cookie.get("httpOnly", False)),
                "sameSite": cookie.get("sameSite") or "Lax",
                "expires": cookie.get("expires", -1),
            }
        )
    if not filtered:
        return ""
    return json.dumps(filtered, ensure_ascii=False, separators=(",", ":"))


def _fetch_duole_records_from_web(

    share_url: str,
    cookie: str,
    target_sheets: List[str],
    logger: Any = None,
) -> Dict[str, List[Dict[str, Any]]]:
    driver = _create_duole_edge_driver()
    try:
        _inject_duole_cookie(driver, cookie)
        try:
            driver.get(f"{share_url}?from=from_copylink")
        except Exception as exc:
            if TimeoutException is None or not isinstance(exc, TimeoutException):
                raise
        if not bool(getattr(config, "DUOLE_WEB_USE_BATCH", False)):
            _log_info(logger, "[多乐] 使用逐 Sheet 抓取模式")
            _wait_for_duole_api(driver)
            return _fetch_duole_sheet_rows_one_by_one(driver, target_sheets, logger=logger)
        retries = max(int(getattr(config, "DUOLE_WEB_BATCH_RETRIES", 1) or 1), 1)
        for attempt in range(1, retries + 1):
            try:
                batch_rows = _fetch_duole_sheet_rows_batch(driver, target_sheets)
                if batch_rows:
                    return batch_rows
                _log_warning(logger, f"[多乐] 批量抓取返回空结果，第 {attempt}/{retries} 次")
            except Exception as exc:
                _log_warning(logger, f"[多乐] 批量抓取失败，第 {attempt}/{retries} 次: {type(exc).__name__}: {exc}")
            if attempt < retries:
                _wait_for_duole_api(driver)
        per_sheet_rows = _fetch_duole_sheet_rows_one_by_one(driver, target_sheets, logger=logger)
        if per_sheet_rows and any(per_sheet_rows.get(sheet_name) for sheet_name in target_sheets):
            return per_sheet_rows
        _log_warning(logger, "[多乐] 批量抓取重试耗尽，回退逐 sheet 矩阵抓取")
        _wait_for_duole_api(driver)
        sheet_index_rows = _fetch_duole_sheet_index_web(driver)
        by_name = {row.get("name"): row for row in sheet_index_rows}
        output: Dict[str, List[Dict[str, Any]]] = {}
        for sheet_name in target_sheets:
            meta = by_name.get(sheet_name)
            if not meta:
                output[sheet_name] = []
                continue
            try:
                matrix = _fetch_duole_sheet_matrix_web(driver, meta, sheet_name)
                output[sheet_name] = _parse_duole_sheet_matrix(sheet_name, meta, matrix)
            except Exception as exc:
                _log_warning(logger, f"[多乐] 逐 sheet 矩阵抓取失败 {sheet_name}: {type(exc).__name__}: {exc}")
                output[sheet_name] = []
        return output
    finally:
        driver.quit()


def _create_duole_edge_driver():
    if webdriver is None or EdgeOptions is None:
        raise RuntimeError("selenium 不可用")
    options = EdgeOptions()
    options.page_load_strategy = "eager"
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-background-networking")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--no-first-run")
    options.add_argument("--remote-debugging-port=0")
    options.add_argument("--window-size=1800,1200")
    options.add_argument("--log-level=3")
    driver = webdriver.Edge(options=options)
    driver.set_page_load_timeout(config.DUOLE_WEB_WAIT_SECONDS)
    driver.set_script_timeout(config.DUOLE_WEB_WAIT_SECONDS)
    return driver


def _parse_cookie_pairs(cookie_text: str) -> List[Tuple[str, str]]:
    pairs: List[Tuple[str, str]] = []
    for part in str(cookie_text or "").split("; "):
        if "=" not in part:
            continue
        name, value = part.split("=", 1)
        name = name.strip()
        if name:
            pairs.append((name, value))
    return pairs


def _parse_duole_cookie_payload(cookie_text: str) -> List[Dict[str, Any]]:
    text = str(cookie_text or "").strip()
    if not text:
        return []
    try:
        payload = json.loads(text)
    except Exception:
        payload = None
    if isinstance(payload, list):
        cookies: List[Dict[str, Any]] = []
        for item in payload:
            if not isinstance(item, dict):
                continue
            name = str(item.get("name") or "").strip()
            value = str(item.get("value") or "")
            if not name:
                continue
            cookie = {
                "name": name,
                "value": value,
                "domain": str(item.get("domain") or ""),
                "path": str(item.get("path") or "/"),
            }
            if item.get("secure") is not None:
                cookie["secure"] = bool(item.get("secure"))
            if item.get("httpOnly") is not None:
                cookie["httpOnly"] = bool(item.get("httpOnly"))
            if item.get("sameSite") is not None:
                cookie["sameSite"] = item.get("sameSite")
            if item.get("expires") is not None:
                cookie["expires"] = item.get("expires")
            cookies.append(cookie)
        if cookies:
            return cookies
    cookies = []
    for name, value in _parse_cookie_pairs(text):
        cookies.append(
            {
                "name": name,
                "value": value,
                "domain": ".kdocs.cn",
                "path": "/",
                "url": "https://www.kdocs.cn/",
            }
        )
    return cookies


def _inject_duole_cookie(driver, cookie_text: str) -> None:
    cookies = _parse_duole_cookie_payload(cookie_text)
    if not cookies:
        return
    driver.execute_cdp_cmd("Network.enable", {})
    for cookie in cookies:
        params = {
            "name": cookie["name"],
            "value": cookie["value"],
            "path": cookie.get("path") or "/",
        }
        domain = str(cookie.get("domain") or "").strip()
        url = str(cookie.get("url") or "").strip()
        if domain:
            params["domain"] = domain
        elif url:
            params["url"] = url
        else:
            params["url"] = "https://www.kdocs.cn/"
        for key in ("secure", "httpOnly", "sameSite", "expires"):
            if key in cookie and cookie.get(key) is not None:
                params[key] = cookie.get(key)
        driver.execute_cdp_cmd("Network.setCookie", params)


def _build_duole_web_sheet_configs(target_sheets: List[str]) -> List[Dict[str, Any]]:
    configs: List[Dict[str, Any]] = []
    for sheet_name in target_sheets:
        layout = config.DUOLE_SHEET_LAYOUTS.get(sheet_name, {})
        header_row = max(_safe_int(layout.get("header_row")), 1)
        max_data_rows = max(_safe_int(layout.get("max_data_rows")), 0)
        used_cols = max(_safe_int(layout.get("used_cols")), len(layout.get("headers") or []))
        configs.append(
            {
                "name": sheet_name,
                "used_cols": used_cols,
                "limit_rows": header_row + max_data_rows if max_data_rows > 0 else 0,
            }
        )
    return configs


def _fetch_duole_sheet_rows_batch(driver, target_sheets: List[str]) -> Dict[str, List[Dict[str, Any]]]:
    payload = driver.execute_async_script(
        """
const [sheetConfigs, timeoutMs, done] = arguments;
(async () => {
  try {
    const sleep = (ms) => new Promise((resolve) => setTimeout(resolve, ms));
    const toColumn = (n) => {
      let text = '';
      while (n > 0) {
        const mod = (n - 1) % 26;
        text = String.fromCharCode(65 + mod) + text;
        n = Math.floor((n - 1) / 26);
      }
      return text;
    };

    const start = Date.now();
    let sheets = null;
    while (Date.now() - start < timeoutMs) {
      const api = window.WPSOpenApi;
      if (api && api.Application && api.Application.Sheets) {
        sheets = api.Application.Sheets;
        break;
      }
      await sleep(200);
    }

    if (!sheets) {
      done({ ok: false, error: 'timeout waiting for WPSOpenApi.Application.Sheets' });
      return;
    }

    const count = await Promise.resolve(sheets.Count);
    const byName = {};
    for (let index = 1; index <= count; index += 1) {
      const sheet = sheets.Item(index);
      const name = await Promise.resolve(sheet.Name);
      byName[name] = {
        index,
        sid: await Promise.resolve(sheet.SId),
        type: await Promise.resolve(sheet.Type),
        visible: await Promise.resolve(sheet.Visible),
        used_rows: await Promise.resolve(sheet.UsedRange.Rows.Count),
        used_cols: await Promise.resolve(sheet.UsedRange.Columns.Count),
      };
    }

    const results = [];
    for (const cfg of sheetConfigs) {
      const meta = byName[cfg.name];
      if (!meta) {
        results.push({ name: cfg.name, meta: null, values: [] });
        continue;
      }
      const usedCols = cfg.used_cols && cfg.used_cols > 0
        ? Math.min(meta.used_cols || cfg.used_cols, cfg.used_cols)
        : (meta.used_cols || 0);
      const usedRows = cfg.limit_rows && cfg.limit_rows > 0
        ? Math.min(meta.used_rows || cfg.limit_rows, cfg.limit_rows)
        : (meta.used_rows || 0);
      if (!usedCols || !usedRows) {
        results.push({ name: cfg.name, meta, values: [] });
        continue;
      }
      const range = `A1:${toColumn(usedCols)}${usedRows}`;
      const sheet = sheets.Item(meta.index);
      const values = await Promise.resolve(sheet.Range(range).Value);
      results.push({ name: cfg.name, meta, range, values });
    }

    done({ ok: true, results });
  } catch (err) {
    done({
      ok: false,
      error: String(err),
      stack: err && err.stack ? String(err.stack).slice(0, 1200) : '',
    });
  }
})();
""",
        _build_duole_web_sheet_configs(target_sheets),
        int(config.DUOLE_WEB_WAIT_SECONDS * 1000),
    )
    if not payload.get("ok"):
        raise RuntimeError(payload.get("error") or "duole batch fetch failed")

    output: Dict[str, List[Dict[str, Any]]] = {}
    for item in payload.get("results") or []:
        sheet_name = str(item.get("name") or "")
        meta = item.get("meta")
        values = item.get("values") or []
        if not sheet_name:
            continue
        if not meta or not values:
            output[sheet_name] = []
            continue
        output[sheet_name] = _parse_duole_sheet_matrix(sheet_name, meta, values)
    return output


def _fetch_duole_sheet_rows_one_by_one(driver, target_sheets: List[str], logger: Any = None) -> Dict[str, List[Dict[str, Any]]]:
    _log_info(logger, f"[多乐] 开始逐 Sheet 抓取，共 {len(target_sheets)} 个 Sheet")
    output: Dict[str, List[Dict[str, Any]]] = {}
    sheet_index_rows = _fetch_duole_sheet_index_web(driver)
    by_name = {row.get("name"): row for row in sheet_index_rows}
    retries = max(int(getattr(config, "DUOLE_WEB_SHEET_RETRIES", getattr(config, "DUOLE_WEB_BATCH_RETRIES", 1)) or 1), 1)
    for sheet_name in target_sheets:
        meta = by_name.get(sheet_name)
        if not meta:
            _log_warning(logger, f"[多乐] Sheet 不存在: {sheet_name}")
            output[sheet_name] = []
            continue
        started = datetime.now()
        output[sheet_name] = []
        for attempt in range(1, retries + 1):
            try:
                matrix = _fetch_duole_sheet_matrix_web(driver, meta, sheet_name)
                rows = _parse_duole_sheet_matrix(sheet_name, meta, matrix)
                output[sheet_name] = rows
                elapsed = (datetime.now() - started).total_seconds()
                _log_info(
                    logger,
                    f"[多乐] Sheet 完成 {sheet_name}: used_rows={meta.get('used_rows')} used_cols={meta.get('used_cols')} parsed_rows={len(rows)} attempt={attempt}/{retries} elapsed={elapsed:.1f}s",
                )
                break
            except Exception as exc:
                _log_warning(logger, f"[多乐] Sheet 抓取失败 {sheet_name}: attempt={attempt}/{retries} {type(exc).__name__}: {exc}")
                if attempt < retries:
                    _wait_for_duole_api(driver)
    return output


def _wait_for_duole_api(driver) -> None:
    if WebDriverWait is None:
        raise RuntimeError("selenium WebDriverWait 不可用")
    WebDriverWait(driver, config.DUOLE_WEB_WAIT_SECONDS).until(
        lambda current: current.execute_script(
            "return !!(window.WPSOpenApi && window.WPSOpenApi.documentReadyPromise && window.WPSOpenApi.Application && window.WPSOpenApi.Application.Sheets);"
        )
    )
    ok = driver.execute_async_script(
        """
const done = arguments[0];
(async () => {
  try {
    await window.WPSOpenApi.documentReadyPromise;
    const start = Date.now();
    while (Date.now() - start < 5000) {
      if (window.WPSOpenApi && window.WPSOpenApi.Application && window.WPSOpenApi.Application.Sheets) {
        done(true);
        return;
      }
      await new Promise((resolve) => setTimeout(resolve, 200));
    }
    done(false);
  } catch (err) {
    done(false);
  }
})();
"""
    )
    if not ok:
        raise RuntimeError("WPSOpenApi.documentReadyPromise 未完成")


def _fetch_duole_sheet_index_web(driver) -> List[Dict[str, Any]]:
    payload = driver.execute_async_script(
        """
const done = arguments[0];
(async () => {
  try {
    const api = window.WPSOpenApi;
    await api.documentReadyPromise;
    const sheets = api.Application.Sheets;
    const count = await Promise.resolve(sheets.Count);
    const out = [];
    for (let index = 1; index <= count; index += 1) {
      const sheet = sheets.Item(index);
      out.push({
        index,
        name: await Promise.resolve(sheet.Name),
        sid: await Promise.resolve(sheet.SId),
        type: await Promise.resolve(sheet.Type),
        visible: await Promise.resolve(sheet.Visible),
        used_rows: await Promise.resolve(sheet.UsedRange.Rows.Count),
        used_cols: await Promise.resolve(sheet.UsedRange.Columns.Count),
      });
    }
    done({ ok: true, sheets: out });
  } catch (err) {
    done({ ok: false, error: String(err) });
  }
})();
"""
    )
    if not payload.get("ok"):
        raise RuntimeError(payload.get("error") or "多乐 sheet 索引抓取失败")
    return payload.get("sheets") or []


def _fetch_duole_range_values_web(driver, sheet_index: int, range_a1: str) -> List[List[Any]]:
    payload = driver.execute_async_script(
        """
const [sheetIndex, rangeA1, done] = arguments;
(async () => {
  try {
    const api = window.WPSOpenApi;
    await api.documentReadyPromise;
    const sheet = api.Application.Sheets.Item(sheetIndex);
    const values = await Promise.resolve(sheet.Range(rangeA1).Value);
    done({ ok: true, values });
  } catch (err) {
    done({ ok: false, error: String(err) });
  }
})();
""",
        sheet_index,
        range_a1,
    )
    if not payload.get("ok"):
        raise RuntimeError(payload.get("error") or f"多乐范围读取失败: {range_a1}")
    return payload.get("values") or []


def _get_duole_fetch_row_limit(sheet_name: str, used_rows: int) -> int:
    layout = config.DUOLE_SHEET_LAYOUTS.get(sheet_name, {})
    header_row = max(_safe_int(layout.get("header_row")), 1)
    max_data_rows = max(_safe_int(layout.get("max_data_rows")), 0)
    if max_data_rows <= 0:
        return used_rows
    return min(used_rows, header_row + max_data_rows)


def _fetch_duole_sheet_matrix_web(driver, sheet_meta: Dict[str, Any], sheet_name: str) -> List[List[Any]]:
    used_rows = max(_safe_int(sheet_meta.get("used_rows")), 0)
    used_cols = max(_safe_int(sheet_meta.get("used_cols")), 0)
    layout = config.DUOLE_SHEET_LAYOUTS.get(sheet_name, {})
    layout_used_cols = max(_safe_int(layout.get("used_cols")), len(layout.get("headers") or []))
    if layout_used_cols:
        used_cols = min(used_cols, layout_used_cols) if used_cols else layout_used_cols
    limited_rows = _get_duole_fetch_row_limit(sheet_name, used_rows)
    if limited_rows <= 0 or used_cols <= 0:
        return []
    end_col = get_column_letter(used_cols)
    if limited_rows <= config.DUOLE_WEB_CHUNK_SIZE:
        return _fetch_duole_range_values_web(driver, _safe_int(sheet_meta["index"]), f"A1:{end_col}{limited_rows}")
    values: List[List[Any]] = []
    for start_row in range(1, limited_rows + 1, config.DUOLE_WEB_CHUNK_SIZE):
        end_row = min(start_row + config.DUOLE_WEB_CHUNK_SIZE - 1, limited_rows)
        values.extend(
            _fetch_duole_range_values_web(
                driver,
                _safe_int(sheet_meta["index"]),
                f"A{start_row}:{end_col}{end_row}",
            )
        )
    return values


def _make_unique_headers(values: List[Any]) -> List[str]:
    headers: List[str] = []
    counts: Dict[str, int] = {}
    for index, value in enumerate(values, 1):
        text = str(value).strip() if value is not None else ""
        header = re.sub(r"\s+", " ", text) if text else f"col_{index}"
        counts[header] = counts.get(header, 0) + 1
        if counts[header] > 1:
            header = f"{header}_{counts[header]}"
        headers.append(header)
    return headers


def _build_duole_raw_row(headers: List[str], values: List[Any], source_positions: Optional[List[int]] = None) -> Dict[str, Any]:
    row: Dict[str, Any] = {}
    for index, header in enumerate(headers):
        source_index = source_positions[index] if source_positions and index < len(source_positions) else index
        row[header] = values[source_index] if source_index < len(values) and values[source_index] is not None else ""
    return row


def _format_datetime(value: datetime) -> str:
    if value.hour == 0 and value.minute == 0 and value.second == 0 and value.microsecond == 0:
        return value.strftime("%Y-%m-%d")
    return value.strftime("%Y-%m-%d %H:%M:%S")


def _parse_datetime(value: Any) -> Optional[datetime]:
    text = str(value).strip() if value is not None else ""
    if not text:
        return None

    normalized = (
        text.replace("T", " ")
        .replace("/", "-")
        .replace(".", "-")
        .replace("年", "-")
        .replace("月", "-")
        .replace("日", "")
    )
    normalized = re.sub(r"\s+", " ", normalized).strip()

    for fmt in (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S.%f",
        "%m-%d-%Y",
        "%m-%d-%Y %H:%M",
        "%m-%d-%Y %H:%M:%S",
    ):
        try:
            return datetime.strptime(normalized, fmt)
        except ValueError:
            continue
    return None


def _format_duole_date_value(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return _format_datetime(value)
    text = str(value).strip()
    if not text:
        return ""
    if re.fullmatch(r"\d{8}", text):
        try:
            return datetime.strptime(text, "%Y%m%d").strftime("%Y-%m-%d")
        except ValueError:
            return text
    parsed = _parse_datetime(text)
    if parsed is not None:
        return _format_datetime(parsed)
    try:
        numeric = float(text)
    except Exception:
        return text
    if not (20000 <= numeric <= 60000):
        return text
    try:
        parsed_excel = from_excel(numeric)
    except Exception:
        return text
    if isinstance(parsed_excel, datetime):
        return _format_datetime(parsed_excel)
    return str(parsed_excel)


def _normalize_duole_date_columns(raw: Dict[str, Any], date_columns: List[str]) -> None:
    for column in date_columns:
        if column in raw:
            raw[column] = _format_duole_date_value(raw.get(column))


def _looks_like_duole_id(text: Any) -> bool:
    value = str(text).strip()
    return value.isdigit() and len(value) >= 8


def _looks_like_duole_audience(text: Any) -> bool:
    value = str(text).strip()
    return any(token in value for token in ("男频", "女频", "海外", "LLM"))


def _looks_like_duole_link(text: Any) -> bool:
    value = str(text).strip().lower()
    return any(token in value for token in ("http", "下载", "小程序", "pan.baidu"))


def _looks_like_duole_language(text: Any) -> bool:
    return str(text).strip() in {
        "英语",
        "法语",
        "葡萄牙语",
        "德语",
        "繁体中文",
        "俄语",
        "意大利语",
        "阿拉伯语",
        "西班牙语",
        "日语",
        "韩语",
    }


def _normalize_duole_sheet_row(sheet_name: str, raw: Dict[str, Any], values: List[Any]) -> Dict[str, Any]:
    layout = config.DUOLE_SHEET_LAYOUTS.get(sheet_name, {})
    cleaned = [str(value).strip() if value is not None else "" for value in values]
    fixed_columns = layout.get("fixed_columns") or {}
    if fixed_columns:
        fixed = {name: _duole_fixed_value(cleaned, column) for name, column in fixed_columns.items()}
        if sheet_name == "12.DramaBox英语":
            raw.update(
                {
                    "更新日期": fixed.get("recommend", ""),
                    "上架时间": fixed.get("publish_at", ""),
                    "外语名": fixed.get("title", ""),
                    "备注": fixed.get("note", ""),
                }
            )
        elif sheet_name == "13.DramaBox小语种":
            raw.update(
                {
                    "更新时间": fixed.get("recommend", ""),
                    "上架时间": fixed.get("publish_at", ""),
                    "语种": fixed.get("language", ""),
                    "外语名": fixed.get("title", ""),
                    "备注": fixed.get("note", ""),
                }
            )
    elif sheet_name == "12.DramaBox英语":
        titles = [value for value in cleaned[2:5] if value and not _looks_like_duole_id(value)]
        drama_id = next((value for value in cleaned[2:5] if _looks_like_duole_id(value)), "")
        chinese_title = next((value for value in titles if re.search(r"[\u4e00-\u9fff]", value)), "")
        foreign_title = next((value for value in titles if value != chinese_title), "")
        if not foreign_title and titles:
            foreign_title = titles[0]
        raw.update(
            {
                "更新日期": cleaned[0] if len(cleaned) > 0 else "",
                "上架时间": cleaned[1] if len(cleaned) > 1 else "",
                "剧目ID": drama_id,
                "中文名": chinese_title,
                "外语名": foreign_title,
                "受众": cleaned[5] if len(cleaned) > 5 else "",
                "百度网盘链接": cleaned[6] if len(cleaned) > 6 else "",
                "故事简介": cleaned[7] if len(cleaned) > 7 else "",
                "备注": cleaned[8] if len(cleaned) > 8 else "",
            }
        )
    elif sheet_name == "13.DramaBox小语种":
        titles = [
            value for value in cleaned[2:6]
            if value and not _looks_like_duole_id(value) and not _looks_like_duole_language(value)
        ]
        drama_id = next((value for value in cleaned[2:6] if _looks_like_duole_id(value)), "")
        language_name = next((value for value in cleaned[2:5] if _looks_like_duole_language(value)), "")
        chinese_title = next((value for value in titles if re.search(r"[\u4e00-\u9fff]", value)), "")
        foreign_title = next((value for value in titles if value != chinese_title), "")
        if not foreign_title and titles:
            foreign_title = titles[0]
        audience = next((value for value in cleaned[5:8] if _looks_like_duole_audience(value)), "")
        link = next((value for value in cleaned[6:9] if _looks_like_duole_link(value)), "")
        story_candidates = [
            value for value in cleaned
            if value and value not in {language_name, drama_id, chinese_title, foreign_title, audience, link}
        ]
        raw.update(
            {
                "更新时间": cleaned[0] if len(cleaned) > 0 else "",
                "上架时间": cleaned[1] if len(cleaned) > 1 else "",
                "语种": language_name,
                "剧目ID": drama_id,
                "中文名": chinese_title,
                "外语名": foreign_title,
                "受众": audience,
                "百度网盘链接": link,
                "故事简介": max(story_candidates, key=len, default=""),
            }
        )
    _normalize_duole_date_columns(raw, list(layout.get("date_columns") or []))
    return raw


def _duole_fixed_value(values: Sequence[str], column: Any) -> str:
    if isinstance(column, int):
        index = column - 1
    else:
        text = str(column or "").strip().upper()
        if not text:
            return ""
        index = 0
        for char in text:
            if not ("A" <= char <= "Z"):
                return ""
            index = index * 26 + (ord(char) - ord("A") + 1)
        index -= 1
    if index < 0 or index >= len(values):
        return ""
    return values[index]


def _duole_note_matches(note_text: Any) -> bool:
    note = str(note_text or "").strip()
    if not note:
        return False
    lowered = note.lower()
    return any(str(keyword).strip().lower() in lowered for keyword in config.DUOLE_RECOMMEND_NOTE_KEYWORDS)


def _infer_duole_language(sheet_name: str, row: Dict[str, Any]) -> str:
    text = " ".join(str(value) for value in row.values())
    if "英语" in sheet_name or "英语" in text:
        return "英语"
    if "小语种" in sheet_name:
        for language_name in ("法语", "葡萄牙语", "德语", "繁体中文", "俄语", "意大利语"):
            if language_name in text:
                return language_name
    return ""


def _infer_duole_theater(sheet_name: str, row: Dict[str, Any]) -> str:
    combined = " ".join(str(value) for value in row.values())
    if "DramaBox" in sheet_name or "dramabox" in combined.lower():
        return "DramaBox"
    for value in config.THEATER_NAMES.values():
        if value.lower() in combined.lower():
            return value
    return ""


def _choose_duole_title(row: Dict[str, Any]) -> str:
    preferred_tokens = ("剧名", "中文名", "英文名", "外语名", "片名", "title", "name")
    excluded_tokens = ("剧场", "id", "编号", "日期", "时间", "素材", "链接", "理由", "受众", "类型", "备注")
    for token in preferred_tokens:
        for key, value in row.items():
            key_text = str(key).strip().lower()
            value_text = str(value).strip()
            if not value_text or token not in key_text:
                continue
            if any(excluded in key_text for excluded in excluded_tokens) or value_text.isdigit():
                continue
            return value_text
    for key, value in row.items():
        key_text = str(key).lower()
        value_text = str(value).strip()
        if not value_text:
            continue
        if any(token in key_text for token in ("剧", "title", "名称", "片名")):
            if any(excluded in key_text for excluded in excluded_tokens) or value_text.isdigit():
                continue
            return value_text
    for value in row.values():
        text = str(value).strip()
        if len(text) >= 4 and not text.isdigit():
            return text
    return ""


def _fill_down_duole_dates(parsed_rows: List[Dict[str, Any]]) -> None:
    last_date = ""
    for row in parsed_rows:
        current_date = str(row.get("日期", "")).strip()
        if current_date:
            last_date = current_date
        elif last_date:
            row["日期"] = last_date


def _parse_duole_sheet_matrix(sheet_name: str, sheet_meta: Dict[str, Any], matrix: List[List[Any]]) -> List[Dict[str, Any]]:
    if not matrix:
        return []
    layout = config.DUOLE_SHEET_LAYOUTS.get(sheet_name, {})
    header_row_number = _safe_int(layout.get("header_row")) or 1
    headers = list(layout.get("headers") or _make_unique_headers(list(matrix[header_row_number - 1])))
    source_positions = list(layout.get("source_positions") or [])
    parsed_rows: List[Dict[str, Any]] = []
    for row_number, values in enumerate(matrix[header_row_number:], header_row_number + 1):
        raw = _build_duole_raw_row(headers, list(values), source_positions)
        raw = _normalize_duole_sheet_row(sheet_name, raw, list(values))
        if not any(str(value).strip() for value in raw.values()):
            continue
        title = raw.get(str(layout.get("title_column") or ""), "") or _choose_duole_title(raw)
        language_name = (
            str(layout.get("language_name") or "").strip()
            or str(raw.get(str(layout.get("language_column") or ""), "")).strip()
            or _infer_duole_language(sheet_name, raw)
        )
        theater_name = (
            str(layout.get("theater_name") or "").strip()
            or str(raw.get(str(layout.get("theater_column") or ""), "")).strip()
            or _infer_duole_theater(sheet_name, raw)
        )
        if not title or not language_name:
            continue
        std = {
            "sheet_name": sheet_name,
            "row_number": row_number,
            "sheet_index": sheet_meta.get("index"),
            "sheet_sid": sheet_meta.get("sid"),
            "sheet_type": sheet_meta.get("type"),
            "used_rows": sheet_meta.get("used_rows"),
            "used_cols": sheet_meta.get("used_cols"),
            "language": language_name,
            "language_name": language_name,
            "theater": normalize_theater_name(theater_name),
            "theater_name": normalize_theater_name(theater_name),
            "title_raw": str(title).strip(),
        }
        std.update(raw)
        parsed_rows.append(std)
    if sheet_name == "2.推荐剧单":
        _fill_down_duole_dates(parsed_rows)
    return parsed_rows


