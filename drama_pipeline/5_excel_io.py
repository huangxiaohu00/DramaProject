from __future__ import annotations

from dataclasses import asdict, is_dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def build_output_path(root: Path, date_text: str, stage: int, name: str) -> Path:
    return Path(root) / date_text / f"{stage}_{name}.xlsx"


def write_workbook(
    output_path: Path | str,
    sheet_rows: Mapping[str, Iterable[Any]],
    sheet_headers: Mapping[str, Iterable[str]] | None = None,
) -> None:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    explicit_headers = {
        str(name): [str(header) for header in headers]
        for name, headers in (sheet_headers or {}).items()
    }

    workbook = openpyxl.Workbook()
    first = True
    for sheet_name, raw_rows in sheet_rows.items():
        rows = [serialize_row(row) for row in raw_rows]
        worksheet = workbook.active if first else workbook.create_sheet()
        first = False
        worksheet.title = safe_sheet_name(sheet_name)
        headers = ordered_headers(rows, explicit_headers.get(str(sheet_name), []))
        if not headers:
            headers = list(explicit_headers.get(str(sheet_name), []))

        for column_index, header in enumerate(headers, 1):
            cell = worksheet.cell(1, column_index, header)
            cell.font = Font(bold=True)

        for row_index, row in enumerate(rows, 2):
            for column_index, header in enumerate(headers, 1):
                worksheet.cell(row_index, column_index, row.get(header, ""))

        autosize_columns(worksheet, headers, rows)

    workbook.save(path)
    workbook.close()


def write_statistics_workbook(
    output_path: Path | str,
    summary_rows: Iterable[Any],
    source_rows: Iterable[Any],
    stage_rows: Iterable[Any],
) -> None:
    write_workbook(
        output_path,
        {
            "汇总统计": list(summary_rows),
            "来源统计": list(source_rows),
            "处理统计": list(stage_rows),
        },
        sheet_headers={
            "汇总统计": ["分类", "指标", "条数"],
            "来源统计": ["阶段", "来源", "维度", "值", "条数"],
            "处理统计": ["阶段", "维度", "值", "条数"],
        },
    )


def serialize_row(row: Any) -> Dict[str, Any]:
    if hasattr(row, "to_dict"):
        return dict(row.to_dict())
    if is_dataclass(row):
        return asdict(row)
    if isinstance(row, dict):
        return dict(row)
    raise TypeError(f"Unsupported row type: {type(row)!r}")


def ordered_headers(rows: List[Dict[str, Any]], preferred: Iterable[str] | None = None) -> List[str]:
    preferred_headers = [str(header) for header in (preferred or [])]
    headers = set()
    for row in rows:
        headers.update(row.keys())
    ordered = [header for header in preferred_headers if header in headers]
    remaining = sorted(header for header in headers if header not in ordered)
    if ordered or remaining:
        return ordered + remaining
    return preferred_headers


def autosize_columns(worksheet, headers: List[str], rows: List[Dict[str, Any]]) -> None:
    for column_index, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row in rows[:500]:
            value = row.get(header, "")
            max_length = max(max_length, len(str(value)))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 10), 60)


def safe_sheet_name(name: str) -> str:
    cleaned = str(name).replace("/", "_").replace("\\", "_").replace("?", "_").replace("*", "_")
    cleaned = cleaned.replace("[", "_").replace("]", "_").replace(":", "_")
    return cleaned[:31] or "Sheet"


def read_workbook_rows(path: Path | str, sheet_name: str) -> List[Dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
        rows: List[Dict[str, Any]] = []
        for row_index in range(2, worksheet.max_row + 1):
            item = {}
            for column_index, header in enumerate(headers, 1):
                if header:
                    item[str(header)] = worksheet.cell(row_index, column_index).value
            rows.append(item)
        return rows
    finally:
        workbook.close()
