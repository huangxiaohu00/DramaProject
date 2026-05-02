from __future__ import annotations

import importlib
import os
import sys
import time
from copy import deepcopy
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Tuple

import openpyxl
from openpyxl.styles import Font


config = importlib.import_module("drama_pipeline.2_config")


WORKBOOK_NAME = "drama_pipeline_config.xlsx"
ADULT_FILTER_WORKBOOK_NAME = "adult_filter.xlsx"
SETTINGS_SHEET = "settings"
ACCOUNTS_SHEET = "accounts"
FEISHU_TABLES_SHEET = "feishu_tables"
QUOTAS_SHEET = "language_quotas"
DUOLE_SHEETS_SHEET = "duole_sheets"
ADULT_FILTER_SHEET = "adult_filter"

SETTINGS_HEADERS = ["section", "key", "value", "required", "notes"]
ACCOUNT_HEADERS = ["platform", "account", "credential", "enabled", "purpose"]
FEISHU_TABLE_HEADERS = ["language", "table_id", "enabled"]
QUOTA_HEADERS = ["language", "theater", "quota"]
DUOLE_SHEET_HEADERS = ["sheet_name", "limit", "enabled"]
ADULT_FILTER_HEADERS = ["语言", "剧场", "剧集名称"]


@dataclass
class RuntimeConfig:
    mobo_drama_auth: str = ""
    mobo_auths: List[str] = field(default_factory=list)
    beidou_drama_auth: str = ""
    beidou_auths: List[str] = field(default_factory=list)
    beidou_agent_id: int = 0
    feishu_app_id: str = ""
    feishu_app_secret: str = ""
    feishu_app_token: str = ""
    feishu_tables: Dict[str, str] = field(default_factory=dict)
    feishu_beidou_hot_spreadsheet_token: str = ""
    feishu_beidou_hot_sheet_name: str = ""
    duole_cookie: str = ""
    duole_share_url: str = ""
    duole_target_sheets: List[str] = field(default_factory=list)
    duole_sheet_limits: Dict[str, int] = field(default_factory=dict)
    material_check_enabled: bool = False
    material_cookie: str = ""
    material_lookback_days: int = 0
    video_threshold: int = 0
    material_shortlist_multiplier: int = 5
    material_prefetch_chunk_size: int = 5
    material_prefetch_workers: int = 5
    material_prefetch_pause_min_seconds: float = 0.5
    material_prefetch_pause_max_seconds: float = 1.0
    material_max_expansion_waves_per_pool: int = 20
    material_post_validation_enabled: bool = False
    material_failed_cooldown_days: int = 7
    today_collect_workers: int = 4
    max_ai_anime_per_language: int = 0
    english_local_translated_ratio: float = 0.5
    language_theater_quotas: Dict[str, Dict[str, int]] = field(default_factory=dict)


@dataclass
class RuntimeContext:
    app_dir: Path
    config_path: Path
    runtime_config: RuntimeConfig
    validation_rows: List[Dict[str, Any]] = field(default_factory=list)


class PipelineLogger:
    def __init__(self, workflow: str, app_dir: Path | str | None = None, run_date: str = ""):
        self.workflow = str(workflow).strip() or "pipeline"
        self.app_dir = resolve_app_dir(app_dir)
        self.run_date = str(run_date).strip()
        self.entries: List[str] = []
        self.problem = False
        self.log_path: Path | None = None

    def set_run_date(self, run_date: str | None) -> None:
        text = str(run_date or "").strip()
        if text:
            self.run_date = text

    def step(self, message: str) -> None:
        self._write("STEP", message)

    def info(self, message: str) -> None:
        self._write("INFO", message)

    def warning(self, message: str) -> None:
        self.problem = True
        self._write("WARN", message)

    def error(self, message: str) -> None:
        self.problem = True
        self._write("ERROR", message)

    def exception(self, exc: Exception, prefix: str = "") -> None:
        text = f"{type(exc).__name__}: {exc}"
        if prefix:
            text = f"{prefix}: {text}"
        self.error(text)

    def flush_if_needed(self) -> Path | None:
        if not self.problem:
            return None
        log_path = self._default_log_path()
        log_path.parent.mkdir(parents=True, exist_ok=True)
        log_path.write_text("\n".join(self.entries) + "\n", encoding="utf-8")
        self.log_path = log_path
        return log_path

    def _write(self, level: str, message: str) -> None:
        stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{stamp}] [{level}] {message}"
        self.entries.append(line)
        _write_text_line(sys.stdout, line)

    def _default_log_path(self) -> Path:
        file_name = f"{self.workflow}_run.log"
        if self.run_date:
            return self.app_dir / self.workflow / self.run_date / file_name
        return self.app_dir / self.workflow / file_name


def resolve_app_dir(app_dir: Path | str | None = None) -> Path:
    if app_dir is not None:
        return Path(app_dir).resolve()
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def _reconfigure_text_stream(
    stream: Any,
    encoding: str,
    errors: str = "strict",
    *,
    line_buffering: bool = False,
    write_through: bool = False,
) -> None:
    reconfigure = getattr(stream, "reconfigure", None)
    if callable(reconfigure):
        try:
            reconfigure(
                encoding=encoding,
                errors=errors,
                line_buffering=line_buffering,
                write_through=write_through,
            )
        except TypeError:
            try:
                reconfigure(encoding=encoding, errors=errors)
            except Exception:
                return
        except Exception:
            return


def _write_text_line(stream: Any, text: str) -> None:
    buffer = getattr(stream, "buffer", None)
    if buffer is not None:
        try:
            buffer.write((text + "\n").encode("utf-8", errors="replace"))
            buffer.flush()
            return
        except Exception:
            pass
    try:
        stream.write(text + "\n")
        flush = getattr(stream, "flush", None)
        if callable(flush):
            flush()
    except Exception:
        return


def _set_windows_console_utf8() -> None:
    if sys.platform != "win32":
        return
    try:
        import ctypes

        kernel32 = ctypes.windll.kernel32
        kernel32.SetConsoleCP(65001)
        kernel32.SetConsoleOutputCP(65001)
    except Exception:
        return


def configure_utf8_runtime() -> None:
    os.environ.setdefault("PYTHONUTF8", "1")
    os.environ.setdefault("PYTHONIOENCODING", "utf-8")
    _set_windows_console_utf8()
    _reconfigure_text_stream(sys.stdin, "utf-8", errors="replace")
    _reconfigure_text_stream(sys.stdout, "utf-8", errors="replace", line_buffering=True, write_through=True)
    _reconfigure_text_stream(sys.stderr, "utf-8", errors="replace", line_buffering=True, write_through=True)


def is_frozen_exe() -> bool:
    return bool(getattr(sys, "frozen", False))


def pause_before_exit(seconds: int = 5) -> None:
    if not is_frozen_exe() or seconds <= 0:
        return
    _write_text_line(sys.stdout, f"处理结束，窗口将在 {seconds} 秒后自动关闭...")
    try:
        time.sleep(seconds)
    except Exception:
        return


def default_config_path(app_dir: Path | str | None = None) -> Path:
    return resolve_app_dir(app_dir) / WORKBOOK_NAME


def default_adult_filter_path(app_dir: Path | str | None = None) -> Path:
    return resolve_app_dir(app_dir) / ADULT_FILTER_WORKBOOK_NAME


def normalize_language_name(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    if not text:
        return ""
    aliases = {
        "英语": "英语",
        "英文": "英语",
        "en": "英语",
        "english": "英语",
        "法语": "法语",
        "法文": "法语",
        "fr": "法语",
        "french": "法语",
        "葡语": "葡萄牙语",
        "葡萄牙语": "葡萄牙语",
        "pt": "葡萄牙语",
        "portuguese": "葡萄牙语",
        "德语": "德语",
        "德文": "德语",
        "de": "德语",
        "german": "德语",
        "繁中": "繁体中文",
        "繁体中文": "繁体中文",
        "繁体": "繁体中文",
        "zh-tw": "繁体中文",
        "traditional chinese": "繁体中文",
        "俄语": "俄语",
        "俄文": "俄语",
        "ru": "俄语",
        "russian": "俄语",
        "意大利语": "意大利语",
        "意语": "意大利语",
        "it": "意大利语",
        "italian": "意大利语",
    }
    direct = aliases.get(text)
    if direct:
        return direct
    lowered = text.lower()
    direct = aliases.get(lowered)
    if direct:
        return direct
    if text in config.LANGUAGE_ORDER:
        return text
    return ""


def runtime_config_from_module(config_module=None) -> RuntimeConfig:
    module = config_module or config
    return RuntimeConfig(
        mobo_drama_auth=str(getattr(module, "MOBO_DRAMA_AUTH", "") or "").strip(),
        mobo_auths=[str(item).strip() for item in list(getattr(module, "MOBO_AUTHS", []) or []) if str(item).strip()],
        beidou_drama_auth=str(getattr(module, "BEIDOU_DRAMA_AUTH", "") or "").strip(),
        beidou_auths=[str(item).strip() for item in list(getattr(module, "BEIDOU_AUTHS", []) or []) if str(item).strip()],
        beidou_agent_id=int(getattr(module, "BEIDOU_AGENT_ID", 0) or 0),
        feishu_app_id=str(getattr(module, "FEISHU_APP_ID", "") or "").strip(),
        feishu_app_secret=str(getattr(module, "FEISHU_APP_SECRET", "") or "").strip(),
        feishu_app_token=str(getattr(module, "FEISHU_APP_TOKEN", "") or "").strip(),
        feishu_tables={
            normalize_language_name(language): str(table_id).strip()
            for language, table_id in dict(getattr(module, "FEISHU_TABLES", {}) or {}).items()
            if normalize_language_name(language) and str(table_id).strip()
        },
        feishu_beidou_hot_spreadsheet_token=str(
            getattr(module, "FEISHU_BEIDOU_HOT_SPREADSHEET_TOKEN", "") or ""
        ).strip(),
        feishu_beidou_hot_sheet_name=str(getattr(module, "FEISHU_BEIDOU_HOT_SHEET_NAME", "") or "").strip(),
        duole_cookie=str(getattr(module, "DUOLE_COOKIE", "") or "").strip(),
        duole_share_url=str(getattr(module, "DUOLE_SHARE_URL", "") or "").strip(),
        duole_target_sheets=[str(item).strip() for item in list(getattr(module, "DUOLE_TARGET_SHEETS", []) or []) if str(item).strip()],
        duole_sheet_limits={
            str(name).strip(): int(limit)
            for name, limit in dict(getattr(module, "DUOLE_SHEET_LIMITS", {}) or {}).items()
            if str(name).strip()
        },
        material_check_enabled=bool(getattr(module, "MATERIAL_CHECK_ENABLED", False)),
        material_cookie=str(getattr(module, "MATERIAL_COOKIE", "") or "").strip(),
        material_lookback_days=int(getattr(module, "MATERIAL_LOOKBACK_DAYS", 0) or 0),
        video_threshold=int(getattr(module, "VIDEO_THRESHOLD", 0) or 0),
        material_shortlist_multiplier=int(getattr(module, "MATERIAL_SHORTLIST_MULTIPLIER", 5) or 5),
        material_prefetch_chunk_size=int(getattr(module, "MATERIAL_PREFETCH_CHUNK_SIZE", 5) or 5),
        material_prefetch_workers=int(getattr(module, "MATERIAL_PREFETCH_WORKERS", 5) or 5),
        material_prefetch_pause_min_seconds=float(getattr(module, "MATERIAL_PREFETCH_PAUSE_MIN_SECONDS", 0.5) or 0.5),
        material_prefetch_pause_max_seconds=float(getattr(module, "MATERIAL_PREFETCH_PAUSE_MAX_SECONDS", 1.0) or 1.0),
        material_max_expansion_waves_per_pool=int(getattr(module, "MATERIAL_MAX_EXPANSION_WAVES_PER_POOL", 20) or 20),
        material_post_validation_enabled=bool(getattr(module, "MATERIAL_POST_VALIDATION_ENABLED", False)),
        material_failed_cooldown_days=int(getattr(module, "MATERIAL_FAILED_COOLDOWN_DAYS", 7) or 7),
        today_collect_workers=int(getattr(module, "TODAY_COLLECT_WORKERS", 4) or 4),
        max_ai_anime_per_language=int(getattr(module, "MAX_AI_ANIME_PER_LANGUAGE", 0) or 0),
        english_local_translated_ratio=float(getattr(module, "ENGLISH_LOCAL_TRANSLATED_RATIO", 0.5) or 0.5),
        language_theater_quotas=deepcopy(dict(getattr(module, "LANGUAGE_THEATER_QUOTAS", {}) or {})),
    )


def create_runtime_workbook(
    output_path: Path | str,
    runtime_config: RuntimeConfig | None = None,
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    data = runtime_config or runtime_config_from_module()

    workbook = openpyxl.Workbook()
    try:
        _write_sheet(workbook.active, SETTINGS_SHEET, SETTINGS_HEADERS, _settings_rows(data))
        _write_sheet(workbook.create_sheet(), ACCOUNTS_SHEET, ACCOUNT_HEADERS, _account_rows(data))
        _write_sheet(workbook.create_sheet(), FEISHU_TABLES_SHEET, FEISHU_TABLE_HEADERS, _feishu_table_rows(data))
        _write_sheet(workbook.create_sheet(), QUOTAS_SHEET, QUOTA_HEADERS, _quota_rows(data))
        _write_sheet(workbook.create_sheet(), DUOLE_SHEETS_SHEET, DUOLE_SHEET_HEADERS, _duole_sheet_rows(data))
        workbook.save(path)
    finally:
        workbook.close()
    return path


def ensure_adult_filter_workbook(path: Path | str) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    if target.exists():
        return target
    workbook = openpyxl.Workbook()
    try:
        worksheet = workbook.active
        worksheet.title = ADULT_FILTER_SHEET
        for column_index, header in enumerate(ADULT_FILTER_HEADERS, 1):
            cell = worksheet.cell(1, column_index, header)
            cell.font = Font(bold=True)
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(column_index)].width = 18
        workbook.save(target)
    finally:
        workbook.close()
    return target


def load_runtime_config(path: Path | str) -> RuntimeConfig:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        settings = _load_settings_sheet(workbook)
        quotas = _load_quota_sheet(workbook)
        duole_sheets = _load_duole_sheet_sheet(workbook)
        feishu_tables = _load_feishu_tables_sheet(workbook)
        accounts = _load_accounts_sheet(workbook)
        return RuntimeConfig(
            mobo_drama_auth=settings.get("today.mobo_drama_auth", ""),
            mobo_auths=accounts.get("mobo", []),
            beidou_drama_auth=settings.get("today.beidou_drama_auth", ""),
            beidou_auths=accounts.get("beidou", []),
            beidou_agent_id=_to_int(settings.get("shared.beidou_agent_id"), config.BEIDOU_AGENT_ID),
            feishu_app_id=settings.get("today.feishu_app_id", ""),
            feishu_app_secret=settings.get("today.feishu_app_secret", ""),
            feishu_app_token=settings.get("today.feishu_app_token", ""),
            feishu_tables=feishu_tables,
            feishu_beidou_hot_spreadsheet_token=settings.get(
                "today.feishu_beidou_hot_spreadsheet_token",
                str(getattr(config, "FEISHU_BEIDOU_HOT_SPREADSHEET_TOKEN", "") or ""),
            ),
            feishu_beidou_hot_sheet_name=settings.get(
                "today.feishu_beidou_hot_sheet_name",
                str(getattr(config, "FEISHU_BEIDOU_HOT_SHEET_NAME", "") or ""),
            ),
            duole_cookie=settings.get("today.duole_cookie", ""),
            duole_share_url=settings.get("today.duole_share_url", str(getattr(config, "DUOLE_SHARE_URL", "") or "")),
            duole_target_sheets=list(duole_sheets.keys()) or list(getattr(config, "DUOLE_TARGET_SHEETS", []) or []),
            duole_sheet_limits=duole_sheets or dict(getattr(config, "DUOLE_SHEET_LIMITS", {}) or {}),
            material_check_enabled=_to_bool(
                settings.get("today.material_check_enabled"),
                bool(getattr(config, "MATERIAL_CHECK_ENABLED", False)),
            ),
            material_cookie=settings.get("today.material_cookie", ""),
            material_lookback_days=_to_int(
                settings.get("today.material_lookback_days"),
                int(getattr(config, "MATERIAL_LOOKBACK_DAYS", 0) or 0),
            ),
            video_threshold=_to_int(
                settings.get("today.video_threshold"),
                int(getattr(config, "VIDEO_THRESHOLD", 0) or 0),
            ),
            material_shortlist_multiplier=_to_int(
                settings.get("today.material_shortlist_multiplier"),
                int(getattr(config, "MATERIAL_SHORTLIST_MULTIPLIER", 5) or 5),
            ),
            material_prefetch_chunk_size=_to_int(
                settings.get("today.material_prefetch_chunk_size"),
                int(getattr(config, "MATERIAL_PREFETCH_CHUNK_SIZE", 5) or 5),
            ),
            material_prefetch_workers=_to_int(
                settings.get("today.material_prefetch_workers"),
                int(getattr(config, "MATERIAL_PREFETCH_WORKERS", 5) or 5),
            ),
            material_prefetch_pause_min_seconds=_to_float(
                settings.get("today.material_prefetch_pause_min_seconds"),
                float(getattr(config, "MATERIAL_PREFETCH_PAUSE_MIN_SECONDS", 0.5) or 0.5),
            ),
            material_prefetch_pause_max_seconds=_to_float(
                settings.get("today.material_prefetch_pause_max_seconds"),
                float(getattr(config, "MATERIAL_PREFETCH_PAUSE_MAX_SECONDS", 1.0) or 1.0),
            ),
            material_max_expansion_waves_per_pool=_to_int(
                settings.get("today.material_max_expansion_waves_per_pool"),
                int(getattr(config, "MATERIAL_MAX_EXPANSION_WAVES_PER_POOL", 20) or 20),
            ),
            material_post_validation_enabled=_to_bool(
                settings.get("today.material_post_validation_enabled"),
                bool(getattr(config, "MATERIAL_POST_VALIDATION_ENABLED", False)),
            ),
            material_failed_cooldown_days=_to_int(
                settings.get("today.material_failed_cooldown_days"),
                int(getattr(config, "MATERIAL_FAILED_COOLDOWN_DAYS", 7) or 7),
            ),
            today_collect_workers=_to_int(
                settings.get("today.today_collect_workers"),
                int(getattr(config, "TODAY_COLLECT_WORKERS", 4) or 4),
            ),
            max_ai_anime_per_language=_to_int(
                settings.get("today.max_ai_anime_per_language"),
                int(getattr(config, "MAX_AI_ANIME_PER_LANGUAGE", 0) or 0),
            ),
            english_local_translated_ratio=_to_float(
                settings.get("today.english_local_translated_ratio"),
                float(getattr(config, "ENGLISH_LOCAL_TRANSLATED_RATIO", 0.5) or 0.5),
            ),
            language_theater_quotas=quotas or deepcopy(dict(getattr(config, "LANGUAGE_THEATER_QUOTAS", {}) or {})),
        )
    finally:
        workbook.close()


def validate_runtime_config(runtime_config: RuntimeConfig) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    def add(level: str, area: str, item: str, message: str, value: Any = "") -> None:
        rows.append({"级别": level, "区域": area, "项目": item, "问题": message, "当前值": _display_value(value)})

    required_credentials = (
        ("today", "mobo_drama_auth", runtime_config.mobo_drama_auth, "Mobo 剧库授权为空，Mobo 新剧可能无法采集"),
        ("today", "beidou_drama_auth", runtime_config.beidou_drama_auth, "北斗任务授权为空，北斗新剧可能无法采集"),
        ("today", "feishu_app_id", runtime_config.feishu_app_id, "飞书 app_id 为空，已发布表可能无法采集"),
        ("today", "feishu_app_secret", runtime_config.feishu_app_secret, "飞书 app_secret 为空，已发布表可能无法采集"),
        ("today", "feishu_app_token", runtime_config.feishu_app_token, "飞书 app_token 为空，已发布表可能无法采集"),
        ("today", "duole_cookie", runtime_config.duole_cookie, "多乐 cookie 为空，多乐推荐榜可能无法采集"),
    )
    for area, item, value, message in required_credentials:
        if not str(value or "").strip():
            add("WARN", area, item, message, value)

    if not runtime_config.mobo_auths:
        add("WARN", "accounts", "mobo", "Mobo 订单账号为空，收入/订单校验信号可能缺失")
    if not runtime_config.beidou_auths:
        add("WARN", "accounts", "beidou", "北斗订单账号为空，收入/订单校验信号可能缺失")
    if int(runtime_config.beidou_agent_id or 0) <= 0:
        add("ERROR", "shared", "beidou_agent_id", "北斗 agent_id 必须为正整数", runtime_config.beidou_agent_id)

    known_languages = set(getattr(config, "LANGUAGE_ORDER", []) or [])
    if not runtime_config.feishu_tables:
        add("WARN", "feishu_tables", "table_id", "飞书语言表为空，已发布过滤可能失效")
    for language, table_id in sorted(runtime_config.feishu_tables.items()):
        if known_languages and language not in known_languages:
            add("ERROR", "feishu_tables", language, "飞书表语种不在系统语种列表中", language)
        if not str(table_id or "").strip():
            add("ERROR", "feishu_tables", language, "飞书 table_id 为空", table_id)

    if not runtime_config.duole_target_sheets:
        add("WARN", "duole_sheets", "sheet_name", "多乐目标 sheet 为空，多乐推荐源将缺失")
    for sheet_name in runtime_config.duole_target_sheets:
        limit = int(runtime_config.duole_sheet_limits.get(sheet_name, 0) or 0)
        if limit <= 0:
            add("ERROR", "duole_sheets", sheet_name, "多乐 sheet limit 必须为正整数", limit)

    if int(runtime_config.video_threshold or 0) < 0:
        add("ERROR", "today", "video_threshold", "素材达标阈值不能为负数", runtime_config.video_threshold)
    if int(runtime_config.material_lookback_days or 0) < 0:
        add("ERROR", "today", "material_lookback_days", "素材回看天数不能为负数", runtime_config.material_lookback_days)
    if int(runtime_config.material_shortlist_multiplier or 0) < 1:
        add("ERROR", "today", "material_shortlist_multiplier", "素材初选倍数必须大于等于 1", runtime_config.material_shortlist_multiplier)
    if int(runtime_config.material_prefetch_chunk_size or 0) < 1:
        add("ERROR", "today", "material_prefetch_chunk_size", "素材批次大小必须大于等于 1", runtime_config.material_prefetch_chunk_size)
    material_workers = int(runtime_config.material_prefetch_workers or 0)
    if material_workers < 1 or material_workers > 5:
        add("ERROR", "today", "material_prefetch_workers", "素材并发线程必须在 1 到 5 之间，避免增加风控风险", runtime_config.material_prefetch_workers)
    collect_workers = int(runtime_config.today_collect_workers or 0)
    if collect_workers < 1 or collect_workers > 4:
        add("ERROR", "today", "today_collect_workers", "今日源数据采集线程必须在 1 到 4 之间", runtime_config.today_collect_workers)
    pause_min = float(runtime_config.material_prefetch_pause_min_seconds or 0.0)
    pause_max = float(runtime_config.material_prefetch_pause_max_seconds or 0.0)
    if pause_min < 0:
        add("ERROR", "today", "material_prefetch_pause_min_seconds", "素材批次最小间隔不能为负数", runtime_config.material_prefetch_pause_min_seconds)
    if pause_max < pause_min:
        add("ERROR", "today", "material_prefetch_pause_max_seconds", "素材批次最大间隔不能小于最小间隔", runtime_config.material_prefetch_pause_max_seconds)
    if int(runtime_config.material_max_expansion_waves_per_pool or 0) < 1:
        add("ERROR", "today", "material_max_expansion_waves_per_pool", "素材扩容轮次上限必须大于等于 1", runtime_config.material_max_expansion_waves_per_pool)
    if int(runtime_config.material_failed_cooldown_days or 0) < 0:
        add("ERROR", "today", "material_failed_cooldown_days", "不达标素材冷却天数不能为负数", runtime_config.material_failed_cooldown_days)
    if int(runtime_config.max_ai_anime_per_language or 0) < 0:
        add("ERROR", "today", "max_ai_anime_per_language", "AI/漫剧语种上限不能为负数", runtime_config.max_ai_anime_per_language)
    ratio = float(runtime_config.english_local_translated_ratio or 0.0)
    if ratio < 0 or ratio > 1:
        add("ERROR", "today", "english_local_translated_ratio", "英语本土/翻译比例必须在 0 到 1 之间", ratio)

    seen_quota_keys: set[Tuple[str, str]] = set()
    for language, theater_quotas in sorted(runtime_config.language_theater_quotas.items()):
        if known_languages and language not in known_languages:
            add("ERROR", "language_quotas", language, "配额语种不在系统语种列表中", language)
        if not isinstance(theater_quotas, Mapping) or not theater_quotas:
            add("WARN", "language_quotas", language, "该语种没有剧场配额", theater_quotas)
            continue
        language_total = 0
        for theater, quota in sorted(theater_quotas.items()):
            key = (language, theater)
            if key in seen_quota_keys:
                add("ERROR", "language_quotas", f"{language}/{theater}", "剧场配额重复", quota)
            seen_quota_keys.add(key)
            quota_value = int(quota or 0)
            if quota_value < 0:
                add("ERROR", "language_quotas", f"{language}/{theater}", "剧场配额不能为负数", quota)
            language_total += max(quota_value, 0)
        if language_total <= 0:
            add("ERROR", "language_quotas", language, "该语种总配额必须大于 0", language_total)

    return rows


def has_blocking_validation_errors(validation_rows: Iterable[Mapping[str, Any]]) -> bool:
    return any(str(row.get("级别", "")).upper() == "ERROR" for row in validation_rows)


def write_validation_report(output_path: Path | str, validation_rows: Iterable[Mapping[str, Any]]) -> None:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = list(validation_rows)
    workbook = openpyxl.Workbook()
    try:
        _write_sheet(
            workbook.active,
            "配置校验",
            ["级别", "区域", "项目", "问题", "当前值"],
            rows or [{"级别": "OK", "区域": "runtime", "项目": "all", "问题": "配置校验通过", "当前值": ""}],
        )
        workbook.save(path)
    finally:
        workbook.close()


def _display_value(value: Any) -> str:
    if isinstance(value, (list, tuple, set)):
        return ",".join(str(item) for item in value)
    if isinstance(value, Mapping):
        return ";".join(f"{key}={val}" for key, val in value.items())
    return str(value or "")


def apply_runtime_config(
    runtime_config: RuntimeConfig,
    config_module=None,
    app_dir: Path | str | None = None,
) -> None:
    module = config_module or config
    resolved_app_dir = resolve_app_dir(app_dir)
    setattr(module, "MOBO_DRAMA_AUTH", runtime_config.mobo_drama_auth)
    setattr(module, "MOBO_AUTHS", list(runtime_config.mobo_auths))
    setattr(module, "BEIDOU_DRAMA_AUTH", runtime_config.beidou_drama_auth)
    setattr(module, "BEIDOU_AUTHS", list(runtime_config.beidou_auths))
    setattr(module, "BEIDOU_AGENT_ID", int(runtime_config.beidou_agent_id or 0))
    setattr(module, "FEISHU_APP_ID", runtime_config.feishu_app_id)
    setattr(module, "FEISHU_APP_SECRET", runtime_config.feishu_app_secret)
    setattr(module, "FEISHU_APP_TOKEN", runtime_config.feishu_app_token)
    setattr(module, "FEISHU_TABLES", dict(runtime_config.feishu_tables))
    setattr(module, "FEISHU_BEIDOU_HOT_SPREADSHEET_TOKEN", runtime_config.feishu_beidou_hot_spreadsheet_token)
    setattr(module, "FEISHU_BEIDOU_HOT_SHEET_NAME", runtime_config.feishu_beidou_hot_sheet_name)
    setattr(module, "DUOLE_COOKIE", runtime_config.duole_cookie)
    setattr(module, "DUOLE_SHARE_URL", runtime_config.duole_share_url)
    setattr(module, "DUOLE_TARGET_SHEETS", list(runtime_config.duole_target_sheets))
    setattr(module, "DUOLE_SHEET_LIMITS", dict(runtime_config.duole_sheet_limits))
    setattr(module, "MATERIAL_CHECK_ENABLED", bool(runtime_config.material_check_enabled))
    setattr(module, "MATERIAL_COOKIE", runtime_config.material_cookie)
    setattr(module, "MATERIAL_LOOKBACK_DAYS", int(runtime_config.material_lookback_days or 0))
    setattr(module, "VIDEO_THRESHOLD", int(runtime_config.video_threshold or 0))
    setattr(module, "MATERIAL_SHORTLIST_MULTIPLIER", int(runtime_config.material_shortlist_multiplier or 1))
    setattr(module, "MATERIAL_PREFETCH_CHUNK_SIZE", int(runtime_config.material_prefetch_chunk_size or 1))
    setattr(module, "MATERIAL_PREFETCH_WORKERS", min(max(int(runtime_config.material_prefetch_workers or 1), 1), 5))
    setattr(module, "MATERIAL_PREFETCH_PAUSE_MIN_SECONDS", float(runtime_config.material_prefetch_pause_min_seconds or 0.0))
    setattr(module, "MATERIAL_PREFETCH_PAUSE_MAX_SECONDS", float(runtime_config.material_prefetch_pause_max_seconds or 0.0))
    setattr(module, "MATERIAL_MAX_EXPANSION_WAVES_PER_POOL", max(int(runtime_config.material_max_expansion_waves_per_pool or 1), 1))
    setattr(module, "MATERIAL_POST_VALIDATION_ENABLED", bool(runtime_config.material_post_validation_enabled))
    setattr(module, "MATERIAL_FAILED_COOLDOWN_DAYS", max(int(runtime_config.material_failed_cooldown_days or 0), 0))
    setattr(module, "TODAY_COLLECT_WORKERS", min(max(int(runtime_config.today_collect_workers or 1), 1), 4))
    setattr(module, "MAX_AI_ANIME_PER_LANGUAGE", int(runtime_config.max_ai_anime_per_language or 0))
    setattr(module, "ENGLISH_LOCAL_TRANSLATED_RATIO", float(runtime_config.english_local_translated_ratio or 0.0))
    setattr(module, "LANGUAGE_THEATER_QUOTAS", deepcopy(runtime_config.language_theater_quotas))
    setattr(module, "OUTPUT_ROOT", resolved_app_dir)
    setattr(module, "TODAY_OUTPUT_ROOT", resolved_app_dir / "today_recommend")
    setattr(module, "ORDER_OUTPUT_ROOT", resolved_app_dir)


def bootstrap_runtime(
    config_module=None,
    app_dir: Path | str | None = None,
    config_path: Path | str | None = None,
    create_if_missing: bool = False,
    logger: PipelineLogger | None = None,
) -> RuntimeContext:
    module = config_module or config
    resolved_app_dir = resolve_app_dir(app_dir)
    resolved_config_path = Path(config_path) if config_path is not None else default_config_path(resolved_app_dir)
    if not resolved_config_path.exists():
        if not create_if_missing:
            raise FileNotFoundError(f"runtime config workbook not found: {resolved_config_path}")
        create_runtime_workbook(resolved_config_path, runtime_config_from_module(module))
        if logger is not None:
            logger.warning(f"未找到配置文件，已生成模板: {resolved_config_path}")
    runtime_config = load_runtime_config(resolved_config_path)
    validation_rows = validate_runtime_config(runtime_config)
    validation_path = resolved_app_dir / "config_validation.xlsx"
    write_validation_report(validation_path, validation_rows)
    for row in validation_rows:
        if logger is not None:
            message = f"配置校验[{row.get('级别')}]: {row.get('区域')}.{row.get('项目')} - {row.get('问题')}"
            if str(row.get("级别", "")).upper() == "ERROR":
                logger.error(message)
            else:
                logger.warning(message)
    if has_blocking_validation_errors(validation_rows):
        raise ValueError(f"runtime config validation failed, see: {validation_path}")
    apply_runtime_config(runtime_config, config_module=module, app_dir=resolved_app_dir)
    ensure_adult_filter_workbook(default_adult_filter_path(resolved_app_dir))
    if logger is not None:
        logger.info(f"已加载运行时配置: {resolved_config_path}")
    return RuntimeContext(
        app_dir=resolved_app_dir,
        config_path=resolved_config_path,
        runtime_config=runtime_config,
        validation_rows=validation_rows,
    )


def _write_sheet(worksheet, title: str, headers: Iterable[str], rows: Iterable[Mapping[str, Any]]) -> None:
    worksheet.title = title
    for column_index, header in enumerate(headers, 1):
        cell = worksheet.cell(1, column_index, header)
        cell.font = Font(bold=True)
    for row_index, row in enumerate(rows, 2):
        for column_index, header in enumerate(headers, 1):
            worksheet.cell(row_index, column_index, row.get(header, ""))
    for column_index, header in enumerate(headers, 1):
        width = max(len(str(header)), 14)
        for row_index in range(2, min(worksheet.max_row, 100) + 1):
            width = max(width, len(str(worksheet.cell(row_index, column_index).value or "")) + 2)
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(column_index)].width = min(width, 80)


def _settings_rows(runtime_config: RuntimeConfig) -> List[Dict[str, Any]]:
    return [
        _setting("today", "mobo_drama_auth", runtime_config.mobo_drama_auth, True, "今日推荐-Mobo 新剧/推荐接口"),
        _setting("today", "beidou_drama_auth", runtime_config.beidou_drama_auth, True, "今日推荐-北斗新剧/收入榜接口"),
        _setting("today", "feishu_app_id", runtime_config.feishu_app_id, True, "飞书 app_id"),
        _setting("today", "feishu_app_secret", runtime_config.feishu_app_secret, True, "飞书 app_secret"),
        _setting("today", "feishu_app_token", runtime_config.feishu_app_token, True, "飞书 bitable app_token"),
        _setting("today", "feishu_beidou_hot_spreadsheet_token", runtime_config.feishu_beidou_hot_spreadsheet_token, True, "飞书北斗热门表 spreadsheet_token"),
        _setting("today", "feishu_beidou_hot_sheet_name", runtime_config.feishu_beidou_hot_sheet_name, True, "飞书北斗热门 sheet 名称"),
        _setting("today", "duole_cookie", runtime_config.duole_cookie, True, "多乐 cookie"),
        _setting("today", "duole_share_url", runtime_config.duole_share_url, True, "多乐共享表 URL"),
        _setting("today", "material_check_enabled", runtime_config.material_check_enabled, False, "是否启用素材达标校验"),
        _setting("today", "material_cookie", runtime_config.material_cookie, False, "素材接口 cookie"),
        _setting("today", "material_lookback_days", runtime_config.material_lookback_days, False, "素材回看天数"),
        _setting("today", "video_threshold", runtime_config.video_threshold, False, "素材达标视频数阈值"),
        _setting("today", "material_shortlist_multiplier", runtime_config.material_shortlist_multiplier, False, "素材初选倍数，建议 5"),
        _setting("today", "material_prefetch_chunk_size", runtime_config.material_prefetch_chunk_size, False, "素材每批查询数，建议 5"),
        _setting("today", "material_prefetch_workers", runtime_config.material_prefetch_workers, False, "素材每批并发线程，上限 5"),
        _setting("today", "material_prefetch_pause_min_seconds", runtime_config.material_prefetch_pause_min_seconds, False, "素材批次间最小间隔秒"),
        _setting("today", "material_prefetch_pause_max_seconds", runtime_config.material_prefetch_pause_max_seconds, False, "素材批次间最大间隔秒"),
        _setting("today", "material_max_expansion_waves_per_pool", runtime_config.material_max_expansion_waves_per_pool, False, "单语言剧场素材扩容最大轮次"),
        _setting("today", "material_post_validation_enabled", runtime_config.material_post_validation_enabled, False, "是否启用素材后置校验"),
        _setting("today", "material_failed_cooldown_days", runtime_config.material_failed_cooldown_days, False, "近 N 天素材不达标剧集直接过滤"),
        _setting("today", "today_collect_workers", runtime_config.today_collect_workers, False, "今日源数据采集线程数，上限 4"),
        _setting("today", "max_ai_anime_per_language", runtime_config.max_ai_anime_per_language, False, "单语种 AI/漫剧最大推荐数"),
        _setting("today", "english_local_translated_ratio", runtime_config.english_local_translated_ratio, False, "英语本土剧与翻译剧目标占比"),
        _setting("shared", "beidou_agent_id", runtime_config.beidou_agent_id, True, "北斗 agent_id"),
    ]


def _account_rows(runtime_config: RuntimeConfig) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for index, credential in enumerate(runtime_config.mobo_auths, 1):
        rows.append(
            {
                "platform": "mobo",
                "account": f"mobo_{index}",
                "credential": credential,
                "enabled": True,
                "purpose": "昨日订单",
            }
        )
    for index, credential in enumerate(runtime_config.beidou_auths, 1):
        rows.append(
            {
                "platform": "beidou",
                "account": f"beidou_{index}",
                "credential": credential,
                "enabled": True,
                "purpose": "昨日订单",
            }
        )
    return rows


def _feishu_table_rows(runtime_config: RuntimeConfig) -> List[Dict[str, Any]]:
    rows = []
    for language in config.LANGUAGE_ORDER:
        rows.append(
            {
                "language": language,
                "table_id": runtime_config.feishu_tables.get(language, ""),
                "enabled": bool(runtime_config.feishu_tables.get(language)),
            }
        )
    return rows


def _quota_rows(runtime_config: RuntimeConfig) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for language in config.LANGUAGE_ORDER:
        quotas = runtime_config.language_theater_quotas.get(language, {})
        for theater, quota in quotas.items():
            rows.append({"language": language, "theater": theater, "quota": int(quota)})
    return rows


def _duole_sheet_rows(runtime_config: RuntimeConfig) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    seen = set()
    for sheet_name in runtime_config.duole_target_sheets:
        seen.add(sheet_name)
        rows.append(
            {
                "sheet_name": sheet_name,
                "limit": int(runtime_config.duole_sheet_limits.get(sheet_name, 0) or 0),
                "enabled": True,
            }
        )
    for sheet_name, limit in runtime_config.duole_sheet_limits.items():
        if sheet_name in seen:
            continue
        rows.append({"sheet_name": sheet_name, "limit": int(limit or 0), "enabled": False})
    return rows


def _setting(section: str, key: str, value: Any, required: bool, notes: str) -> Dict[str, Any]:
    return {
        "section": section,
        "key": key,
        "value": "" if value is None else value,
        "required": "Y" if required else "",
        "notes": notes,
    }


def _load_settings_sheet(workbook) -> Dict[str, str]:
    if SETTINGS_SHEET not in workbook.sheetnames:
        return {}
    rows = _iter_sheet_rows(workbook[SETTINGS_SHEET])
    output: Dict[str, str] = {}
    for row in rows:
        section = str(row.get("section") or "").strip()
        key = str(row.get("key") or "").strip()
        if not section or not key:
            continue
        output[f"{section}.{key}"] = str(row.get("value") or "").strip()
    return output


def _load_accounts_sheet(workbook) -> Dict[str, List[str]]:
    output: Dict[str, List[str]] = {}
    if ACCOUNTS_SHEET not in workbook.sheetnames:
        return output
    for row in _iter_sheet_rows(workbook[ACCOUNTS_SHEET]):
        if not _to_bool(row.get("enabled"), True):
            continue
        platform = str(row.get("platform") or "").strip().lower()
        credential = str(row.get("credential") or "").strip()
        if not platform or not credential:
            continue
        output.setdefault(platform, []).append(credential)
    return output


def _load_feishu_tables_sheet(workbook) -> Dict[str, str]:
    output: Dict[str, str] = {}
    if FEISHU_TABLES_SHEET not in workbook.sheetnames:
        return output
    for row in _iter_sheet_rows(workbook[FEISHU_TABLES_SHEET]):
        if not _to_bool(row.get("enabled"), True):
            continue
        language = normalize_language_name(row.get("language"))
        table_id = str(row.get("table_id") or "").strip()
        if language and table_id:
            output[language] = table_id
    return output


def _load_quota_sheet(workbook) -> Dict[str, Dict[str, int]]:
    output: Dict[str, Dict[str, int]] = {}
    if QUOTAS_SHEET not in workbook.sheetnames:
        return output
    for row in _iter_sheet_rows(workbook[QUOTAS_SHEET]):
        language = normalize_language_name(row.get("language"))
        theater = config.normalize_theater_name(row.get("theater"))
        quota = _to_int(row.get("quota"), 0)
        if language and theater and quota > 0:
            output.setdefault(language, {})[theater] = quota
    return output


def _load_duole_sheet_sheet(workbook) -> Dict[str, int]:
    output: Dict[str, int] = {}
    if DUOLE_SHEETS_SHEET not in workbook.sheetnames:
        return output
    for row in _iter_sheet_rows(workbook[DUOLE_SHEETS_SHEET]):
        if not _to_bool(row.get("enabled"), True):
            continue
        sheet_name = str(row.get("sheet_name") or "").strip()
        limit = _to_int(row.get("limit"), 0)
        if sheet_name and limit > 0:
            output[sheet_name] = limit
    return output


def _iter_sheet_rows(worksheet) -> List[Dict[str, Any]]:
    headers = [str(value).strip() if value is not None else "" for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())]
    rows: List[Dict[str, Any]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        row: Dict[str, Any] = {}
        has_value = False
        for index, header in enumerate(headers):
            if not header:
                continue
            value = values[index] if index < len(values) else None
            row[header] = value
            if value not in (None, ""):
                has_value = True
        if has_value:
            rows.append(row)
    return rows


def update_duole_cookie_in_config(cookie_value: str, config_path: Path | str | None = None) -> None:
    """将新的 duole_cookie 写入 drama_pipeline_config.xlsx 的 settings sheet。"""
    path = Path(config_path) if config_path else default_config_path()
    if not path.exists():
        return
    workbook = openpyxl.load_workbook(str(path))
    try:
        if SETTINGS_SHEET not in workbook.sheetnames:
            return
        ws = workbook[SETTINGS_SHEET]
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        key_col = None
        value_col = None
        for i, h in enumerate(headers, 1):
            if h == "key":
                key_col = i
            elif h == "value":
                value_col = i
        if key_col is None or value_col is None:
            return
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            cell_key = row[key_col - 1].value
            if cell_key is not None and str(cell_key).strip() == "duole_cookie":
                row[value_col - 1].value = cookie_value
                break
        workbook.save(str(path))
    finally:
        workbook.close()


def _to_bool(value: Any, default: bool = False) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"1", "true", "y", "yes", "on"}:
        return True
    if text in {"0", "false", "n", "no", "off"}:
        return False
    return default


def _to_int(value: Any, default: int = 0) -> int:
    if value is None or value == "":
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _to_float(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default
