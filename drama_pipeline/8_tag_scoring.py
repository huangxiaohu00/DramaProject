from __future__ import annotations

import importlib
import math
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence, Tuple

import openpyxl


config = importlib.import_module("drama_pipeline.2_config")


@dataclass(frozen=True)
class TagGlobalWeight:
    tag: str
    score: float
    orders: float


@dataclass(frozen=True)
class TagLevelEvidence:
    level: str
    raw_score: float
    orders: float
    level_confidence: float
    sample_confidence: float
    final_confidence: float
    calibrated_score: float


@dataclass(frozen=True)
class TagWeightResult:
    tag: str
    weight: float
    status: str
    evidence: List[TagLevelEvidence] = field(default_factory=list)


@dataclass(frozen=True)
class DramaTagScore:
    score: float
    status: str
    matched_tags: List[str]
    details: Dict[str, Any] = field(default_factory=dict)


@dataclass
class _LevelWeight:
    score: float
    orders: float


class TagWeightScorer:
    def __init__(self, workbook_path: Path | str = config.TAG_SCORING_WORKBOOK):
        self.workbook_path = Path(workbook_path)
        self.normalized_tags: Dict[str, str] = {}
        self.language_weights: Dict[Tuple[str, str], _LevelWeight] = {}
        self.theater_weights: Dict[Tuple[str, str], _LevelWeight] = {}
        self.language_theater_weights: Dict[Tuple[str, str, str], _LevelWeight] = {}
        self.global_weights: Dict[str, TagGlobalWeight] = {}
        if self.workbook_path.exists():
            self._load()

    def score_record(self, record: Any) -> DramaTagScore:
        theater = getattr(record, "theater", "")
        raw = getattr(record, "raw", {}) or {}
        if isinstance(raw, dict):
            merged_theaters = raw.get("merged_theaters") or []
            if isinstance(merged_theaters, (list, tuple)):
                theaters = [str(item).strip() for item in merged_theaters if str(item).strip()]
            else:
                theaters = []
        else:
            theaters = []
        theaters.extend(_split_theater_names(theater))
        return self.score_tags_for_theaters(record.language, theaters, getattr(record, "tags", []))

    def score_tags(self, language: str, theater: str, tags: Iterable[Any]) -> DramaTagScore:
        return self.score_tags_for_theaters(language, [theater] if theater else [], tags)

    def score_tags_for_theaters(self, language: str, theaters: Sequence[str], tags: Iterable[Any]) -> DramaTagScore:
        normalized_tags = []
        for tag in tags:
            normalized = self.normalize_tag(tag)
            if normalized and normalized not in normalized_tags:
                normalized_tags.append(normalized)

        if not normalized_tags:
            return DramaTagScore(score=config.TAG_DEFAULT_SCORE, status="无标签兜底", matched_tags=[], details={})

        tag_results = [self.best_tag_weight(language, theaters, tag) for tag in normalized_tags]
        usable = [item for item in tag_results if item.evidence]
        if not usable:
            return DramaTagScore(score=config.TAG_DEFAULT_SCORE, status="标签无权重兜底", matched_tags=[], details={})

        usable.sort(key=lambda item: item.weight, reverse=True)
        score = 0.0
        matched_tags: List[str] = []
        for result, decay in zip(usable, config.TAG_COMBINE_DECAYS):
            score += result.weight * decay
            matched_tags.append(result.tag)

        details = {
            "tag_weights": {
                result.tag: {
                    "weight": round(result.weight, 4),
                    "status": result.status,
                    "evidence": [evidence.__dict__ for evidence in result.evidence],
                }
                for result in usable
            }
        }
        return DramaTagScore(
            score=round(min(score, config.TAG_SCORE_CAP), 4),
            status="标签命中",
            matched_tags=matched_tags,
            details=details,
        )

    def best_tag_weight(self, language: str, theaters: Sequence[str], tag: Any) -> TagWeightResult:
        candidates = [""]
        for theater in theaters:
            normalized = config.normalize_theater_name(theater)
            if normalized and normalized not in candidates:
                candidates.append(normalized)
        best: TagWeightResult | None = None
        for theater in candidates:
            current = self.tag_weight(language, theater, tag)
            if best is None or current.weight > best.weight or (current.weight == best.weight and len(current.evidence) > len(best.evidence)):
                best = current
        return best or self.tag_weight(language, "", tag)

    def tag_weight(self, language: str, theater: str, tag: Any) -> TagWeightResult:
        normalized = self.normalize_tag(tag)
        global_weight = self.global_weight(normalized)
        evidence: List[TagLevelEvidence] = []

        self._append_evidence(
            evidence,
            "language_theater",
            self.language_theater_weights.get((language, theater, normalized)),
            global_weight.score,
        )
        self._append_evidence(evidence, "language", self.language_weights.get((language, normalized)), global_weight.score)
        self._append_evidence(evidence, "theater", self.theater_weights.get((theater, normalized)), global_weight.score)
        if normalized in self.global_weights:
            self._append_evidence(evidence, "global", _LevelWeight(global_weight.score, global_weight.orders), global_weight.score)

        weighted_sum = sum(item.calibrated_score * item.final_confidence for item in evidence)
        confidence_sum = sum(item.final_confidence for item in evidence)
        if confidence_sum <= 0:
            return TagWeightResult(tag=normalized, weight=config.TAG_DEFAULT_SCORE, status="标签无权重兜底", evidence=[])
        return TagWeightResult(
            tag=normalized,
            weight=round(weighted_sum / confidence_sum, 4),
            status="标签命中",
            evidence=evidence,
        )

    def global_weight(self, tag: Any) -> TagGlobalWeight:
        normalized = self.normalize_tag(tag)
        return self.global_weights.get(normalized, TagGlobalWeight(normalized, config.TAG_DEFAULT_SCORE, 0))

    def normalize_tag(self, tag: Any) -> str:
        text = "" if tag is None else str(tag).strip()
        if not text:
            return ""
        normalized = self.normalized_tags.get(text, self.normalized_tags.get(text.lower(), text))
        if normalized in config.TAG_EXCLUDED_BROAD_TAGS:
            return ""
        return normalized

    def _load(self) -> None:
        workbook = openpyxl.load_workbook(self.workbook_path, read_only=True, data_only=True)
        try:
            self._load_normalized_tags(workbook)
            self.language_weights = self._load_level_sheet(workbook, "按语言偏好标签", ("语言", "标签"))
            self.theater_weights = self._load_level_sheet(workbook, "按剧场偏好标签", ("剧场", "标签"))
            self.language_theater_weights = self._load_level_sheet(workbook, "按语言剧场偏好标签", ("语言", "剧场", "标签"))
            self.global_weights = self._build_global_weights(workbook)
        finally:
            workbook.close()

    def _load_normalized_tags(self, workbook: Any) -> None:
        if "标签归一化检查" not in workbook.sheetnames:
            return
        worksheet = workbook["标签归一化检查"]
        header = _header_map(worksheet)
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            raw = _cell(row, header, "raw_tag")
            normalized = _cell(row, header, "normalized_tag")
            if raw and normalized:
                self.normalized_tags[str(raw).strip()] = str(normalized).strip()
                self.normalized_tags[str(raw).strip().lower()] = str(normalized).strip()

    def _load_level_sheet(self, workbook: Any, sheet_name: str, key_columns: Tuple[str, ...]) -> Dict[Tuple[str, ...], _LevelWeight]:
        if sheet_name not in workbook.sheetnames:
            return {}
        worksheet = workbook[sheet_name]
        header = _header_map(worksheet)
        grouped: Dict[Tuple[str, ...], List[Tuple[float, float]]] = defaultdict(list)
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            values = []
            for column in key_columns:
                value = _cell(row, header, column)
                if column == "标签":
                    value = self.normalize_tag(value)
                values.append(str(value).strip() if value is not None else "")
            if not all(values):
                continue
            weight = _normalize_theme_score(_safe_float(_cell(row, header, "推荐权重")))
            orders = _safe_float(_cell(row, header, "订单数"))
            grouped[tuple(values)].append((weight, orders))
        return {key: _merge_level_weights(values) for key, values in grouped.items()}

    def _build_global_weights(self, workbook: Any) -> Dict[str, TagGlobalWeight]:
        if "按语言偏好标签" not in workbook.sheetnames:
            return {}
        worksheet = workbook["按语言偏好标签"]
        header = _header_map(worksheet)
        totals: Dict[str, Tuple[float, float]] = defaultdict(lambda: (0.0, 0.0))
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            tag = self.normalize_tag(_cell(row, header, "标签"))
            if not tag:
                continue
            amount = _safe_float(_cell(row, header, "金额"))
            orders = _safe_float(_cell(row, header, "订单数"))
            current_amount, current_orders = totals[tag]
            totals[tag] = (current_amount + amount, current_orders + orders)
        if not totals:
            return {}

        max_amount = max(amount for amount, _orders in totals.values()) or 1.0
        max_orders = max(orders for _amount, orders in totals.values()) or 1.0
        raw_scores: Dict[str, float] = {}
        for tag, (amount, orders) in totals.items():
            amount_norm = amount / max_amount
            order_norm = orders / max_orders
            raw_scores[tag] = config.TAG_AMOUNT_WEIGHT * amount_norm + config.TAG_ORDER_WEIGHT * order_norm
        max_raw = max(raw_scores.values()) or 1.0

        output: Dict[str, TagGlobalWeight] = {}
        for tag, raw_score in raw_scores.items():
            _amount, orders = totals[tag]
            preference = raw_score / max_raw
            sample_confidence = sample_confidence_from_orders(orders)
            preference_score = _normalize_theme_score(preference)
            score = config.TAG_DEFAULT_SCORE + sample_confidence * (preference_score - config.TAG_DEFAULT_SCORE)
            output[tag] = TagGlobalWeight(tag=tag, score=round(score, 4), orders=orders)
        return output

    def _append_evidence(
        self,
        evidence: List[TagLevelEvidence],
        level: str,
        level_weight: _LevelWeight | None,
        global_score: float,
    ) -> None:
        if level_weight is None:
            return
        level_confidence = config.TAG_LEVEL_CONFIDENCE[level]
        sample_confidence = sample_confidence_from_orders(level_weight.orders)
        final_confidence = level_confidence * sample_confidence
        calibrated = global_score + final_confidence * (level_weight.score - global_score)
        evidence.append(
            TagLevelEvidence(
                level=level,
                raw_score=level_weight.score,
                orders=level_weight.orders,
                level_confidence=level_confidence,
                sample_confidence=round(sample_confidence, 6),
                final_confidence=round(final_confidence, 6),
                calibrated_score=round(calibrated, 4),
            )
        )


def sample_confidence_from_orders(orders: float) -> float:
    if orders <= 0:
        return 0.0
    cap = max(float(config.TAG_SAMPLE_CONFIDENCE_ORDER_CAP), 1.0)
    return min(1.0, math.log1p(float(orders)) / math.log1p(cap))


def _header_map(worksheet: Any) -> Dict[str, int]:
    return {
        str(worksheet.cell(1, column).value).strip(): column - 1
        for column in range(1, worksheet.max_column + 1)
        if worksheet.cell(1, column).value is not None
    }


def _cell(row: Tuple[Any, ...], header: Dict[str, int], name: str) -> Any:
    index = header.get(name)
    if index is None or index >= len(row):
        return None
    return row[index]


def _safe_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _normalize_theme_score(value: float) -> float:
    clamped = max(0.0, float(value))
    if clamped <= 1.0:
        normalized = config.TAG_SCORE_FLOOR + (config.TAG_SCORE_CAP - config.TAG_SCORE_FLOOR) * clamped
    else:
        normalized = clamped
    return round(max(config.TAG_SCORE_FLOOR, min(config.TAG_SCORE_CAP, normalized)), 4)


def _merge_level_weights(values: List[Tuple[float, float]]) -> _LevelWeight:
    total_orders = sum(orders for _score, orders in values)
    if total_orders > 0:
        score = sum(score * orders for score, orders in values) / total_orders
    else:
        score = sum(score for score, _orders in values) / max(len(values), 1)
    return _LevelWeight(score=round(score, 4), orders=total_orders)


def _split_theater_names(value: Any) -> List[str]:
    text = "" if value is None else str(value).strip()
    if not text:
        return []
    output = [text]
    for separator in (",", "，", "|", "/"):
        if separator in text:
            output = [item.strip() for item in text.split(separator) if item.strip()]
            break
    seen = set()
    deduped: List[str] = []
    for item in output:
        normalized = config.normalize_theater_name(item)
        key = normalized or item
        if key in seen:
            continue
        seen.add(key)
        deduped.append(key)
    return deduped
